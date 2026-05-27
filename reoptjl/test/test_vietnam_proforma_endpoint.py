# REopt®, Copyright (c) Alliance for Sustainable Energy, LLC. See also https://github.com/NREL/REopt_API/blob/master/LICENSE.
import uuid
from io import BytesIO

from django.test import TransactionTestCase
from openpyxl import load_workbook
from tastypie.test import ResourceTestCaseMixin

from reoptjl.models import (
    APIMeta,
    ElectricLoadInputs,
    ElectricLoadOutputs,
    ElectricTariffInputs,
    ElectricTariffOutputs,
    ElectricUtilityOutputs,
    FinancialInputs,
    FinancialOutputs,
    PVOutputs,
    Settings,
    SiteInputs,
    SiteOutputs,
)


class TestVietnamProformaEndpoint(ResourceTestCaseMixin, TransactionTestCase):

    def test_results_endpoint_can_return_vietnam_proforma_xlsx(self):
        run_uuid = _create_completed_vietnam_result()

        resp = self.api_client.get(
            f"/v3/job/{run_uuid}/results",
            data={
                "vietnam_proforma": "true",
                "esco_energy_discount_fraction": "0.9",
            },
        )

        self.assertHttpOK(resp)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            f'attachment; filename="vietnam_proforma_{run_uuid}.xlsx"',
            resp["Content-Disposition"],
        )

        workbook = load_workbook(BytesIO(resp.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary",
                "System Sizing",
                "Results Comparison",
                "Annual Production",
                "Dispatch Profile",
                "Load Duration",
                "Cash Flow",
                "Tax Schedule",
                "Debt Service",
                "Developer Financials",
                "Assumptions",
            ],
        )
        self.assertEqual(workbook["Summary"]["A1"].value, "Total Capex (USD)")
        self.assertEqual(workbook["Summary"]["B1"].value, 100000)
        self.assertEqual(
            workbook["Assumptions"]["A2"].value,
            "esco_energy_discount_fraction",
        )
        self.assertEqual(workbook["Assumptions"]["B2"].value, 0.9)

    def test_results_endpoint_applies_vietnam_proforma_query_overrides(self):
        run_uuid = _create_completed_vietnam_result()

        resp = self.api_client.get(
            f"/v3/job/{run_uuid}/results",
            data={
                "vietnam_proforma": "true",
                "esco_energy_discount_fraction": "0.9",
                "owner_discount_rate_fraction": "0.12",
                "debt_fraction": "0.5",
                "debt_interest_rate_fraction": "0.09",
                "debt_term_years": "12",
                "annual_om_usd": "2000",
                "pv_capex_usd": "200000",
                "bess_capex_usd": "50000",
                "demand_savings_esco_share": "0.75",
                "grid_charging_enabled": "true",
            },
        )

        self.assertHttpOK(resp)
        workbook = load_workbook(BytesIO(resp.content), data_only=True)

        self.assertEqual(workbook["Summary"]["B1"].value, 250000)
        self.assertEqual(workbook["Summary"]["B2"].value, 125000)
        self.assertEqual(workbook["Summary"]["B3"].value, 125000)
        self.assertEqual(workbook["Cash Flow"]["F2"].value, 2000)
        self.assertEqual(workbook["Cash Flow"]["C2"].value, 3750)
        assumptions = {
            workbook["Assumptions"].cell(row=row, column=1).value:
            workbook["Assumptions"].cell(row=row, column=2).value
            for row in range(1, workbook["Assumptions"].max_row + 1)
        }
        self.assertEqual(assumptions["owner_discount_rate_fraction"], 0.12)
        self.assertEqual(assumptions["debt_fraction"], 0.5)
        self.assertEqual(assumptions["annual_om_usd"], 2000.0)
        self.assertEqual(assumptions["grid_charging_enabled"], True)


def _create_completed_vietnam_result():
    run_uuid = uuid.uuid4()
    meta = APIMeta.objects.create(
        run_uuid=run_uuid,
        api_version=3,
        status="optimal",
    )

    Settings.objects.create(meta=meta)
    SiteInputs.objects.create(meta=meta, latitude=10.0, longitude=106.0)
    FinancialInputs.objects.create(meta=meta, owner_discount_rate_fraction=0.11)
    ElectricLoadInputs.objects.create(meta=meta, loads_kw=[1, 2], year=2025)
    ElectricTariffInputs.objects.create(
        meta=meta,
        tou_energy_rates_per_kwh=[1000, 2000],
    )

    FinancialOutputs.objects.create(meta=meta, year_one_om_costs_before_tax=1000)
    ElectricTariffOutputs.objects.create(
        meta=meta,
        year_one_bill_before_tax_bau=50000,
        year_one_bill_before_tax=30000,
        year_one_demand_cost_before_tax_bau=8000,
        year_one_demand_cost_before_tax=3000,
    )
    ElectricUtilityOutputs.objects.create(meta=meta)
    ElectricLoadOutputs.objects.create(meta=meta)
    SiteOutputs.objects.create(meta=meta)
    PVOutputs.objects.create(
        meta=meta,
        size_kw=100,
        installed_cost_per_kw=1000,
        electric_to_load_series_kw=[1, 2],
    )

    return run_uuid
