from openpyxl import load_workbook

from proforma_vietnam.cash_flow import calculate_vietnam_esco_cash_flow


DPPA_YEAR_ONE_KEYS = (
    "c_dn_vnd",
    "c_dppa_vnd",
    "c_cl_vnd",
    "c_bl_vnd",
    "cfd_strike_revenue_vnd",
    "cfd_fmp_offset_vnd",
    "generator_fmp_revenue_vnd",
)


def compare_reference_workbook(workbook_or_path, tolerance_fraction=0.01):
    workbook = _load_workbook(workbook_or_path)
    inputs = _read_key_value_sheet(workbook["Inputs"])
    reference_summary = _read_key_value_sheet(workbook["Reference Summary"])
    reference_rows = _read_table_sheet(workbook["Reference Annual Cash Flow"])

    cash_flow_result = calculate_vietnam_esco_cash_flow(**_cash_flow_inputs(inputs))

    discrepancies = []
    _compare_mapping(
        discrepancies,
        "summary",
        cash_flow_result["summary"],
        reference_summary,
        tolerance_fraction,
    )

    actual_rows_by_year = {
        row["year"]: row
        for row in cash_flow_result["annual_cash_flows"]
    }
    for reference_row in reference_rows:
        year = reference_row["year"]
        actual_row = actual_rows_by_year.get(year)
        if actual_row is None:
            discrepancies.append(f"year {year}: missing actual row")
            continue
        _compare_mapping(
            discrepancies,
            f"year {year}",
            actual_row,
            reference_row,
            tolerance_fraction,
            skip_keys={"year"},
        )

    return discrepancies


def compare_dppa_reference_workbook(workbook_or_path, tolerance_fraction=0.01):
    """Compare a grid_dppa_cfd reference workbook against the implementation.

    Required sheets: ``Inputs`` (same shape as the Phase 2 reference), plus
    ``DPPA Year One`` (key/value sheet with the seven year-one settlement
    primitives in DPPA_YEAR_ONE_KEYS plus ``fee_escalation_rate`` and
    ``cfd_strike_escalation_rate``), ``Reference Summary``, and
    ``Reference Annual Cash Flow`` (extra columns ``c_dn_vnd``,
    ``cfd_net_vnd``, ``generator_revenue_vnd`` honored when present).
    """
    workbook = _load_workbook(workbook_or_path)
    inputs = _read_key_value_sheet(workbook["Inputs"])
    dppa_year_one = _read_key_value_sheet(workbook["DPPA Year One"])
    reference_summary = _read_key_value_sheet(workbook["Reference Summary"])
    reference_rows = _read_table_sheet(workbook["Reference Annual Cash Flow"])

    cash_flow_kwargs = _cash_flow_inputs(inputs)
    cash_flow_kwargs["dppa_settlement"] = _dppa_settlement_from_workbook(dppa_year_one)
    cash_flow_result = calculate_vietnam_esco_cash_flow(**cash_flow_kwargs)

    discrepancies = []
    _compare_mapping(
        discrepancies,
        "summary",
        cash_flow_result["summary"],
        reference_summary,
        tolerance_fraction,
    )

    actual_rows_by_year = {
        row["year"]: row
        for row in cash_flow_result["annual_cash_flows"]
    }
    for reference_row in reference_rows:
        year = reference_row["year"]
        actual_row = actual_rows_by_year.get(year)
        if actual_row is None:
            discrepancies.append(f"year {year}: missing actual row")
            continue
        _compare_mapping(
            discrepancies,
            f"year {year}",
            actual_row,
            reference_row,
            tolerance_fraction,
            skip_keys={"year"},
        )

    return discrepancies


def _dppa_settlement_from_workbook(dppa_year_one):
    year_one = {key: float(dppa_year_one[key]) for key in DPPA_YEAR_ONE_KEYS}
    return {
        "type": "grid_dppa_cfd",
        "esco_energy_revenue_vnd": 0.0,
        "year_one": year_one,
        "escalation": {
            "fee_escalation_rate": float(dppa_year_one.get("fee_escalation_rate", 0.0)),
            "cfd_strike_escalation_rate": float(
                dppa_year_one.get("cfd_strike_escalation_rate", 0.0)
            ),
        },
        "hourly_breakout": [],
        "monthly_breakout": [],
    }


def _load_workbook(workbook_or_path):
    if isinstance(workbook_or_path, str):
        return load_workbook(workbook_or_path, data_only=True)
    return workbook_or_path


def _read_key_value_sheet(worksheet):
    values = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key, value = row[0], row[1]
        if key is not None:
            values[key] = value
    return values


def _read_table_sheet(worksheet):
    headers = [
        cell.value
        for cell in worksheet[1]
        if cell.value is not None
    ]
    rows = []

    for worksheet_row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in worksheet_row):
            continue
        rows.append({
            header: worksheet_row[index]
            for index, header in enumerate(headers)
        })

    return rows


def _cash_flow_inputs(inputs):
    converted = dict(inputs)
    converted["project_served_pv_kwh"] = _parse_number_list(converted["project_served_pv_kwh"])
    converted["evn_energy_rates_vnd_per_kwh"] = _parse_number_list(
        converted["evn_energy_rates_vnd_per_kwh"]
    )
    return converted


def _parse_number_list(value):
    if isinstance(value, list):
        return value
    return [
        float(part.strip())
        for part in str(value).split(",")
        if part.strip()
    ]


def _compare_mapping(
    discrepancies,
    label,
    actual_values,
    reference_values,
    tolerance_fraction,
    skip_keys=None,
):
    skip_keys = skip_keys or set()

    for key, reference_value in reference_values.items():
        if key in skip_keys:
            continue
        actual_value = actual_values.get(key)
        if not _within_tolerance(actual_value, reference_value, tolerance_fraction):
            discrepancies.append(
                (
                    f"{label} {key}: actual={actual_value}, "
                    f"reference={reference_value}, tolerance={tolerance_fraction:.2%}"
                )
            )


def _within_tolerance(actual_value, reference_value, tolerance_fraction):
    if reference_value in (None, ""):
        return actual_value in (None, "")
    if actual_value is None:
        return False
    if reference_value == 0:
        return abs(actual_value) <= tolerance_fraction

    return abs(actual_value - reference_value) / abs(reference_value) <= tolerance_fraction
