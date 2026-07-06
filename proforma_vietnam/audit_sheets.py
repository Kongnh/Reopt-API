"""Audit-grade workbook sheets: live Excel formulas + engine tie-out.

These sheets make the Vietnam proforma workbook self-auditable for a
third-party (investor, lender, or model auditor):

- ``Cover``            — contents, conventions, colour legend, overall status.
- ``Assumptions``      — every input grouped, with unit + source, as *named
                         cells* so the formula sheets read like the design doc.
- ``Model Basis``      — methodology, settlement math, simplifications register.
- ``Pro Forma (Audit)``— the full multi-year cash flow rebuilt with live Excel
                         formulas (SAM cash-flow convention: line items down,
                         years across). Only engine outputs that cannot be
                         derived in-sheet (year-1 dispatch/settlement bases,
                         the replacement schedule, tie-out rows) are hardcoded
                         and shaded as inputs. A Checks block compares every
                         Excel-computed metric against the Python engine and
                         shows PASS/REVIEW.
- ``FX Sensitivity``   — USD-reported equity returns under annual VND
                         depreciation, live formulas over the audit sheet.

The Excel formulas replicate ``cash_flow.py`` (and ``tax_model.py``) exactly —
including the CIT holiday clock and the FIFO 5-year tax-loss carryforward —
so an auditor can trace every published figure from named inputs to the
metric without leaving Excel.
"""

from datetime import date

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from proforma_vietnam.cash_flow import calculate_fx_sensitivity
from proforma_vietnam.dppa_settlement import (
    DPPA_TYPE_GRID_CFD,
    DPPA_TYPE_PHYSICAL_PRIVATE_WIRE,
)
from proforma_vietnam.structures import DIRECT_OWNERSHIP, DPPA, PHYSICAL_DPPA

NAVY = "1F3864"
ACCENT = "2E74B5"
INPUT_COLOR = "FFF2CC"   # hardcoded engine outputs / user inputs
CHECK_COLOR = "E2EFDA"   # tie-out block

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
SECTION_FILL = PatternFill("solid", fgColor=ACCENT)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor=INPUT_COLOR)
CHECK_FILL = PatternFill("solid", fgColor=CHECK_COLOR)
NOTE_FONT = Font(italic=True, size=9, color="595959")
BOLD_FONT = Font(bold=True)

FMT_AMOUNT = "#,##0"
FMT_AMOUNT_2 = "#,##0.00"
FMT_PERCENT = "0.00%"
FMT_RATIO = "0.00"
FMT_FACTOR = "0.0000"
FMT_YEARS = "0.0"

PRO_FORMA_SHEET = "Pro Forma (Audit)"

# assumptions.json keys already rendered in a curated Assumptions section; the
# remainder is echoed raw at the bottom of the sheet so the file is complete.
CURATED_ASSUMPTION_KEYS = {
    "case_name", "run_uuid", "country", "tariff_year", "voltage_level",
    "tou_schedule", "exchange_rate_vnd_per_usd", "evn_energy_escalation_rate",
    "evn_capacity_escalation_rate", "esco_energy_discount_fraction",
    "demand_savings_esco_share", "grid_charging_enabled",
    "battery_replacement_year", "annual_om_usd", "pv_capex_usd",
    "bess_capex_usd", "om_escalation_rate", "pv_degradation_rate",
    "pv_depreciation_years", "debt_fraction", "debt_interest_rate_fraction",
    "debt_term_years", "construction_months", "principal_grace_years",
    "target_min_dscr", "contract_years", "contract_residual_value_usd",
    "vat_rate_fraction", "vat_refund_year",
    "owner_discount_rate_fraction", "analysis_years",
    "case_config", "dppa",
}

STORAGE_CASE_ROWS = [
    ("Storage minimum power", "min_kw", "kW"),
    ("Storage maximum power", "max_kw", "kW"),
    ("Storage minimum energy", "min_kwh", "kWh"),
    ("Storage maximum energy", "max_kwh", "kWh"),
    ("Storage cost per kW", "installed_cost_per_kw", "USD/kW"),
    ("Storage cost per kWh", "installed_cost_per_kwh", "USD/kWh"),
    ("Storage fixed cost", "installed_cost_constant", "USD"),
    ("Replacement cost per kW", "replace_cost_per_kw", "USD/kW"),
    ("Replacement cost per kWh", "replace_cost_per_kwh", "USD/kWh"),
    ("Replacement fixed cost", "replace_cost_constant", "USD"),
    ("Inverter replacement year", "inverter_replacement_year", "year"),
    ("Battery replacement year", "battery_replacement_year", "year"),
    ("Fixed-cost replacement year", "cost_constant_replacement_year", "year"),
    ("O&M (fraction of installed cost)", "om_cost_fraction_of_installed_cost", "per year"),
    ("Can charge from grid", "can_grid_charge", "yes/no"),
]

# Absolute tolerances for the tie-out checks. Excel and the engine execute the
# same double-precision arithmetic, so real deltas are ~1e-9; the tolerances
# only absorb IRR root-finding differences and display rounding.
TOL_AMOUNT = 1.0        # USD
TOL_RATE = 0.0005       # 5 bp on IRR / percent metrics
TOL_RATIO = 0.005       # DSCR
TOL_YEARS = 0.05        # payback


def _fmt_num(value):
    """Render a float for embedding inside an Excel formula string."""
    return repr(float(value))


def _define_name(workbook, name, sheet_title, cell):
    workbook.defined_names[name] = DefinedName(
        name, attr_text=f"'{sheet_title}'!${cell[0]}${cell[1]}"
    )


# ---------------------------------------------------------------------------
# Assumptions
# ---------------------------------------------------------------------------

def write_assumptions_sheet(worksheet, workbook, assumptions, derivation):
    """Grouped assumptions with units, sources and workbook-scope named cells.

    Returns True when the engine derivation block was available (i.e. the
    formula sheets can be built on top of these names).
    """
    derivation = derivation or {}
    dppa = assumptions.get("dppa") if assumptions else None
    # is_dppa is strictly the grid-CfD settlement (drives the CfD assumption
    # rows / labels). The private wire is its own branch.
    is_dppa = derivation.get("structure") == DPPA or bool(
        dppa and dppa.get("type", "none") == DPPA_TYPE_GRID_CFD
    )
    is_physical = derivation.get("structure") == PHYSICAL_DPPA or bool(
        dppa and dppa.get("type", "none") == DPPA_TYPE_PHYSICAL_PRIVATE_WIRE
    )
    is_direct = derivation.get("structure") == DIRECT_OWNERSHIP or bool(
        assumptions and assumptions.get("direct_ownership")
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("B1:E1")
    title = worksheet.cell(row=1, column=2, value="Assumptions & Model Inputs")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    worksheet.row_dimensions[1].height = 24
    worksheet.cell(
        row=2, column=2,
        value="Shaded cells are hardcoded inputs (case file or engine output); "
              "named cells drive every formula on the Pro Forma (Audit) sheet.",
    ).font = NOTE_FONT

    state = {"row": 4}

    def section(label):
        row = state["row"]
        for column in range(2, 6):
            worksheet.cell(row=row, column=column).fill = SECTION_FILL
        cell = worksheet.cell(row=row, column=2, value=label)
        cell.font = SECTION_FONT
        state["row"] = row + 1

    def entry(label, value, unit="", source="", name=None, fmt=None, formula=None):
        row = state["row"]
        worksheet.cell(row=row, column=2, value=label)
        value_cell = worksheet.cell(row=row, column=3)
        if formula is not None:
            value_cell.value = formula
        else:
            value_cell.value = value
            value_cell.fill = INPUT_FILL
        if fmt:
            value_cell.number_format = fmt
        # Never write empty strings: openpyxl serialises them as an inlineStr
        # cell with no <is> child, which Excel rejects as a corrupt workbook.
        if unit:
            worksheet.cell(row=row, column=4, value=unit).font = NOTE_FONT
        if source:
            worksheet.cell(row=row, column=5, value=source).font = NOTE_FONT
        if name:
            _define_name(workbook, name, worksheet.title, ("C", row))
        state["row"] = row + 1

    d = derivation

    def get(key, alt=None):
        """Engine derivation is authoritative; fall back to the case assumptions."""
        value = d.get(key)
        if value is not None:
            return value
        return assumptions.get(alt or key)

    if is_dppa:
        structure_label = "Grid DPPA with CfD (ND57/2025)"
    elif is_physical:
        structure_label = "Physical (private-wire) DPPA (ND57 Điều 25; Decree 243/2026)"
    elif is_direct:
        structure_label = "Direct ownership — factory self-invest (avoided EVN bill)"
    else:
        structure_label = "ESCO discount-to-EVN (behind-the-meter)"

    section("Project & Run")
    entry("Case name", assumptions.get("case_name", "Vietnam case"), source="case.json")
    entry("Financing structure", structure_label,
          source="proforma_vietnam.structures")
    if assumptions.get("run_uuid"):
        entry("REopt run UUID", assumptions["run_uuid"], source="REopt API")
    entry("Report prepared", date.today().isoformat())
    entry("Analysis period", d.get("project_years", assumptions.get("analysis_years", 25)),
          unit="years", source="case.json financial.analysis_years",
          name="PROJECT_YEARS", fmt="0")
    if assumptions.get("tariff_year"):
        entry("Tariff year", assumptions.get("tariff_year"), source="case.json tariff.year", fmt="0")
    if assumptions.get("voltage_level"):
        entry("Voltage level", str(assumptions.get("voltage_level")), source="case.json tariff.voltage_level")
    if assumptions.get("tou_schedule"):
        entry("TOU schedule", assumptions.get("tou_schedule"), source="case.json tariff.tou_schedule")

    case_config = assumptions.get("case_config") or {}
    _write_case_definition_sections(section, entry, case_config)

    section("Currency & FX")
    entry("Model currency", "USD",
          source="EVN tariff converted VND→USD before REopt; see Model Basis")
    entry("Contract exchange rate",
          d.get("exchange_rate_vnd_per_usd") or assumptions.get("exchange_rate_vnd_per_usd"),
          unit="VND per USD", source="case.json tariff.exchange_rate_vnd_per_usd",
          name="FX_VND_PER_USD", fmt=FMT_AMOUNT)
    entry("FX treatment", "Held flat over the analysis period",
          source="Simplification — see FX Sensitivity sheet")

    section("Capital Costs")
    entry("PV capex", get("pv_capex_usd"), unit="USD",
          source="REopt PV size_kw × installed_cost_per_kw",
          name="PV_CAPEX", fmt=FMT_AMOUNT)
    entry("BESS capex", get("bess_capex_usd"), unit="USD",
          source="REopt ElectricStorage initial_capital_cost",
          name="BESS_CAPEX", fmt=FMT_AMOUNT)
    entry("Other capex", d.get("other_capex_usd", 0), unit="USD",
          source="case.json (developer overrides)",
          name="OTHER_CAPEX", fmt=FMT_AMOUNT)
    entry("Total investment", None, unit="USD", source="Formula: PV + BESS + Other",
          name="TOTAL_CAPEX", fmt=FMT_AMOUNT,
          formula="=PV_CAPEX+BESS_CAPEX+OTHER_CAPEX")

    section("Operating Costs")
    entry("Year-1 O&M", get("annual_om_year1_usd", "annual_om_usd"), unit="USD/yr",
          source="REopt Financial.year_one_om_costs_before_tax",
          name="OM_YEAR1", fmt=FMT_AMOUNT)
    entry("O&M escalation", get("om_escalation_rate") or 0.0, unit="per year",
          source="vietnam_defaults.json / case.json financial.om_escalation_rate",
          name="ESC_OM", fmt=FMT_PERCENT)
    if assumptions.get("battery_replacement_year"):
        entry("Battery replacement year", assumptions["battery_replacement_year"],
              unit="year", source="case.json technologies.storage.battery_replacement_year",
              fmt="0")

    section("Tariff & Escalation")
    entry("EVN energy escalation", get("evn_energy_escalation_rate"), unit="per year",
          source="case.json tariff.evn_energy_escalation_rate",
          name="ESC_ENERGY", fmt=FMT_PERCENT)
    entry("EVN capacity (demand) escalation", get("evn_capacity_escalation_rate"),
          unit="per year", source="case.json tariff.evn_capacity_escalation_rate",
          name="ESC_CAPACITY", fmt=FMT_PERCENT)
    entry("PV degradation", get("pv_degradation_rate") or 0.0, unit="per year",
          source="REopt PV.degradation_fraction",
          name="PV_DEGRADATION", fmt=FMT_PERCENT)
    tariff_config = case_config.get("tariff") or {}
    if tariff_config.get("currency"):
        entry("Tariff currency fed to REopt", str(tariff_config["currency"]).upper(),
              source="case.json tariff.currency")
    if tariff_config.get("two_component_pilot_enabled") is not None:
        entry("Two-component tariff pilot", tariff_config["two_component_pilot_enabled"],
              unit="yes/no", source="case.json tariff.two_component_pilot_enabled")

    section("Contract Terms")
    entry("ESCO energy price (fraction of EVN tariff)",
          get("esco_energy_discount_fraction"), unit="fraction",
          source="case.json esco_contract.esco_energy_discount_fraction",
          name="ESCO_DISCOUNT", fmt=FMT_PERCENT)
    entry("Demand savings share to ESCO",
          get("esco_demand_savings_share", "demand_savings_esco_share"),
          unit="fraction", source="case.json esco_contract.demand_savings_esco_share",
          name="DEMAND_SHARE", fmt=FMT_PERCENT)
    if assumptions.get("grid_charging_enabled") is not None:
        entry("Grid charging enabled", assumptions["grid_charging_enabled"],
              unit="yes/no", source="case.json esco_contract.grid_charging_enabled")
    if is_dppa and dppa:
        volume = dppa.get("cfd_contract_volume_kwh_per_hour")
        annual_volume = sum(volume) if isinstance(volume, list) else (volume or 0.0) * 8760
        entry("CfD strike price", dppa.get("cfd_strike_per_kwh_vnd"), unit="VND/kWh",
              source="case.json dppa.cfd_strike_per_kwh_vnd", fmt=FMT_AMOUNT)
        entry("CfD strike escalation",
              (d.get("dppa_escalation") or {}).get("cfd_strike_escalation_rate", 0.0),
              unit="per year", source="case.json dppa.cfd_strike_escalation_rate",
              name="ESC_STRIKE", fmt=FMT_PERCENT)
        entry("DPPA fee / FMP escalation",
              (d.get("dppa_escalation") or {}).get("fee_escalation_rate", 0.0),
              unit="per year", source="case.json dppa.fee_escalation_rate",
              name="ESC_FEE", fmt=FMT_PERCENT)
        entry("Annual CfD contract volume", annual_volume, unit="kWh/yr",
              source="case.json dppa.cfd_contract_volume_kwh_per_hour", fmt=FMT_AMOUNT)
        entry("Transmission loss factor k (price-only)",
              dppa.get("transmission_loss_factor_k"), unit="ratio",
              source="ND57; CFMP = FMP × k", fmt=FMT_FACTOR)
        entry("Distribution loss factor K_pp", dppa.get("distribution_loss_factor_kpp"),
              unit="ratio", source="ND57 by voltage; Q_adj = Q_meter / K_pp",
              fmt="0.000000")
        entry("DPPA system service fee", dppa.get("c_dppa_service_fee_vnd_per_kwh"),
              unit="VND/kWh", source="2025 EVN published fee", fmt=FMT_AMOUNT_2)
        entry("Settlement adder C_CL", dppa.get("c_cl_settlement_adder_vnd_per_kwh"),
              unit="VND/kWh", source="2025 EVN published adder", fmt=FMT_AMOUNT_2)
        entry("Allocation fraction δ", dppa.get("allocation_fraction_delta"),
              unit="fraction", source="case.json dppa.allocation_fraction_delta",
              fmt=FMT_RATIO)
        if dppa.get("fmp_series_path"):
            entry("FMP / CFMP price series", dppa.get("fmp_series_path"),
                  source="case.json dppa.fmp_series_path (8760-hour VND/kWh)")

    physical = d.get("physical_dppa")
    if physical:
        # ND57 Điều 25 private wire. Gated on the engine derivation so ESCO /
        # grid-CfD cases carry no PPA names or rows.
        section("Physical DPPA (Private Wire — ND57 Điều 25)")
        entry("PPA price (matched energy)", physical.get("ppa_price_usd_per_kwh"),
              unit="USD/kWh",
              source="case.json dppa.ppa_price_vnd_per_kwh (freely negotiated; "
                     "Decree 243/2026 removed the ceiling) at contract FX",
              name="PPA_PRICE", fmt="#,##0.00000")
        entry("PPA price escalation", physical.get("ppa_price_escalation_rate"),
              unit="per year",
              source="case.json dppa.ppa_price_escalation_rate (default: flat PPA)",
              name="PPA_ESC", fmt=FMT_PERCENT)

    surplus = d.get("surplus_export")
    if surplus:
        # Decree 243/2026 rooftop surplus-export (ESCO + physical private wire).
        # Gated on the engine derivation so disabled cases carry no surplus names
        # or rows.
        section("Surplus Export (Decree 243/2026)")
        entry("Year-1 surplus sold to EVN", surplus.get("sold_kwh_year1"),
              unit="kWh/yr",
              source="min(PV grid export + curtailed, cap × PV output)",
              name="SURPLUS_KWH_Y1", fmt=FMT_AMOUNT)
        entry("Surplus export price", surplus.get("price_usd_per_kwh"),
              unit="USD/kWh",
              source="Decree 243 market price, capped at Decision 988 regional "
                     "ceiling, at contract FX",
              name="SURPLUS_PRICE", fmt="#,##0.00000")
        entry("Surplus price escalation", surplus.get("price_escalation_rate"),
              unit="per year",
              source="case.json surplus_export.price_escalation_rate "
                     "(default: EVN energy escalation)",
              name="SURPLUS_ESC", fmt=FMT_PERCENT)
        entry("Surplus export cap", surplus.get("cap_fraction"),
              unit="of PV output",
              source="Decree 243/2026 (50%; >50% negotiable to 2030)",
              name="SURPLUS_CAP", fmt=FMT_PERCENT)

    section("Financing")
    entry("Debt fraction", get("debt_fraction"), unit="of total capex",
          source="vietnam_defaults.json / case.json financial.debt_fraction",
          name="DEBT_FRACTION", fmt=FMT_PERCENT)
    entry("Debt interest rate", get("debt_interest_rate_fraction"), unit="per year",
          source="vietnam_defaults.json / case.json financial.debt_interest_rate_fraction",
          name="DEBT_RATE", fmt=FMT_PERCENT)
    if d.get("debt_currency") == "USD":
        # USD-denominated debt (default is VND). Gated on the engine derivation
        # so VND cases carry no currency name and stay byte-identical. The
        # resolved rate is already surfaced by the DEBT_RATE cell above; this
        # only labels the currency the FX Sensitivity sheet treats as FX-fixed.
        entry("Debt currency", d.get("debt_currency"), unit="",
              source="case.json financial.debt_currency (default VND; USD debt "
                     "service is FX-fixed — see FX Sensitivity)",
              name="DEBT_CURRENCY")
    entry("Debt term", get("debt_term_years"), unit="years",
          source="vietnam_defaults.json / case.json financial.debt_term_years",
          name="DEBT_TERM_YEARS", fmt="0")
    debt_sizing = d.get("debt_sizing")
    if debt_sizing:
        # DSCR-driven debt sizing: the loan is min(fraction-based, DSCR-
        # supported). Gated on the engine derivation so fraction-sized cases
        # carry no sizing names or rows and stay byte-identical. The DSCR-
        # supported principal is the engine's converged fixed point (a
        # non-closed-form iteration, so it is carried as an input value like the
        # other engine constants); DEBT_PRINCIPAL then reproduces the min() rule
        # live, and every downstream formula (IDC, COD balance, debt schedule,
        # equity) rides the sized principal.
        entry("Target minimum DSCR (covenant)", debt_sizing.get("target_min_dscr"),
              unit="x",
              source="case.json financial.target_min_dscr (DSCR-sized debt; "
                     "level-payment sizing on base-case CFADS)",
              name="TARGET_MIN_DSCR", fmt=FMT_RATIO)
        entry("Fraction-based debt", None, unit="USD",
              source="Formula: total investment × debt fraction (pre-DSCR cap)",
              name="FRACTION_DEBT", fmt=FMT_AMOUNT,
              formula="=TOTAL_CAPEX*DEBT_FRACTION")
        entry("DSCR-supported debt", debt_sizing.get("supported_principal_usd"),
              unit="USD",
              source="Engine fixed point: D × min_DSCR(D) / target (converged so "
                     "min DSCR = covenant when this binds)",
              name="SUPPORTED_DEBT", fmt=FMT_AMOUNT)
        entry("Debt principal (DSCR-sized)", None, unit="USD",
              source="Formula: min(fraction-based, DSCR-supported)",
              name="DEBT_PRINCIPAL", fmt=FMT_AMOUNT,
              formula="=MIN(FRACTION_DEBT,SUPPORTED_DEBT)")
    else:
        entry("Debt principal", None, unit="USD", source="Formula: total investment × debt fraction",
              name="DEBT_PRINCIPAL", fmt=FMT_AMOUNT,
              formula="=TOTAL_CAPEX*DEBT_FRACTION")
    entry("Equity investment", None, unit="USD", source="Formula: total investment − debt",
          name="EQUITY_INVESTMENT", fmt=FMT_AMOUNT,
          formula="=TOTAL_CAPEX-DEBT_PRINCIPAL")
    construction = d.get("construction")
    if construction:
        # Construction period + capitalized IDC + principal grace. Gated on the
        # engine derivation so overnight-build cases carry no construction
        # names or rows and stay byte-identical.
        entry("Construction period", construction.get("construction_months"),
              unit="months", source="case.json financial.construction_months",
              name="CONSTRUCTION_MONTHS", fmt="0")
        entry("Interest during construction (IDC)", None, unit="USD",
              source="Formula: debt principal × rate × months/12 ÷ 2 (even "
                     "drawdown, simple interest); capitalized per Circular 45",
              name="IDC", fmt=FMT_AMOUNT,
              formula="=DEBT_PRINCIPAL*DEBT_RATE*CONSTRUCTION_MONTHS/12/2")
        entry("Debt balance at COD", None, unit="USD",
              source="Formula: debt principal + IDC (IDC debt-funded, rolled up)",
              name="COD_DEBT_BALANCE", fmt=FMT_AMOUNT,
              formula="=DEBT_PRINCIPAL+IDC")
        entry("Principal grace period", construction.get("principal_grace_years"),
              unit="years",
              source="case.json financial.principal_grace_years (interest-only; "
                     "principal amortizes over term − grace)",
              name="PRINCIPAL_GRACE_YEARS", fmt="0")
        entry("Annual debt payment (level)", None, unit="USD/yr",
              source="Annuity on the COD balance over (term − grace) years",
              name="DEBT_PAYMENT", fmt=FMT_AMOUNT,
              formula=(
                  "=IF(OR(COD_DEBT_BALANCE<=0,"
                  "DEBT_TERM_YEARS-PRINCIPAL_GRACE_YEARS<=0),0,"
                  "IF(DEBT_RATE=0,"
                  "COD_DEBT_BALANCE/(DEBT_TERM_YEARS-PRINCIPAL_GRACE_YEARS),"
                  "COD_DEBT_BALANCE*(DEBT_RATE*(1+DEBT_RATE)"
                  "^(DEBT_TERM_YEARS-PRINCIPAL_GRACE_YEARS))"
                  "/((1+DEBT_RATE)^(DEBT_TERM_YEARS-PRINCIPAL_GRACE_YEARS)-1)))"
              ))
    else:
        entry("Annual debt payment (level)", None, unit="USD/yr",
              source="Standard annuity over the debt term",
              name="DEBT_PAYMENT", fmt=FMT_AMOUNT,
              formula=(
                  "=IF(OR(DEBT_PRINCIPAL<=0,DEBT_TERM_YEARS<=0),0,"
                  "IF(DEBT_RATE=0,DEBT_PRINCIPAL/DEBT_TERM_YEARS,"
                  "DEBT_PRINCIPAL*(DEBT_RATE*(1+DEBT_RATE)^DEBT_TERM_YEARS)"
                  "/((1+DEBT_RATE)^DEBT_TERM_YEARS-1)))"
              ))
    entry("Owner discount rate", get("owner_discount_rate_fraction"), unit="per year",
          source="case.json financial.owner_discount_rate_fraction",
          name="DISC_RATE", fmt=FMT_PERCENT)

    section("Tax & Depreciation (Vietnam)")
    cit = d.get("cit", {})
    regime_label = {
        "re_producer": "RE producer — Law 67/2025 preferential (10% / 15y)",
        "standard_with_holiday": "Standard + holiday (conservative ESCO default)",
        "standard_flat": "Standard flat 20% (self-invest factory; no new-project incentive)",
    }.get(cit.get("regime"), "Standard + holiday (conservative ESCO default)")
    entry("CIT regime", regime_label,
          source="proforma_vietnam.cash_flow (structure-dependent; explicit override wins)")
    entry("CIT standard rate", cit.get("standard_rate"), unit="of taxable income",
          source="Law 67/2025/QH15; vietnam_defaults.json",
          name="CIT_RATE", fmt=FMT_PERCENT)
    if cit.get("preferential_rate") is not None:
        entry("CIT preferential rate (RE producer)", cit.get("preferential_rate"),
              unit="of taxable income",
              source="Law 67/2025/QH15 + Decree 320/2025/NĐ-CP (first-15-year window)",
              name="CIT_PREF_RATE", fmt=FMT_PERCENT)
        entry("CIT preferential-rate period", cit.get("preferential_years"),
              unit="years (from year 1)", source="Law 67/2025/QH15",
              name="CIT_PREF_YEARS", fmt="0")
    entry("CIT holiday", cit.get("holiday_years"), unit="years",
          source="Law 67/2025 (from first profitable year); Circular 78/2014 Art. 18 shape",
          name="CIT_HOLIDAY_YEARS", fmt="0")
    entry("CIT reduced-rate period", cit.get("reduced_rate_years"), unit="years",
          source="Law 67/2025; Circular 78/2014 Art. 18 shape",
          name="CIT_REDUCED_YEARS", fmt="0")
    entry("CIT reduction during period", cit.get("reduced_rate_fraction"),
          unit="of applicable base rate", source="Law 67/2025 (50% of base rate)",
          name="CIT_REDUCED_FRACTION", fmt=FMT_PERCENT)
    entry("Tax-loss carryforward limit", cit.get("loss_carryforward_years"),
          unit="years", source="Law 67/2025; Circular 78/2014 Art. 9 shape",
          name="CIT_LOSS_CF_YEARS", fmt="0")
    entry("Incentive clock cap", (cit.get("incentive_start_cap_index") or 3) + 1,
          unit="year (latest start)", source="Law 67/2025; Circular 78/2014 Art. 18 shape",
          name="CIT_CLOCK_CAP_YEAR", fmt="0")
    entry("PV depreciation (straight-line)", d.get("pv_depreciation_years"),
          unit="years", source="Circular 45/2013/TT-BTC (7–20 yr band)",
          name="PV_DEP_YEARS", fmt="0")
    entry("BESS depreciation (straight-line)", d.get("bess_depreciation_years"),
          unit="years", source="Circular 45/2013/TT-BTC",
          name="BESS_DEP_YEARS", fmt="0")

    contract_term = d.get("contract_term")
    if contract_term:
        # ESCO contract tenor + end-of-term asset transfer (Task 4e). Gated on
        # the engine derivation so full-horizon (default) cases carry no tenor
        # names or rows and stay byte-identical.
        section("Contract Tenor & Asset Transfer (Task 4e)")
        entry("Contract tenor (asset-transfer year T)",
              contract_term.get("contract_years"), unit="years",
              source="case.json financial.contract_years (operations cease at "
                     "end of year T; asset transfers to the host)",
              name="CONTRACT_YEARS", fmt="0")
        entry("Contractual residual / buyout value",
              contract_term.get("residual_value_usd"), unit="USD",
              source="case.json financial.contract_residual_value_usd (host's "
                     "end-of-term transfer payment, at contract FX)",
              name="RESIDUAL_VALUE", fmt=FMT_AMOUNT)

    vat = d.get("vat")
    if vat:
        # Input VAT on capex (Task 4f). Gated on the engine derivation so default
        # (no-VAT) cases carry no VAT names or rows and stay byte-identical.
        section("Input VAT on Capex (Task 4f)")
        entry("Input VAT rate (user input)", vat.get("rate"), unit="of capex base",
              source="case.json financial.vat_rate_fraction (VAT Law 48/2024/QH15, "
                     "eff. 2025-07-01; standard rate 10%, user-supplied)",
              name="VAT_RATE", fmt=FMT_PERCENT)
        entry("VAT refund year", vat.get("refund_year"), unit="year (0 = COD year)",
              source="case.json financial.vat_refund_year (investment-project "
                     "refund timing; equity-funded, out of debt / CFADS / DSCR)",
              name="VAT_REFUND_YEAR", fmt="0")

    section("Year-1 Engine Outputs (hardcoded — dispatch × tariff, not derivable in-sheet)")
    entry("Year-1 BAU EVN bill", d.get("bau_evn_bill_year1_usd"), unit="USD",
          source="REopt ElectricTariff.year_one_bill_before_tax_bau",
          name="BAU_BILL_Y1", fmt=FMT_AMOUNT)
    entry("Year-1 optimized EVN bill", d.get("optimized_evn_bill_year1_usd"), unit="USD",
          source="REopt ElectricTariff.year_one_bill_before_tax",
          name="OPT_BILL_Y1", fmt=FMT_AMOUNT)
    entry("Year-1 BAU demand charge", d.get("bau_demand_charge_year1_usd"), unit="USD",
          source="REopt year_one_demand_cost_before_tax_bau",
          name="BAU_DEMAND_Y1", fmt=FMT_AMOUNT)
    entry("Year-1 optimized demand charge", d.get("optimized_demand_charge_year1_usd"),
          unit="USD", source="REopt year_one_demand_cost_before_tax",
          name="OPT_DEMAND_Y1", fmt=FMT_AMOUNT)
    entry("Year-1 demand savings base", d.get("base_demand_savings_usd"), unit="USD",
          source="max(BAU − optimized demand charge, 0)",
          name="BASE_DEMAND_SAVINGS", fmt=FMT_AMOUNT)
    entry("Year-1 served-energy retail value", d.get("base_served_retail_value_usd"),
          unit="USD", source="Σ project-served kWh × EVN TOU rate (8760 h)",
          name="BASE_SERVED_RETAIL", fmt=FMT_AMOUNT)
    if is_dppa:
        dp = d.get("dppa_year_one_usd", {})
        entry("Year-1 C_DN (spot energy)", dp.get("c_dn"), unit="USD",
              source="settle_dppa_year_one: Σ Q_Khc × CFMP × K_pp",
              name="DPPA_C_DN_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 C_DPPA (system fee)", dp.get("c_dppa"), unit="USD",
              source="Σ Q_Khc × f_dppa", name="DPPA_C_DPPA_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 C_CL (settlement adder)", dp.get("c_cl"), unit="USD",
              source="Σ Q_Khc × f_cl", name="DPPA_C_CL_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 C_BL (retail shortfall)", dp.get("c_bl"), unit="USD",
              source="Σ shortfall × EVN TOU rate", name="DPPA_C_BL_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 CfD strike leg", dp.get("cfd_strike_revenue"), unit="USD",
              source="Σ P_c × min(Q_c, Q_Khc)", name="DPPA_CFD_STRIKE_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 CfD market offset", dp.get("cfd_fmp_offset"), unit="USD",
              source="Σ FMP × min(Q_c, Q_Khc)", name="DPPA_CFD_OFFSET_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 generator FMP revenue", dp.get("generator_fmp_revenue"), unit="USD",
              source="Σ Q_re_meter × FMP", name="DPPA_FMP_REV_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 matched-energy retail value", dp.get("matched_retail_value"),
              unit="USD", source="Σ Q_Khc × EVN TOU rate (degradation repurchase)",
              name="DPPA_MATCHED_RETAIL_Y1", fmt=FMT_AMOUNT)
    elif is_physical:
        entry("Year-1 matched energy (PPA volume)",
              (physical or {}).get("matched_kwh_year1"), unit="kWh/yr",
              source="Σ project-served kWh (PV→load + battery→load)",
              name="PPA_MATCHED_KWH_Y1", fmt=FMT_AMOUNT)
        entry("Year-1 grid arbitrage base", d.get("base_grid_arbitrage_revenue_usd", 0),
              unit="USD", source="REopt net grid arbitrage × ESCO share "
                                 "(0 under a private wire)",
              name="BASE_GRID_ARB", fmt=FMT_AMOUNT)
    elif is_direct:
        # Factory self-invest: the benefit line is the full avoided EVN bill
        # (BAU − optimized), rebuilt on the Pro Forma sheet from BAU_BILL_Y1 /
        # OPT_BILL_Y1 / BASE_SERVED_RETAIL above; no ESCO discount base applies.
        profitable_host = bool((d.get("direct_ownership") or {}).get("assume_profitable_host"))
        entry("Profitable-host tax convention",
              "Yes — project losses shield the host's other profits (negative CIT)"
              if profitable_host else
              "No — standalone 5-year FIFO loss carryforward",
              source="case.json direct_ownership.assume_profitable_host "
                     "(default: yes for direct ownership)")
    else:
        entry("Year-1 ESCO energy revenue base", d.get("base_energy_revenue_usd"),
              unit="USD", source="Σ served kWh × EVN TOU rate × ESCO discount",
              name="BASE_ENERGY_REV", fmt=FMT_AMOUNT)
        entry("Year-1 grid arbitrage base", d.get("base_grid_arbitrage_revenue_usd", 0),
              unit="USD", source="REopt net grid arbitrage × ESCO share",
              name="BASE_GRID_ARB", fmt=FMT_AMOUNT)

    # Completeness guarantee: any assumptions.json scalar not already shown in
    # a curated group above is echoed raw here, so nothing is silently dropped.
    leftover = {
        key: value
        for key, value in (assumptions or {}).items()
        if key not in CURATED_ASSUMPTION_KEYS
        and not isinstance(value, (dict, list))
        and value is not None
    }
    if leftover:
        section("Other Assumptions (assumptions.json echo)")
        for key in sorted(leftover):
            entry(key, leftover[key], source="assumptions.json")

    worksheet.column_dimensions["A"].width = 2
    worksheet.column_dimensions["B"].width = 44
    worksheet.column_dimensions["C"].width = 16
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 52
    worksheet.freeze_panes = "A4"
    return bool(derivation)


def _write_case_definition_sections(section, entry, case_config):
    """Site, load-profile, PV and storage definitions straight from case.json."""
    if not case_config:
        return
    site = case_config.get("site") or {}
    load_profile = case_config.get("load_profile") or {}
    if site or load_profile:
        section("Site & Load Profile (case.json)")
        entry("Latitude", site.get("latitude"), unit="deg",
              source="case.json site.latitude", fmt="0.0000")
        entry("Longitude", site.get("longitude"), unit="deg",
              source="case.json site.longitude", fmt="0.0000")
        if load_profile.get("path"):
            entry("Load profile file", load_profile["path"],
                  source="case.json load_profile.path (8760 hourly kW)")
        if load_profile.get("year"):
            entry("Load year", load_profile["year"], fmt="0",
                  source="case.json load_profile.year")

    technologies = case_config.get("technologies") or {}
    pv = technologies.get("pv") or {}
    if pv:
        section("PV Technology (case.json)")
        entry("PV minimum size", pv.get("min_kw"), unit="kW",
              source="case.json technologies.pv.min_kw", fmt=FMT_AMOUNT)
        entry("PV maximum size", pv.get("max_kw"), unit="kW",
              source="case.json technologies.pv.max_kw", fmt=FMT_AMOUNT)
        entry("PV installed cost", pv.get("installed_cost_per_kw"), unit="USD/kW",
              source="case.json technologies.pv.installed_cost_per_kw", fmt=FMT_AMOUNT)
        entry("PV O&M cost", pv.get("om_cost_per_kw"), unit="USD/kW/yr",
              source="case.json technologies.pv.om_cost_per_kw", fmt=FMT_AMOUNT_2)
        entry("PV degradation (case input)", pv.get("degradation_fraction"),
              unit="per year", source="case.json technologies.pv.degradation_fraction",
              fmt=FMT_PERCENT)
        for key, value in (pv.get("pvwatts") or {}).items():
            entry(f"PVWatts {key}", value,
                  source="case.json technologies.pv.pvwatts (production factor inputs)")

    storage = technologies.get("storage") or {}
    if storage:
        section("Storage Technology (case.json)")
        for label, key, unit in STORAGE_CASE_ROWS:
            if storage.get(key) is None:
                continue
            fmt = FMT_PERCENT if key == "om_cost_fraction_of_installed_cost" else FMT_AMOUNT
            entry(label, storage[key], unit=unit,
                  source=f"case.json technologies.storage.{key}", fmt=fmt)


# ---------------------------------------------------------------------------
# Pro Forma (Audit)
# ---------------------------------------------------------------------------

class _ProFormaWriter:
    """Row-oriented writer: line items down, Year 0..N across (SAM layout)."""

    def __init__(self, worksheet, years):
        self.ws = worksheet
        self.years = years          # analysis years (N); columns C..C+N are Year 0..N
        self.row = 1
        self.rows = {}              # key -> row index

    def col(self, year):
        return get_column_letter(3 + year)

    @property
    def first_col(self):
        return self.col(0)

    @property
    def last_col(self):
        return self.col(self.years)

    def year_range(self, row, start=1, end=None):
        end = self.years if end is None else end
        return f"{self.col(start)}{row}:{self.col(end)}{row}"

    def full_range(self, row):
        return f"{self.col(0)}{row}:{self.col(self.years)}{row}"

    def skip(self, count=1):
        self.row += count

    def section(self, label):
        row = self.row
        for year in range(-2, self.years + 1):
            self.ws.cell(row=row, column=3 + year).fill = SECTION_FILL
        cell = self.ws.cell(row=row, column=1, value=label)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        self.row += 1
        return row

    def line(self, key, label, unit, *, y0=None, formula=None, values=None,
             fmt=FMT_AMOUNT, fill=None, bold=False, start=1):
        """Write one line item.

        ``formula(year, col)`` returns the formula for each year >= ``start``;
        ``values`` hardcodes a list indexed by year 0..N; ``y0`` sets Year 0
        explicitly (value or formula string).
        """
        row = self.row
        self.rows[key] = row
        label_cell = self.ws.cell(row=row, column=1, value=label)
        if unit:  # empty strings serialise as corrupt inlineStr cells
            self.ws.cell(row=row, column=2, value=unit).font = NOTE_FONT
        if bold:
            label_cell.font = BOLD_FONT
        if y0 is not None:
            cell = self.ws.cell(row=row, column=3, value=y0)
            cell.number_format = fmt
            if fill:
                cell.fill = fill
            if bold:
                cell.font = BOLD_FONT
        for year in range(start, self.years + 1):
            cell = self.ws.cell(row=row, column=3 + year)
            if values is not None:
                cell.value = values[year] if year < len(values) else 0
            elif formula is not None:
                cell.value = formula(year, self.col(year))
            cell.number_format = fmt
            if fill:
                cell.fill = fill
            if bold:
                cell.font = BOLD_FONT
        self.row += 1
        return row

    def metric(self, key, label, formula_or_value, fmt, note="", fill=None):
        row = self.row
        self.rows[key] = row
        self.ws.cell(row=row, column=1, value=label)
        cell = self.ws.cell(row=row, column=3, value=formula_or_value)
        cell.number_format = fmt
        cell.font = BOLD_FONT
        if fill:
            cell.fill = fill
        if note:
            self.ws.cell(row=row, column=4, value=note).font = NOTE_FONT
        self.row += 1
        return row


def write_pro_forma_audit_sheet(worksheet, cash_flow_result, assumptions):
    """Rebuild the cash flow with live Excel formulas + engine tie-out.

    Returns a refs dict used by the FX Sensitivity sheet and the Cover status:
    ``{"equity_cf_row", "year_row", "years", "status_cells": [...]}``.
    """
    d = cash_flow_result["derivation"]
    summary = cash_flow_result["summary"]
    annual = cash_flow_result["annual_cash_flows"]
    years = d["project_years"]
    is_dppa = d["structure"] == DPPA
    is_physical = d["structure"] == PHYSICAL_DPPA
    is_direct = d["structure"] == DIRECT_OWNERSHIP
    assume_profitable_host = bool((d.get("direct_ownership") or {}).get("assume_profitable_host"))
    construction = d.get("construction")
    # Circular 45 replacement capitalization (default). Present only when a
    # replacement is actually capitalized; the legacy "expense" flag and
    # no-replacement cases carry no block and render the legacy expensed formulas.
    battery_replacement = d.get("battery_replacement")
    # ESCO contract tenor + end-of-term asset transfer (Task 4e). Present only
    # when contract_years is set; drives operating-line truncation beyond year T,
    # the NBV disposal block and the year-T transfer proceeds.
    contract_term = d.get("contract_term")
    # Input VAT on capex (Task 4f). Present only when vat_rate_fraction is set;
    # drives the year-0 outflow / refund-year inflow rows and their equity-cash-
    # flow tie-out. No-VAT cases carry no block and stay byte-identical.
    vat = d.get("vat")
    w = _ProFormaWriter(worksheet, years)

    worksheet.sheet_view.showGridLines = False
    title = worksheet.cell(
        row=1, column=1,
        value="Pro Forma (Audit) — every white cell is a live Excel formula",
    )
    title.font = Font(bold=True, size=13, color=NAVY)
    worksheet.cell(
        row=2, column=1,
        value="Shaded cells are hardcoded engine outputs (REopt dispatch / 8760-h "
              "settlement / tie-out reference). All other cells derive from the "
              "named cells on the Assumptions sheet. Currency: USD at the fixed "
              "contract FX rate.",
    ).font = NOTE_FONT
    w.row = 4

    # --- timeline -----------------------------------------------------------
    w.section("TIMELINE & INDEXATION")
    r_year = w.line("year", "Year", "index", y0=0,
                    formula=lambda y, c: y, fmt="0", bold=True)

    def year_ref(col):
        return f"{col}${r_year}"

    def trunc(c, expr):
        # ESCO contract tenor (Task 4e): every ESCO operating line is zero for
        # years beyond the contract tenor T (the asset has transferred to the
        # host). No-op when no tenor is set, so non-tenor workbooks stay
        # byte-identical.
        if contract_term is None:
            return expr
        return f"IF({year_ref(c)}<=CONTRACT_YEARS,{expr},0)"

    r_fac_energy = w.line(
        "fac_energy", "EVN energy escalation factor", "index",
        formula=lambda y, c: f"=(1+ESC_ENERGY)^({year_ref(c)}-1)", fmt=FMT_FACTOR)
    r_fac_capacity = w.line(
        "fac_capacity", "EVN capacity escalation factor", "index",
        formula=lambda y, c: f"=(1+ESC_CAPACITY)^({year_ref(c)}-1)", fmt=FMT_FACTOR)
    r_fac_om = w.line(
        "fac_om", "O&M escalation factor", "index",
        formula=lambda y, c: f"=(1+ESC_OM)^({year_ref(c)}-1)", fmt=FMT_FACTOR)
    r_fac_deg = w.line(
        "fac_deg", "PV degradation factor", "index",
        formula=lambda y, c: f"=(1-PV_DEGRADATION)^({year_ref(c)}-1)", fmt=FMT_FACTOR)
    if is_dppa:
        r_fac_fee = w.line(
            "fac_fee", "DPPA fee / FMP escalation factor", "index",
            formula=lambda y, c: f"=(1+ESC_FEE)^({year_ref(c)}-1)", fmt=FMT_FACTOR)
        r_fac_strike = w.line(
            "fac_strike", "CfD strike escalation factor", "index",
            formula=lambda y, c: f"=(1+ESC_STRIKE)^({year_ref(c)}-1)", fmt=FMT_FACTOR)
    w.skip()

    # --- revenue ------------------------------------------------------------
    w.section("REVENUE (USD)")
    if is_dppa:
        r_fmp_rev = w.line(
            "fmp_rev", "Generator FMP market revenue", "USD",
            formula=lambda y, c: f"=DPPA_FMP_REV_Y1*{c}{r_fac_fee}*{c}{r_fac_deg}")
        r_cfd_strike = w.line(
            "cfd_strike", "CfD strike leg (P_c × Q_cfd)", "USD",
            formula=lambda y, c: f"=DPPA_CFD_STRIKE_Y1*{c}{r_fac_strike}")
        r_cfd_offset = w.line(
            "cfd_offset", "CfD market offset (FMP × Q_cfd)", "USD",
            formula=lambda y, c: f"=DPPA_CFD_OFFSET_Y1*{c}{r_fac_fee}")
        r_cfd_net = w.line(
            "cfd_net", "CfD net settlement (to generator)", "USD",
            formula=lambda y, c: f"={c}{r_cfd_strike}-{c}{r_cfd_offset}")
        r_energy_rev = w.line(
            "energy_rev", "Generator revenue (FMP + CfD net)", "USD",
            formula=lambda y, c: f"={c}{r_fmp_rev}+{c}{r_cfd_net}", bold=True)
    elif is_physical:
        # ND57 Điều 25 private wire: matched energy × freely negotiated PPA price;
        # price escalates at its own rate (flat by default), volume degrades.
        r_energy_rev = w.line(
            "energy_rev", "PPA energy revenue (matched × price)", "USD",
            formula=lambda y, c:
                f"=PPA_MATCHED_KWH_Y1*PPA_PRICE*(1+PPA_ESC)^({year_ref(c)}-1)"
                f"*{c}{r_fac_deg}")
    elif is_direct:
        # Factory self-invest: the benefit is the FULL avoided EVN bill (energy +
        # demand — BAU/optimized are total bills), captured whole with no ESCO
        # discount / 80/20 demand split. Same trajectories as the offtaker block:
        # BAU − optimized, the residual bill grown by the degradation repurchase.
        r_energy_rev = w.line(
            "energy_rev", "Bill savings (avoided EVN bill: BAU − optimized)", "USD",
            formula=lambda y, c:
                f"=BAU_BILL_Y1*{c}{r_fac_energy}"
                f"-(OPT_BILL_Y1+BASE_SERVED_RETAIL*(1-{c}{r_fac_deg}))*{c}{r_fac_energy}",
            bold=True)
    else:
        r_energy_rev = w.line(
            "energy_rev", "ESCO energy revenue (discount-to-EVN)", "USD",
            formula=lambda y, c: "=" + trunc(
                c, f"BASE_ENERGY_REV*{c}{r_fac_energy}*{c}{r_fac_deg}"))
    if is_direct:
        # Bill savings already folds in demand + any grid arbitrage (BAU/optimized
        # are the total EVN bills), so there is no separate demand-share or
        # arbitrage line; only the optional surplus leg is added.
        surplus = d.get("surplus_export")
        if surplus:
            r_surplus = w.line(
                "surplus_rev", "Surplus export revenue (Decree 243)", "USD",
                formula=lambda y, c:
                    f"=SURPLUS_KWH_Y1*{c}{r_fac_deg}*SURPLUS_PRICE"
                    f"*(1+SURPLUS_ESC)^({year_ref(c)}-1)")
            r_revenue = w.line(
                "revenue", "Total developer revenue", "USD",
                formula=lambda y, c: f"={c}{r_energy_rev}+{c}{r_surplus}", bold=True)
        else:
            r_revenue = w.line(
                "revenue", "Total developer revenue", "USD",
                formula=lambda y, c: f"={c}{r_energy_rev}", bold=True)
    else:
        r_dem_savings = w.line(
            "dem_savings", "Demand charge savings (total)", "USD",
            formula=lambda y, c: f"=BASE_DEMAND_SAVINGS*{c}{r_fac_capacity}")
        r_dem_rev = w.line(
            "dem_rev", "ESCO share of demand savings", "USD",
            formula=lambda y, c: "=" + trunc(c, f"{c}{r_dem_savings}*DEMAND_SHARE"))
        if is_dppa:
            r_arb_rev = None
            r_revenue = w.line(
                "revenue", "Total developer revenue", "USD",
                formula=lambda y, c: f"={c}{r_energy_rev}+{c}{r_dem_rev}", bold=True)
        else:
            r_arb_rev = w.line(
                "arb_rev", "Grid arbitrage revenue", "USD",
                formula=lambda y, c: "=" + trunc(c, f"BASE_GRID_ARB*{c}{r_fac_energy}"))
            surplus = d.get("surplus_export")
            if surplus:
                # Volume degrades with PV output (r_fac_deg); the price escalates
                # independently (annual market re-set proxy). Matches the engine's
                # surplus_export_revenue exactly so the tie-out holds.
                r_surplus = w.line(
                    "surplus_rev", "Surplus export revenue (Decree 243)", "USD",
                    formula=lambda y, c: "=" + trunc(
                        c,
                        f"SURPLUS_KWH_Y1*{c}{r_fac_deg}*SURPLUS_PRICE"
                        f"*(1+SURPLUS_ESC)^({year_ref(c)}-1)"))
                r_revenue = w.line(
                    "revenue", "Total developer revenue", "USD",
                    formula=lambda y, c:
                        f"={c}{r_energy_rev}+{c}{r_dem_rev}+{c}{r_arb_rev}+{c}{r_surplus}",
                    bold=True)
            else:
                r_revenue = w.line(
                    "revenue", "Total developer revenue", "USD",
                    formula=lambda y, c: f"={c}{r_energy_rev}+{c}{r_dem_rev}+{c}{r_arb_rev}",
                    bold=True)
    w.skip()

    # --- operating costs ----------------------------------------------------
    w.section("OPERATING COSTS (USD)")
    r_om = w.line(
        "om", "O&M", "USD",
        formula=lambda y, c: "=" + trunc(c, f"OM_YEAR1*{c}{r_fac_om}"))
    replacement = list(d.get("replacement_costs_by_year_usd") or [])
    replacement_by_year = [0.0] + replacement + [0.0] * years
    r_repl = w.line(
        "repl", "Battery replacement (engine schedule)", "USD",
        values=replacement_by_year[:years + 1], fill=INPUT_FILL)
    r_ebitda = w.line(
        "ebitda", "EBITDA", "USD",
        formula=lambda y, c: f"={c}{r_revenue}-{c}{r_om}-{c}{r_repl}", bold=True)
    w.skip()

    # --- debt ---------------------------------------------------------------
    w.section("DEBT SCHEDULE (USD)")
    # Years >= 2 reference the closing-balance row, which is not written yet;
    # they are filled in right after r_close is known.
    opening_balance_ref = "=COD_DEBT_BALANCE" if construction else "=DEBT_PRINCIPAL"
    r_open = w.line(
        "debt_open", "Opening debt balance", "USD",
        formula=lambda y, c: opening_balance_ref if y == 1 else None)
    r_interest = w.line(
        "interest", "Interest", "USD",
        formula=lambda y, c:
            f"=IF({year_ref(c)}<=DEBT_TERM_YEARS,{c}{r_open}*DEBT_RATE,0)")
    if construction:
        # Principal grace: interest-only rows through year g, amortization
        # starting year g+1 — reproduces the engine's grace debt schedule.
        r_principal = w.line(
            "principal", "Principal repayment", "USD",
            formula=lambda y, c:
                f"=IF(OR({year_ref(c)}<=PRINCIPAL_GRACE_YEARS,"
                f"{year_ref(c)}>DEBT_TERM_YEARS),0,"
                f"MIN(MAX(DEBT_PAYMENT-{c}{r_interest},0),{c}{r_open}))")
    else:
        r_principal = w.line(
            "principal", "Principal repayment", "USD",
            formula=lambda y, c:
                f"=IF({year_ref(c)}<=DEBT_TERM_YEARS,"
                f"MIN(MAX(DEBT_PAYMENT-{c}{r_interest},0),{c}{r_open}),0)")
    r_ds = w.line(
        "debt_service", "Debt service", "USD",
        formula=lambda y, c: f"={c}{r_interest}+{c}{r_principal}", bold=True)
    r_close = w.line(
        "debt_close", "Closing debt balance", "USD",
        formula=lambda y, c: f"={c}{r_open}-{c}{r_principal}")
    # fix the opening-balance back-reference now that r_close is known
    for year in range(2, years + 1):
        worksheet.cell(
            row=r_open, column=3 + year,
            value=f"={w.col(year - 1)}{r_close}",
        )
    w.skip()

    # --- depreciation & tax --------------------------------------------------
    w.section("DEPRECIATION & VIETNAM CIT (USD)")
    if construction:
        # Capitalized IDC joins the depreciable base pro-rata by capex share
        # (Circular 45 borrowing-cost capitalization) and rides each class's
        # existing schedule — ties out to the engine's IDC-inclusive bases.
        r_dep_pv = w.line(
            "dep_pv", "PV depreciation (straight-line)", "USD",
            formula=lambda y, c: "=" + trunc(
                c,
                f"IF({year_ref(c)}<=PV_DEP_YEARS,"
                f"(PV_CAPEX+IDC*PV_CAPEX/(PV_CAPEX+BESS_CAPEX))/PV_DEP_YEARS,0)"))
        r_dep_bess = w.line(
            "dep_bess", "BESS depreciation (straight-line)", "USD",
            formula=lambda y, c: "=" + trunc(
                c,
                f"IF({year_ref(c)}<=BESS_DEP_YEARS,"
                f"(BESS_CAPEX+IDC*BESS_CAPEX/(PV_CAPEX+BESS_CAPEX))/BESS_DEP_YEARS,0)"))
    else:
        r_dep_pv = w.line(
            "dep_pv", "PV depreciation (straight-line)", "USD",
            formula=lambda y, c: "=" + trunc(
                c, f"IF({year_ref(c)}<=PV_DEP_YEARS,PV_CAPEX/PV_DEP_YEARS,0)"))
        r_dep_bess = w.line(
            "dep_bess", "BESS depreciation (straight-line)", "USD",
            formula=lambda y, c: "=" + trunc(
                c, f"IF({year_ref(c)}<=BESS_DEP_YEARS,BESS_CAPEX/BESS_DEP_YEARS,0)"))
    # Circular 45 replacement capitalization (default). Each replacement asset is
    # depreciated straight-line over the BESS class life from its in-service year;
    # the live formula reads the replacement cost off the engine schedule row and
    # divides by the named BESS_DEP_YEARS, active only within its window. Gated on
    # the engine derivation so expense-mode / no-replacement workbooks keep the
    # legacy two-class total-depreciation and EBT formulas byte-for-byte.
    repl_dep_rows = []
    if battery_replacement:
        for schedule in battery_replacement["schedules"]:
            in_service = schedule["in_service_year"]
            cost_cell = f"${w.col(in_service)}${r_repl}"
            repl_dep_rows.append(w.line(
                f"dep_repl_{in_service}",
                f"Replacement depreciation (in-service yr {in_service})", "USD",
                formula=lambda y, c, r=in_service, cost=cost_cell: "=" + trunc(
                    c,
                    f"IF(AND({year_ref(c)}>={r},{year_ref(c)}<={r}+BESS_DEP_YEARS-1),"
                    f"{cost}/BESS_DEP_YEARS,0)")))
    r_dep = w.line(
        "dep", "Total depreciation", "USD",
        formula=lambda y, c: "=" + "+".join(
            [f"{c}{r_dep_pv}", f"{c}{r_dep_bess}"]
            + [f"{c}{row}" for row in repl_dep_rows]
        ))
    # Task 4e asset transfer at the end of contract year T: net book value of
    # every asset class = capitalized cost − cumulative straight-line
    # depreciation through year T (charges beyond T are not taken; the remainder
    # is recovered here, not written off). Disposal gain/(loss) = residual − NBV
    # enters year-T taxable income below (EBT), routed through the case's regime.
    # Gated on the engine derivation so full-horizon cases carry no NBV rows.
    r_disposal = None
    if contract_term:
        w.section("ASSET TRANSFER AT CONTRACT END (Task 4e)")
        if construction:
            pv_basis = "(PV_CAPEX+IDC*PV_CAPEX/(PV_CAPEX+BESS_CAPEX))"
            bess_basis = "(BESS_CAPEX+IDC*BESS_CAPEX/(PV_CAPEX+BESS_CAPEX))"
        else:
            pv_basis = "PV_CAPEX"
            bess_basis = "BESS_CAPEX"
        r_nbv_pv = w.metric(
            "nbv_pv", "NBV of initial PV at transfer",
            f"={pv_basis}-{pv_basis}/PV_DEP_YEARS*MIN(CONTRACT_YEARS,PV_DEP_YEARS)",
            FMT_AMOUNT, note="capitalized cost − cumulative depreciation through T")
        r_nbv_bess = w.metric(
            "nbv_bess", "NBV of initial BESS at transfer",
            f"={bess_basis}-{bess_basis}/BESS_DEP_YEARS"
            "*MIN(CONTRACT_YEARS,BESS_DEP_YEARS)",
            FMT_AMOUNT)
        nbv_rows = [r_nbv_pv, r_nbv_bess]
        if battery_replacement:
            # Each capitalized replacement (Task 4d) contributes its undepreciated
            # remainder: cost − annual charge × charges taken through T.
            for schedule in battery_replacement["schedules"]:
                in_service = schedule["in_service_year"]
                cost_cell = f"${w.col(in_service)}${r_repl}"
                nbv_rows.append(w.metric(
                    f"nbv_repl_{in_service}",
                    f"NBV of replacement (in-service yr {in_service}) at transfer",
                    f"={cost_cell}-{cost_cell}/BESS_DEP_YEARS"
                    f"*(MIN(CONTRACT_YEARS,{in_service}+BESS_DEP_YEARS-1)"
                    f"-{in_service}+1)",
                    FMT_AMOUNT))
        r_nbv_total = w.metric(
            "nbv_total", "Net book value at transfer (total)",
            "=" + "+".join(f"$C${r}" for r in nbv_rows), FMT_AMOUNT)
        r_disposal = w.metric(
            "disposal_gain", "Disposal gain/(loss) at transfer (residual − NBV)",
            f"=RESIDUAL_VALUE-$C${r_nbv_total}", FMT_AMOUNT,
            note="enters year-T taxable income (EBT)")
        w.skip()

    def disposal_term(c):
        # Year-T only: the disposal gain/(loss) joins taxable income through the
        # existing regime/carryforward machinery. Empty when no tenor is set.
        if r_disposal is None:
            return ""
        return f"+IF({year_ref(c)}=CONTRACT_YEARS,$C${r_disposal},0)"

    if battery_replacement:
        # Capitalize mode: EBITDA already subtracted the replacement cash cost, so
        # add it back and deduct the capitalized depreciation instead (total
        # depreciation above now includes the replacement schedules).
        r_ebt = w.line(
            "ebt", "Taxable income before loss relief (EBT)", "USD",
            formula=lambda y, c:
                f"={c}{r_ebitda}+{c}{r_repl}-{c}{r_dep}-{c}{r_interest}"
                + disposal_term(c),
            bold=True)
    else:
        r_ebt = w.line(
            "ebt", "Taxable income before loss relief (EBT)", "USD",
            formula=lambda y, c:
                f"={c}{r_ebitda}-{c}{r_dep}-{c}{r_interest}" + disposal_term(c),
            bold=True)
    if is_direct and assume_profitable_host:
        # Profitable-host convention (DIRECT_OWNERSHIP default): the factory has
        # other taxable profits, so the project's deductions offset them at the
        # flat standard rate every year — a negative EBT yields a negative CIT
        # (an immediate tax shield). No first-profit holiday, no reduced-rate
        # window and no loss carryforward apply. Ties out to the engine's flat
        # calculate_cit(immediate_loss_relief=True) exactly, EBT sign regardless.
        r_cit = w.line(
            "cit", "CIT payable (flat 20%; profitable-host shield)", "USD",
            formula=lambda y, c: f"={c}{r_ebt}*CIT_RATE", bold=True)
    else:
        r_newloss = w.line(
            "newloss", "New tax loss arising", "USD",
            formula=lambda y, c: f"=MAX(-{c}{r_ebt},0)")
        r_income_pos = w.line(
            "income_pos", "Assessable income (pre-relief)", "USD",
            formula=lambda y, c: f"=MAX({c}{r_ebt},0)")

        # FIFO tax-loss carryforward with 5-year expiry (Circular 78/2014 Art. 9):
        # avail[j] = unused loss aged j years; usage consumes oldest vintages first.
        r_avail = {}
        r_used = {}
        for age in range(1, 6):
            r_avail[age] = w.line(
                f"loss_avail_{age}", f"Loss vintage aged {age}y — available", "USD",
                formula=None)
        oldest_first = [5, 4, 3, 2, 1]
        for age in oldest_first:
            def used_formula(y, c, age=age):
                older_terms = "".join(
                    f"-{c}{r_used[a]}" for a in oldest_first[:oldest_first.index(age)]
                )
                return (f"=MIN({c}{r_avail[age]},"
                        f"MAX({c}{r_income_pos}{older_terms},0))")
            r_used[age] = w.line(
                f"loss_used_{age}", f"Loss vintage aged {age}y — utilised", "USD",
                formula=used_formula)
        # availability recursion (needs r_used, so filled after)
        for age in range(1, 6):
            for year in range(1, years + 1):
                cell = worksheet.cell(row=r_avail[age], column=3 + year)
                prev = w.col(year - 1)
                if age == 1:
                    cell.value = 0 if year == 1 else f"={prev}{r_newloss}"
                else:
                    cell.value = (
                        0 if year == 1
                        else f"=MAX({prev}{r_avail[age - 1]}-{prev}{r_used[age - 1]},0)"
                    )
                cell.number_format = FMT_AMOUNT
        r_relief = w.line(
            "relief", "Loss relief utilised (FIFO, ≤5y old)", "USD",
            formula=lambda y, c: "=" + "+".join(f"{c}{r_used[a]}" for a in oldest_first))
        r_taxbase = w.line(
            "taxbase", "Taxable income after loss relief", "USD",
            formula=lambda y, c: f"={c}{r_income_pos}-{c}{r_relief}", bold=True)
        r_flag = w.line(
            "profit_flag", "Profitable-to-date flag (starts CIT clock)", "0/1",
            formula=lambda y, c: (
                f"=IF({c}{r_ebt}>0,1,0)" if y == 1
                else f"=IF(OR({w.col(y - 1)}{w.rows['profit_flag']}=1,{c}{r_ebt}>0),1,0)"
            ), fmt="0")
        r_clock = w.metric(
            "cit_clock", "CIT clock start year (first profit, capped)",
            f"=MIN(IFERROR(MATCH(1,{w.year_range(r_flag)},0),CIT_CLOCK_CAP_YEAR),"
            "CIT_CLOCK_CAP_YEAR)",
            "0", note="Circular 78/2014 Art. 18")
        clock_ref = f"$C${r_clock}"
        cit = d.get("cit", {})
        if cit.get("preferential_rate") is not None:
            # Law 67/2025 re_producer: the applicable base rate is the preferential
            # rate for the first CIT_PREF_YEARS years (counted from year 1), then
            # the standard rate. The 50% reduction multiplies whichever base rate
            # applies. The standard regime keeps CIT_RATE as the base throughout
            # (formula below collapses to the legacy expression, bit-for-bit).
            r_cit_base = w.line(
                "cit_base_rate", "Applicable base CIT rate (preferential window)", "%",
                formula=lambda y, c:
                    f"=IF({year_ref(c)}<=CIT_PREF_YEARS,CIT_PREF_RATE,CIT_RATE)",
                fmt=FMT_PERCENT)

            def base_ref(c):
                return f"{c}{r_cit_base}"
        else:
            def base_ref(c):
                return "CIT_RATE"
        r_cit_rate = w.line(
            "cit_rate", "Applicable CIT rate", "%",
            formula=lambda y, c: (
                f"=IF({year_ref(c)}<{clock_ref}+CIT_HOLIDAY_YEARS,0,"
                f"IF({year_ref(c)}<{clock_ref}+CIT_HOLIDAY_YEARS+CIT_REDUCED_YEARS,"
                f"{base_ref(c)}*CIT_REDUCED_FRACTION,{base_ref(c)}))"
            ), fmt=FMT_PERCENT)
        r_cit = w.line(
            "cit", "CIT payable", "USD",
            formula=lambda y, c: f"={c}{r_taxbase}*{c}{r_cit_rate}", bold=True)
    w.skip()

    # --- cash flows -----------------------------------------------------------
    w.section("CASH FLOW & COVERAGE (USD)")
    r_cfads = w.line(
        "cfads", "Cash available for debt service (CFADS)", "USD",
        formula=lambda y, c: f"={c}{r_ebitda}-{c}{r_cit}", bold=True)
    r_proj = w.line(
        "project_cf", "Project cash flow (unlevered, post-tax)", "USD",
        y0="=-TOTAL_CAPEX",
        formula=lambda y, c: f"={c}{r_cfads}")
    # Task 4e: the host's end-of-term buyout lands in the developer's year-T
    # equity cash flow as its own disclosed line — an equity-side terminal value,
    # so the PROCEEDS leave CFADS and the unlevered project cash flow unchanged.
    # The disposal gain's tax effect is separate: it enters year-T taxable
    # income, so CIT (and thus CFADS and the project cash flow) already carry it.
    r_transfer = None
    if contract_term:
        r_transfer = w.line(
            "transfer_proceeds", "Asset transfer proceeds (host buyout)", "USD",
            formula=lambda y, c: f"=IF({year_ref(c)}=CONTRACT_YEARS,RESIDUAL_VALUE,0)")
    # Task 4f: input VAT on capex is paid at year 0 (−VAT_RATE × total capex) and
    # refunded (VAT_RATE × total capex) in the refund year. Both are equity-funded
    # and out of CFADS/DSCR/project cash flow — they only join the equity cash
    # flow below. Gated so no-VAT workbooks stay byte-identical.
    r_vat_paid = None
    r_vat_refund = None
    if vat:
        r_vat_paid = w.line(
            "input_vat_paid", "Input VAT paid on capex", "USD",
            y0="=-VAT_RATE*TOTAL_CAPEX",
            formula=lambda y, c: f"=IF({year_ref(c)}=0,-VAT_RATE*TOTAL_CAPEX,0)")
        r_vat_refund = w.line(
            "vat_refund", "VAT refund received", "USD",
            y0="=IF(VAT_REFUND_YEAR=0,VAT_RATE*TOTAL_CAPEX,0)",
            formula=lambda y, c:
                f"=IF({year_ref(c)}=VAT_REFUND_YEAR,VAT_RATE*TOTAL_CAPEX,0)")
    eq_y0 = "=-EQUITY_INVESTMENT"
    if r_vat_paid:
        eq_y0 += f"+{w.col(0)}{r_vat_paid}+{w.col(0)}{r_vat_refund}"
    r_eq = w.line(
        "equity_cf", "Equity cash flow", "USD",
        y0=eq_y0,
        formula=lambda y, c: (
            f"={c}{r_cfads}-{c}{r_ds}"
            + (f"+{c}{r_transfer}" if r_transfer else "")
            + (f"+{c}{r_vat_paid}+{c}{r_vat_refund}" if r_vat_paid else "")
        ), bold=True)
    r_cum = w.line(
        "cum_equity", "Cumulative equity cash flow", "USD",
        y0=f"={w.col(0)}{r_eq}",
        formula=lambda y, c: f"={w.col(y - 1)}{w.rows['cum_equity']}+{c}{w.rows['equity_cf']}")
    r_dscr = w.line(
        "dscr", "DSCR (debt years)", "x",
        formula=lambda y, c: f"=IF({c}{r_ds}>0,{c}{r_cfads}/{c}{r_ds},\"\")",
        fmt=FMT_RATIO)
    w.skip()

    # --- offtaker -------------------------------------------------------------
    w.section("OFFTAKER (BUYER) POSITION (USD)")
    r_bau = w.line(
        "bau_bill", "BAU EVN bill (energy + demand)", "USD",
        formula=lambda y, c: f"=BAU_BILL_Y1*{c}{r_fac_energy}")
    if is_dppa:
        r_c_dn = w.line(
            "c_dn", "C_DN spot energy (Q_Khc × CFMP × K_pp)", "USD",
            formula=lambda y, c: f"=DPPA_C_DN_Y1*{c}{r_fac_fee}*{c}{r_fac_deg}")
        r_c_dppa = w.line(
            "c_dppa", "C_DPPA system service fee", "USD",
            formula=lambda y, c: f"=DPPA_C_DPPA_Y1*{c}{r_fac_fee}*{c}{r_fac_deg}")
        r_c_cl = w.line(
            "c_cl", "C_CL settlement adder", "USD",
            formula=lambda y, c: f"=DPPA_C_CL_Y1*{c}{r_fac_fee}*{c}{r_fac_deg}")
        r_c_bl = w.line(
            "c_bl", "C_BL retail shortfall (incl. degradation repurchase)", "USD",
            formula=lambda y, c:
                f"=(DPPA_C_BL_Y1+DPPA_MATCHED_RETAIL_Y1*(1-{c}{r_fac_deg}))"
                f"*{c}{r_fac_energy}")
        r_post = w.line(
            "post_cost", "Buyer cost with project", "USD",
            formula=lambda y, c:
                f"={c}{r_c_dn}+{c}{r_c_dppa}+{c}{r_c_cl}+{c}{r_c_bl}"
                f"+{c}{w.rows['cfd_net']}+OPT_DEMAND_Y1*{c}{r_fac_capacity}"
                f"+{c}{r_dem_rev}",
            bold=True)
    elif is_direct:
        # The factory IS the investor: it pays no ESCO fee, so its residual cost
        # is just the optimized EVN bill (energy + demand, grown by the
        # degradation repurchase). Savings = BAU − optimized = the bill-savings
        # revenue line above — the buyer view equals the developer view.
        r_post = w.line(
            "post_cost", "Buyer cost with project (residual EVN bill)", "USD",
            formula=lambda y, c:
                f"=(OPT_BILL_Y1+BASE_SERVED_RETAIL*(1-{c}{r_fac_deg}))"
                f"*{c}{r_fac_energy}",
            bold=True)
    else:
        r_post = w.line(
            "post_cost", "Buyer cost with project", "USD",
            formula=lambda y, c:
                f"=(OPT_BILL_Y1+BASE_SERVED_RETAIL*(1-{c}{r_fac_deg}))"
                f"*{c}{r_fac_energy}"
                f"+{c}{r_energy_rev}+{c}{r_dem_rev}+{c}{r_arb_rev}",
            bold=True)
    r_savings = w.line(
        "savings", "Buyer savings vs BAU", "USD",
        formula=lambda y, c: f"={c}{r_bau}-{c}{r_post}", bold=True)
    w.line(
        "savings_pct", "Buyer savings (% of BAU)", "%",
        formula=lambda y, c: f"=IF({c}{r_bau}=0,\"\",{c}{r_savings}/{c}{r_bau})",
        fmt=FMT_PERCENT)
    w.skip()

    # --- return metrics ---------------------------------------------------------
    w.section("RETURN METRICS (Excel formulas)")
    ten = min(10, years)
    r_irr_eq = w.metric(
        "m_equity_irr", "Equity IRR", f"=IRR({w.full_range(r_eq)})", FMT_PERCENT)
    r_irr_proj = w.metric(
        "m_project_irr", "Project IRR (unlevered)",
        f"=IRR({w.full_range(r_proj)})", FMT_PERCENT)
    r_npv = w.metric(
        "m_npv", "Equity NPV @ owner discount rate",
        f"=NPV(DISC_RATE,{w.year_range(r_eq)})+{w.col(0)}{r_eq}", FMT_AMOUNT)
    r_min_dscr = w.metric(
        "m_min_dscr", "Minimum DSCR (debt years)",
        f"=MIN({w.year_range(r_dscr)})", FMT_RATIO,
        note="MIN ignores the blank non-debt years")
    r_avg_dscr = w.metric(
        "m_avg_dscr", "Average DSCR (debt years)",
        f"=AVERAGE({w.year_range(r_dscr)})", FMT_RATIO)
    r_pay_match = w.metric(
        "m_payback_year", "First year cumulative equity CF ≥ 0",
        f"=MATCH(TRUE,INDEX({w.year_range(r_cum)}>=0,0),0)", "0")
    r_payback = w.metric(
        "m_payback", "Simple equity payback",
        f"=IF(ISNA($C${r_pay_match}),\"n/a\",$C${r_pay_match}-1"
        f"+IFERROR(-INDEX({w.col(0)}{r_cum}:{w.last_col}{r_cum},$C${r_pay_match})"
        f"/INDEX({w.year_range(r_eq)},$C${r_pay_match}),0))",
        FMT_YEARS, note="years")
    # Task 4f: when input VAT is on, the year-0 equity cell already carries the
    # true equity outlay (EQUITY_INVESTMENT plus the VAT paid net of any
    # year-0 refund), so the denominator ties to that cell instead of the
    # VAT-exclusive EQUITY_INVESTMENT named range — otherwise a refund landing
    # in an operating year (included in the SUM) would be credited against
    # capital the denominator never charged. No-VAT workbooks keep the
    # original formula byte-for-byte.
    roi_denominator = f"-{w.col(0)}{r_eq}" if r_vat_paid else "EQUITY_INVESTMENT"
    r_roi = w.metric(
        "m_roi", "ROI (cumulative equity CF / equity)",
        f"=SUM({w.year_range(r_eq)})/{roi_denominator}", FMT_PERCENT)
    r_sav10 = w.metric(
        "m_savings_10yr", "Buyer savings, years 1-10 (% of BAU)",
        f"=SUM({w.year_range(r_savings, 1, ten)})/SUM({w.year_range(r_bau, 1, ten)})",
        FMT_PERCENT)
    w.skip()

    # --- engine tie-out ----------------------------------------------------------
    w.section("CHECKS — EXCEL vs ENGINE (per-year tie-out)")
    eq_engine = [-summary["equity_investment_usd"]] + [
        row["equity_cash_flow_usd"] for row in annual
    ]
    if vat:
        # Year-0 input VAT on capex rides the equity cash flow, not the equity
        # investment — mirror the engine so the tie-out reference matches the
        # live equity_cf year-0 cell (a year-0 refund nets it back).
        eq_engine[0] -= vat["vat_amount_usd"]
        if vat["refund_year"] == 0:
            eq_engine[0] += vat["vat_amount_usd"]
    cfads_engine = [None] + [
        row["cash_available_for_debt_service_usd"] for row in annual
    ]
    cit_engine = [None] + [row["cit_usd"] for row in annual]
    r_eq_engine = w.line(
        "eq_engine", "Equity cash flow (engine)", "USD",
        y0=eq_engine[0], values=eq_engine, fill=INPUT_FILL)
    r_cfads_engine = w.line(
        "cfads_engine", "CFADS (engine)", "USD",
        values=cfads_engine, fill=INPUT_FILL)
    r_cit_engine = w.line(
        "cit_engine", "CIT (engine)", "USD",
        values=cit_engine, fill=INPUT_FILL)
    r_delta = w.line(
        "delta", "Max |Excel − engine| this year", "USD",
        y0=f"=ABS({w.col(0)}{r_eq}-{w.col(0)}{r_eq_engine})",
        formula=lambda y, c: (
            f"=MAX(ABS({c}{r_eq}-{c}{r_eq_engine}),"
            f"ABS({c}{r_cfads}-{c}{r_cfads_engine}),"
            f"ABS({c}{r_cit}-{c}{r_cit_engine}))"
        ), fmt="0.0000")
    w.skip()

    header_row = w.row
    for column, header in enumerate(
            ("Check", None, "Excel", "Engine", "Delta", "Status"), start=1):
        cell = worksheet.cell(row=header_row, column=column)
        if header:
            cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    w.row += 1

    status_cells = []

    def check(label, excel_ref, engine_value, tolerance, fmt):
        row = w.row
        worksheet.cell(row=row, column=1, value=label)
        excel_cell = worksheet.cell(row=row, column=3, value=f"={excel_ref}")
        excel_cell.number_format = fmt
        engine_cell = worksheet.cell(row=row, column=4, value=engine_value)
        engine_cell.number_format = fmt
        engine_cell.fill = INPUT_FILL
        delta = worksheet.cell(row=row, column=5, value=f"=ABS(C{row}-D{row})")
        delta.number_format = "0.000000"
        status = worksheet.cell(
            row=row, column=6,
            value=f"=IF(ABS(C{row}-D{row})<={_fmt_num(tolerance)},\"PASS\",\"REVIEW\")",
        )
        status.fill = CHECK_FILL
        status.font = BOLD_FONT
        status_cells.append(f"F{row}")
        w.row += 1

    min_dscr_engine = _minimum_dscr(annual)
    check("Per-year cash flow tie-out (max delta, all years)",
          f"MAX({w.full_range(r_delta)})", 0.0, TOL_AMOUNT, "0.0000")
    if summary.get("equity_irr_fraction") is not None:
        check("Equity IRR", f"$C${r_irr_eq}", summary["equity_irr_fraction"],
              TOL_RATE, FMT_PERCENT)
    if summary.get("project_irr_fraction") is not None:
        check("Project IRR", f"$C${r_irr_proj}", summary["project_irr_fraction"],
              TOL_RATE, FMT_PERCENT)
    check("Equity NPV", f"$C${r_npv}", summary["npv_usd"], TOL_AMOUNT, FMT_AMOUNT)
    if min_dscr_engine is not None:
        check("Minimum DSCR", f"$C${r_min_dscr}", min_dscr_engine, TOL_RATIO, FMT_RATIO)
    if summary.get("average_dscr") is not None:
        check("Average DSCR", f"$C${r_avg_dscr}", summary["average_dscr"],
              TOL_RATIO, FMT_RATIO)
    if summary.get("simple_payback_years") is not None:
        check("Simple equity payback", f"$C${r_payback}",
              summary["simple_payback_years"], TOL_YEARS, FMT_YEARS)
    if summary.get("roi_fraction") is not None:
        check("ROI", f"$C${r_roi}", summary["roi_fraction"], TOL_RATE, FMT_PERCENT)
    if summary.get("buyer_savings_10yr_fraction") is not None:
        check("Buyer savings, years 1-10 (% of BAU)", f"$C${r_sav10}",
              summary["buyer_savings_10yr_fraction"], TOL_RATE, FMT_PERCENT)

    debt_sizing = d.get("debt_sizing")
    if debt_sizing and min_dscr_engine is not None:
        # DSCR debt-sizing fixed-point tie-out. DEBT_PRINCIPAL is the live
        # min(FRACTION_DEBT, SUPPORTED_DEBT), so the debt schedule above is
        # driven by the sized loan. When the covenant binds (SUPPORTED_DEBT <
        # FRACTION_DEBT) the MIN DSCR of that schedule must equal the covenant;
        # when it does not, DEBT_PRINCIPAL falls back to the fraction-based
        # principal and the MIN DSCR clears the covenant. A single live formula
        # self-determines which branch applies.
        row = w.row
        worksheet.cell(row=row, column=1, value="DSCR-sized debt fixed point")
        excel_cell = worksheet.cell(row=row, column=3, value=f"=$C${r_min_dscr}")
        excel_cell.number_format = FMT_RATIO
        target_cell = worksheet.cell(
            row=row, column=4, value=debt_sizing["target_min_dscr"])
        target_cell.number_format = FMT_RATIO
        target_cell.fill = INPUT_FILL
        delta = worksheet.cell(row=row, column=5, value=f"=ABS(C{row}-D{row})")
        delta.number_format = "0.000000"
        status = worksheet.cell(
            row=row, column=6,
            value=(
                "=IF(IF(SUPPORTED_DEBT<FRACTION_DEBT,"
                f"ABS($C${r_min_dscr}-TARGET_MIN_DSCR)<={_fmt_num(TOL_RATIO)},"
                f"AND($C${r_min_dscr}>=TARGET_MIN_DSCR-{_fmt_num(TOL_RATIO)},"
                f"ABS(DEBT_PRINCIPAL-FRACTION_DEBT)<={_fmt_num(TOL_AMOUNT)})),"
                "\"PASS\",\"REVIEW\")"
            ),
        )
        status.fill = CHECK_FILL
        status.font = BOLD_FONT
        status_cells.append(f"F{row}")
        w.row += 1

    # highlight REVIEW statuses in red
    status_range = f"F{header_row + 1}:F{w.row - 1}"
    worksheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"REVIEW"'],
                   font=Font(bold=True, color="9C0006"),
                   fill=PatternFill("solid", fgColor="FFC7CE")),
    )

    worksheet.column_dimensions["A"].width = 48
    worksheet.column_dimensions["B"].width = 8
    for year in range(0, years + 1):
        worksheet.column_dimensions[w.col(year)].width = 13
    worksheet.freeze_panes = f"C{r_year + 1}"

    return {
        "sheet": worksheet.title,
        "years": years,
        "year_row": r_year,
        "equity_cf_row": r_eq,
        "cfads_row": r_cfads,
        "debt_service_row": r_ds,
        "status_cells": status_cells,
        "status_range": status_range,
    }


def _minimum_dscr(annual_rows):
    values = [
        row["dscr"] for row in annual_rows
        if row.get("debt_service_usd") and row.get("dscr") is not None
    ]
    return min(values) if values else None


# ---------------------------------------------------------------------------
# FX Sensitivity
# ---------------------------------------------------------------------------

def write_fx_sensitivity_sheet(worksheet, cash_flow_result, proforma_refs):
    """USD equity returns under annual VND depreciation — live formulas.

    For USD-denominated debt (gated on the engine derivation) the debt service
    is FX-fixed while VND revenue deflates, so the equity helper reproduces the
    engine's decomposition (CFADS_t/(1+d)^t − debt_service_t) and a DSCR-vs-
    depreciation column is added. VND-debt (default) workbooks are unchanged.
    """
    derivation = cash_flow_result.get("derivation", {})
    is_usd_debt = derivation.get("debt_currency") == "USD"
    engine_rows = calculate_fx_sensitivity(cash_flow_result)
    show_dscr = is_usd_debt and any(
        row["min_dscr"] is not None for row in engine_rows
    )
    years = proforma_refs["years"]
    eq_row = proforma_refs["equity_cf_row"]
    cfads_row = proforma_refs["cfads_row"]
    ds_row = proforma_refs["debt_service_row"]
    sheet_ref = f"'{proforma_refs['sheet']}'"

    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:H1" if show_dscr else "A1:F1")
    title = worksheet.cell(row=1, column=1, value="FX Sensitivity — VND depreciation vs USD returns")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    worksheet.row_dimensions[1].height = 24
    if is_usd_debt:
        note = (
            "Cash flows are VND-denominated (EVN tariff / DPPA settlement) but reported in USD at "
            "the fixed contract rate. Debt is USD-denominated, so its service is FX-fixed while the "
            "rest of the equity flow (CFADS) deflates by (1+d)^t: adjusted_t = CFADS_t/(1+d)^t − "
            "debt_service_t, and DSCR erodes as (CFADS_t/(1+d)^t)/debt_service_t. CIT is not "
            "recomputed under drift; FX revaluation of the USD principal is not modelled. The rate "
            "cells are editable; engine columns validate the default scenarios."
        )
    else:
        note = (
            "Cash flows are VND-denominated (EVN tariff / DPPA settlement) but reported in USD "
            "at the fixed contract rate. Each scenario deflates the year-t USD equity cash flow "
            "by (1+d)^t. Debt is assumed VND-denominated, so DSCR is unchanged. The rate cells "
            "are editable; engine columns validate the default scenarios."
        )
    worksheet.cell(row=2, column=1, value=note).font = NOTE_FONT

    header_row = 4
    headers = ["VND depreciation (per year)", "Equity IRR (USD)", "Equity NPV (USD)",
               "Equity IRR (engine)", "Equity NPV (engine)", "Status"]
    if show_dscr:
        headers += ["Min DSCR (USD)", "Min DSCR (engine)"]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    helper_start = header_row + len(engine_rows) + 3
    worksheet.cell(
        row=helper_start - 1, column=1,
        value="Helper — FX-adjusted equity cash flow per scenario (USD): "
              "CF_t / (1 + d)^t",
    ).font = NOTE_FONT
    year_header_row = helper_start
    for year in range(0, years + 1):
        col = get_column_letter(2 + year)
        cell = worksheet.cell(row=year_header_row, column=2 + year, value=year)
        cell.font = BOLD_FONT
        cell.number_format = "0"

    # Second helper block: FX-adjusted DSCR per debt year (USD debt only).
    dscr_helper_start = year_header_row + len(engine_rows) + 2
    if show_dscr:
        worksheet.cell(
            row=dscr_helper_start - 1, column=1,
            value="Helper — FX-adjusted DSCR per debt year (USD): "
                  "(CFADS_t/(1+d)^t) / debt_service_t",
        ).font = NOTE_FONT

    status_cells = []
    for index, engine in enumerate(engine_rows):
        table_row = header_row + 1 + index
        helper_row = year_header_row + 1 + index
        rate_cell = worksheet.cell(row=table_row, column=1, value=engine["vnd_depreciation_rate"])
        rate_cell.number_format = FMT_PERCENT
        rate_cell.fill = INPUT_FILL

        worksheet.cell(row=helper_row, column=1, value=f"=A{table_row}").number_format = FMT_PERCENT
        for year in range(0, years + 1):
            col = get_column_letter(2 + year)
            proforma_col = get_column_letter(3 + year)
            if is_usd_debt and year >= 1:
                # USD debt service is FX-fixed; the whole VND equity leg (equity
                # cash flow + debt service — i.e. CFADS plus the equity-side
                # residual / VAT refund) deflates, then debt service is netted.
                value = (
                    f"=(({sheet_ref}!{proforma_col}{eq_row}"
                    f"+{sheet_ref}!{proforma_col}{ds_row})"
                    f"/(1+$A{table_row})^{col}${year_header_row})"
                    f"-{sheet_ref}!{proforma_col}{ds_row}"
                )
            else:
                value = (
                    f"={sheet_ref}!{proforma_col}{eq_row}"
                    f"/(1+$A{table_row})^{col}${year_header_row}"
                )
            worksheet.cell(row=helper_row, column=2 + year, value=value).number_format = FMT_AMOUNT
        helper_range = f"B{helper_row}:{get_column_letter(2 + years)}{helper_row}"
        first_year_range = f"C{helper_row}:{get_column_letter(2 + years)}{helper_row}"

        irr_cell = worksheet.cell(row=table_row, column=2, value=f"=IRR({helper_range})")
        irr_cell.number_format = FMT_PERCENT
        npv_cell = worksheet.cell(
            row=table_row, column=3,
            value=f"=NPV(DISC_RATE,{first_year_range})+B{helper_row}",
        )
        npv_cell.number_format = FMT_AMOUNT
        engine_irr = worksheet.cell(row=table_row, column=4, value=engine["equity_irr_fraction"])
        engine_irr.number_format = FMT_PERCENT
        engine_irr.fill = INPUT_FILL
        engine_npv = worksheet.cell(row=table_row, column=5, value=engine["npv_usd"])
        engine_npv.number_format = FMT_AMOUNT
        engine_npv.fill = INPUT_FILL

        if show_dscr:
            # Per-debt-year DSCR helper row: (CFADS_t/(1+d)^t)/debt_service_t,
            # blank when debt service is 0 so MIN spans the debt term only.
            dscr_helper_row = dscr_helper_start + index
            worksheet.cell(
                row=dscr_helper_row, column=1, value=f"=A{table_row}"
            ).number_format = FMT_PERCENT
            for year in range(1, years + 1):
                col = get_column_letter(2 + year)
                proforma_col = get_column_letter(3 + year)
                worksheet.cell(
                    row=dscr_helper_row, column=2 + year,
                    value=(
                        f"=IF({sheet_ref}!{proforma_col}{ds_row}>0,"
                        f"({sheet_ref}!{proforma_col}{cfads_row}"
                        f"/(1+$A{table_row})^{col}${year_header_row})"
                        f"/{sheet_ref}!{proforma_col}{ds_row},\"\")"
                    ),
                ).number_format = FMT_RATIO
            dscr_range = f"C{dscr_helper_row}:{get_column_letter(2 + years)}{dscr_helper_row}"
            min_dscr_cell = worksheet.cell(
                row=table_row, column=7, value=f"=MIN({dscr_range})"
            )
            min_dscr_cell.number_format = FMT_RATIO
            engine_dscr = worksheet.cell(
                row=table_row, column=8, value=engine["min_dscr"]
            )
            engine_dscr.number_format = FMT_RATIO
            engine_dscr.fill = INPUT_FILL
            status_formula = (
                f"=IF(AND(ABS(B{table_row}-D{table_row})<={_fmt_num(TOL_RATE)},"
                f"ABS(C{table_row}-E{table_row})<={_fmt_num(TOL_AMOUNT)},"
                f"ABS(G{table_row}-H{table_row})<={_fmt_num(TOL_RATIO)}),"
                f"\"PASS\",\"REVIEW\")"
            )
        else:
            status_formula = (
                f"=IF(AND(ABS(B{table_row}-D{table_row})<={_fmt_num(TOL_RATE)},"
                f"ABS(C{table_row}-E{table_row})<={_fmt_num(TOL_AMOUNT)}),"
                f"\"PASS\",\"REVIEW\")"
            )
        status = worksheet.cell(row=table_row, column=6, value=status_formula)
        status.fill = CHECK_FILL
        status.font = BOLD_FONT
        status_cells.append(f"F{table_row}")

    status_range = f"F{header_row + 1}:F{header_row + len(engine_rows)}"
    worksheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"REVIEW"'],
                   font=Font(bold=True, color="9C0006"),
                   fill=PatternFill("solid", fgColor="FFC7CE")),
    )

    worksheet.column_dimensions["A"].width = 28
    for letter in ("B", "C", "D", "E", "F"):
        worksheet.column_dimensions[letter].width = 20
    if show_dscr:
        for letter in ("G", "H"):
            worksheet.column_dimensions[letter].width = 20

    return {"sheet": worksheet.title, "status_range": status_range}


# ---------------------------------------------------------------------------
# Model Basis
# ---------------------------------------------------------------------------

def _settlement_title_suffix(is_dppa, is_physical, is_direct=False):
    if is_dppa:
        return " — ND57/2025 grid DPPA with CfD"
    if is_physical:
        return " — ND57 Điều 25 physical (private-wire) DPPA"
    if is_direct:
        return " — direct ownership (factory self-invest)"
    return " — ESCO discount-to-EVN"


def _settlement_bullets(is_dppa, is_physical, is_direct=False, assume_profitable_host=True):
    if is_dppa:
        return [
            "Q_adj[h] = Q_re_meter[h] / K_pp × delta — quantity conversion generator→customer uses "
            "K_pp only (CD7 Ví dụ 1); k is price-only.",
            "Q_Khc[h] = min(load[h], Q_adj[h]) — the buyer settles only matched consumption; surplus "
            "stays with the generator at FMP and is never billed to the buyer.",
            "C_DN = Q_Khc × CFMP × K_pp with CFMP = FMP × k;  C_DPPA = Q_Khc × f_dppa;  C_CL = Q_Khc × f_cl;  "
            "C_BL = shortfall × EVN TOU retail.",
            "CfD settles on min(Q_c, Q_Khc) per CD7 Ví dụ 4: CfD[h] = (P_c − FMP[h]) × Q_cfd[h]. The strike "
            "leg escalates at the contracted strike escalation; the market legs at the fee escalation.",
            "Curtailed PV from the self-consumption REopt run is credited as grid export at FMP (the "
            "generator does not curtail under DPPA).",
            "Year-1 settlement is computed hourly; later years scale the year-1 totals by the escalation "
            "and degradation factors shown on the Pro Forma (Audit) sheet.",
        ]
    if is_physical:
        return [
            "The generator sells matched (project-served) energy — PV→load plus battery→load — directly "
            "to the factory over a private line at a freely negotiated PPA price; Decree 243/2026 removed "
            "the ceiling for the directly-traded volume.",
            "No EVN grid settlement chain applies to the contracted energy (no k/K_pp, CFMP, f_dppa/f_cl, "
            "or C_BL decomposition) — the private wire bypasses the grid.",
            "Surplus (PV→grid + would-be-curtailed) is sold to EVN at the Decree 243 market price, capped "
            "at the Decision 988 regional ceiling and at 50% of PV output — the same machinery as the "
            "ESCO surplus leg.",
            "Demand-charge savings vs BAU are shared with the developer at the contracted share, as in the "
            "ESCO case: the factory's grid demand falls identically under a private wire.",
            "Year-1 volumes come from the 8760-hour REopt dispatch; the PPA leg escalates at its own "
            "negotiated rate (flat by default), the surplus/residual legs at EVN escalation, and all volumes "
            "degrade with PV, per the factors on the Pro Forma (Audit) sheet.",
        ]
    if is_direct:
        host_bullet = (
            "Profitable-host convention (default): the factory has other taxable profits, so the project's "
            "deductions offset them at the flat 20% rate every year — a loss year yields a NEGATIVE CIT "
            "(immediate tax shield), with no loss carryforward."
            if assume_profitable_host else
            "Standalone tax treatment (assume_profitable_host = No): a loss year pays no CIT and its loss "
            "carries forward FIFO for up to 5 years, exactly as for a standalone project."
        )
        return [
            "The factory self-invests: it owns the PV/BESS asset, borrows the debt, and pays O&M and "
            "replacements. Its benefit is the FULL avoided EVN bill — the buyer's natural benchmark against "
            "an ESCO or DPPA offer (\"what if we just built it ourselves?\").",
            "Bill savings = BAU EVN bill − optimized residual EVN bill, the full delta (energy + demand). "
            "There is NO ESCO discount and NO 80/20 demand-savings split — the factory captures everything; "
            "the buyer-analysis savings view therefore equals the developer view.",
            "May sell rooftop surplus to EVN under Decree 243/2026 (PV→grid + would-be-curtailed), capped "
            "at the Decision 988 regional ceiling and 50% of PV output — the same machinery as the ESCO "
            "surplus leg; omitted → surplus not monetized (conservative).",
            "CIT: flat standard 20% every year (standard_flat) — no first-profit holiday and no RE-producer "
            "preferential rate. A factory adding rooftop solar to an existing operation gets no new-project "
            "incentive on its general income and is not a licensed RE generator.",
            host_bullet,
            "Incremental taxable income = bill savings + surplus − O&M − replacement − depreciation − "
            "interest; the developer metrics and DSCR are computed on the factory's incremental cash flow.",
        ]
    return [
        "The ESCO is paid a contracted fraction of the time-specific EVN tariff for project-served "
        "energy (PV→load, plus battery→load when the battery cannot grid-charge).",
        "Demand-charge savings vs BAU are split between ESCO and offtaker at the contracted share.",
        "Year-1 revenue is computed from the 8760-hour REopt dispatch × TOU rates; later years scale "
        "by EVN escalation and PV degradation factors shown on the Pro Forma (Audit) sheet.",
    ]


def write_model_basis_sheet(worksheet, assumptions, derivation):
    derivation = derivation or {}
    is_dppa = derivation.get("structure") == DPPA
    is_physical = derivation.get("structure") == PHYSICAL_DPPA
    is_direct = derivation.get("structure") == DIRECT_OWNERSHIP
    assume_profitable_host = bool(
        (derivation.get("direct_ownership") or {}).get("assume_profitable_host")
    )
    fx = derivation.get("exchange_rate_vnd_per_usd") or assumptions.get("exchange_rate_vnd_per_usd")

    # Construction/IDC/grace disclosures, gated on the engine derivation so
    # overnight-build (default) cases stay byte-identical.
    construction = derivation.get("construction")
    debt_bullets = [
        "Debt: level-payment annuity over the debt term (interest + principal split per the Debt "
        "Schedule block).",
    ]
    construction_register_bullets = []
    if construction:
        months = construction.get("construction_months", 0)
        grace = construction.get("principal_grace_years", 0)
        debt_bullets.append(
            f"Construction period ({months} months): all pre-COD flows are collapsed at the model's "
            "year-0 point — operations years and the equity outflow timing are unchanged. This "
            "collapsed-year-0 convention slightly flatters IRR for long construction periods "
            "(disclosed simplification)."
        )
        debt_bullets.append(
            "IDC = debt fraction × capex × debt rate × (months/12) ÷ 2 — even drawdown of the "
            "debt-funded capex over construction, simple interest on the average balance (half the "
            "final draw), no compounding. IDC is debt-funded (rolled up): the COD debt balance is "
            "the drawn principal plus IDC; equity is unchanged. IDC is capitalized into the "
            "depreciable base pro-rata across the asset classes by capex share (VAS / Circular 45 "
            "borrowing-cost capitalization), not expensed."
        )
        if grace:
            debt_bullets.append(
                f"Principal grace: years 1-{grace} are interest-only on the full COD balance; "
                "principal then amortizes over the remaining (term − grace) years as a level-payment "
                "annuity. DSCR during grace reflects interest-only debt service."
            )
        construction_register_bullets.append(
            "Construction period and capitalized IDC are modelled for this case (default is an "
            "overnight build): pre-COD flows stay collapsed at year 0 with the closed-form IDC "
            "above; multi-year construction equity phasing is not modelled."
        )

    # USD-denominated debt disclosures, gated on the engine derivation so
    # VND-debt (default) cases carry no USD-debt bullets and stay byte-identical.
    usd_debt = derivation.get("debt_currency") == "USD"
    usd_debt_register_bullets = []
    if usd_debt:
        debt_bullets.append(
            "Debt currency: USD (international financing, ~5% default rate vs the 8.5% VND commercial-"
            "bank rate). The base case holds the contract FX flat, so the debt schedule, IDC, DSCR "
            "and interest tax deduction run in USD exactly as VND debt would — only the default rate "
            "differs. The FX exposure is surfaced on the FX Sensitivity sheet: under VND depreciation "
            "d, USD debt service is FX-fixed while VND revenue (CFADS) deflates by (1+d)^t, so "
            "adjusted equity = CFADS_t/(1+d)^t − debt_service_t and DSCR erodes as "
            "(CFADS_t/(1+d)^t)/debt_service_t (VND debt keeps a depreciation-invariant DSCR)."
        )
        usd_debt_register_bullets.append(
            "USD-denominated debt FX exposure is quantified on the FX Sensitivity sheet as a "
            "deflation overlay only: CIT is NOT recomputed under FX drift (including the interest "
            "deduction of the USD loan), and VAS FX revaluation gains/losses on the outstanding USD "
            "principal are not modelled."
        )

    # DSCR-driven debt-sizing disclosures, gated on the engine derivation so
    # fraction-sized (default) cases carry no sizing bullets and stay byte-
    # identical.
    debt_sizing = derivation.get("debt_sizing")
    debt_sizing_register_bullets = []
    if debt_sizing:
        target = debt_sizing.get("target_min_dscr")
        binds = debt_sizing.get("binding_constraint") == "dscr"
        debt_bullets.append(
            f"Debt sizing: the loan is min(fraction-based, DSCR-supported) at a {target}x minimum-DSCR "
            "covenant. Because debt service scales linearly with the principal, the DSCR-supported loan "
            "is found by a fixed-point iteration (supported(D) = D × min_DSCR(D) / target, re-running the "
            "full debt / IDC / CIT / CFADS derivation each round) so the CIT and IDC feedbacks are solved, "
            "not approximated. "
            + (
                "This covenant binds: the sized loan sits below the fraction-based principal and the "
                "minimum DSCR over the debt term equals the covenant (see the Debt-sizing fixed-point "
                "check). Equity absorbs the gap (equity = capex − sized debt)."
                if binds else
                "This covenant does not bind: the fraction-based principal already clears it, so the loan "
                "and every downstream metric are identical to the un-sized case."
            )
        )
        debt_sizing_register_bullets.append(
            "DSCR debt sizing uses level-payment sizing (no sculpted repayment profile) on base-case "
            "CFADS only — the contract FX is held flat (no FX-drift overlay) and no downside/stress case "
            "is used to size the loan."
        )

    # Battery replacement CIT treatment. Default (capitalize): the register line
    # disclosing expensing is REMOVED and replaced by a Model Basis bullet
    # describing the Circular 45 capitalized treatment and the truncation
    # convention. Legacy "expense" flag (and no-replacement cases): the expensing
    # register line is kept, byte-for-byte with the pre-change workbook.
    battery_replacement = derivation.get("battery_replacement")
    if battery_replacement:
        replacement_basis_bullets = [
            "Battery replacement is CAPITALIZED, not expensed (VAS / Circular 45/2013): each replacement "
            "battery is a fixed asset depreciated straight-line over the 8-year BESS class life from its "
            "in-service (replacement) year, with each replacement year carrying its own schedule. The "
            "replacement cash outflow is unchanged — only the CIT deduction timing shifts from a single "
            "full deduction to the depreciation stream.",
        ]
        replacement_register_bullets = [
            "Battery replacement depreciation is truncated at the analysis horizon: charges beyond the "
            "final year are simply not taken and the undepreciated remainder is not written off (no "
            "terminal disposal/salvage of the replaced or residual battery is modelled).",
        ]
    else:
        replacement_basis_bullets = []
        replacement_register_bullets = [
            "Battery replacement is expensed in the replacement year, not capitalized and re-depreciated.",
        ]

    # ESCO contract tenor + end-of-term asset transfer (Task 4e). When active the
    # "no terminal/residual value" register line is replaced by the modeled
    # treatment; default (no tenor) keeps that line byte-for-byte.
    contract_term = derivation.get("contract_term")
    contract_basis_bullets = []
    contract_register_bullets = []
    if contract_term:
        tenor = contract_term.get("contract_years")
        contract_basis_bullets.append(
            f"ESCO contract tenor: operations run for {tenor} years, then the PV/BESS asset "
            "transfers to the host at a contractual residual/buyout value. Every ESCO-side line "
            "(energy / demand / arbitrage / surplus revenue, O&M, battery replacement and "
            "depreciation) is zero beyond year T; REopt-scheduled replacements after T are not "
            "incurred (the asset has transferred). The residual payment lands in the developer's "
            "year-T equity cash flow as its own disclosed line."
        )
        contract_basis_bullets.append(
            "Asset-transfer tax: the disposal gain/(loss) = residual − net book value at year T "
            "(each asset class's capitalized cost, IDC included, minus straight-line depreciation "
            "taken through T) enters year-T taxable income through the case's CIT regime and "
            "loss-carryforward machinery — no separate disposal-tax computation. Depreciation "
            "beyond T is not taken; the undepreciated remainder is recovered via the disposal, "
            "not written off separately (see the Pro Forma NBV tie-out). Asymmetry: the transfer "
            "proceeds are equity-side (out of CFADS/DSCR) while this disposal tax flows through CIT "
            "into CFADS and the project cash flow, so the year-T DSCR bears the tax without the "
            "offsetting proceeds — a lender may normalize that year."
        )
        contract_register_bullets.append(
            "Asset-disposal tax convention: Vietnamese CIT generally taxes asset-disposal gains at "
            "the standard rate outside preferential regimes; routing the transfer gain/(loss) "
            "through the case's CIT regime (rather than the standard rate) is a modeling convention."
        )
        contract_register_bullets.append(
            "Post-transfer economics are out of scope: the Buyer Analysis covers the contract term "
            "as-is; no continued-merchant-ESCO alternative, no balloon refinancing (the tenor is "
            "required to be at least the debt term), and no salvage/disposal costs are modelled."
        )
        terminal_value_register_line = (
            "No working-capital or DSRA is modelled; an ESCO contract tenor with end-of-term asset "
            "transfer at a residual/buyout value IS modelled (operations truncated at year T, "
            "NBV-based disposal gain/(loss) through the CIT regime — see the asset-transfer bullets "
            "and the Pro Forma NBV tie-out)."
        )
    else:
        terminal_value_register_line = (
            "No working-capital, DSRA, or terminal/residual value is modelled."
        )

    # Input VAT on capex (Task 4f). When active the "VAT is out of scope" register
    # line is replaced by the modeled capex-VAT timing plus the remaining VAT
    # simplifications; default (no VAT) keeps that line byte-for-byte.
    vat_block = derivation.get("vat")
    vat_basis_bullets = []
    if vat_block:
        vat_rate = vat_block.get("rate")
        vat_refund_year = vat_block.get("refund_year")
        vat_basis_bullets.append(
            "Input VAT on capex (VAT Law 48/2024/QH15, effective 2025-07-01): the project company pays "
            f"input VAT at the user-supplied rate ({vat_rate:.0%} of total capex — the standard rate is "
            "10%, with temporary reductions for some goods that the model does not adjudicate) at "
            f"purchase (year 0) and recovers it in full via the investment-project refund mechanism in "
            f"year {vat_refund_year}. Both flows are shown gross on the developer's equity cash flow "
            "(net-zero timing when refunded in year 0)."
        )
        vat_basis_bullets.append(
            "The refundable input VAT is EQUITY-funded — it does not enter the debt principal, the "
            "debt-fraction base, IDC or DSCR sizing — and the refund is NOT operating revenue (kept out "
            "of CFADS/DSCR as a one-off, non-operating item), so it moves equity IRR and NPV only. "
            "Creditable input VAT is neither income nor expense: no CIT effect and no depreciation-base "
            "effect (assets stay excl.-VAT)."
        )
        vat_register_line = (
            "Capex input-VAT timing IS modelled (paid year 0, refunded in the configured year, "
            "equity-funded, returns-only). Remaining VAT simplifications: the operating-stage VAT float "
            "(output VAT on invoices vs input VAT on O&M) nets to ~zero in annual buckets and stays "
            "disclosed as pass-through; the refund is assumed full and on time (partial / denied / "
            "delayed refunds out of scope); battery replacement stays excl.-VAT like all O&M-stage costs."
        )
    else:
        vat_register_line = "VAT is out of scope (pass-through assumed for both parties)."

    if derivation.get("cit", {}).get("regime") == "re_producer":
        cit_regime_text = (
            "CIT regime: renewable-energy producer (Law 67/2025/QH15 + Decree "
            "320/2025/NĐ-CP) — 10% preferential base rate for the first 15 years "
            "counted from the first revenue-generating year, then 20%."
        )
    elif derivation.get("cit", {}).get("regime") == "standard_flat":
        cit_regime_text = (
            "CIT regime: flat standard 20% every year (standard_flat) — no "
            "first-profit holiday and no RE-producer preferential rate. A factory "
            "self-investing in rooftop solar gets no new-project incentive on its "
            "general income and is not a licensed RE generator. "
            + (
                "Under the profitable-host convention a loss year yields a negative "
                "CIT (immediate shield against the host's other profits), with no "
                "carryforward."
                if assume_profitable_host else
                "Under standalone treatment a loss year pays no CIT and carries "
                "forward FIFO for up to 5 years."
            )
        )
    else:
        cit_regime_text = (
            "CIT regime: standard rate with holiday (Law 67/2025; legacy Circular "
            "78/2014 shape) — 20% base rate throughout. Conservative default for a "
            "service ESCO, whose RE-producer status is a legal question; "
            "grandfathered projects keep old-law incentives."
        )

    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("B1:C1")
    title = worksheet.cell(row=1, column=2, value="Model Basis, Conventions & Simplifications")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    worksheet.row_dimensions[1].height = 24

    sections = [
        ("1. Model architecture", [
            "REopt (NLR optimization engine) selects PV/BESS sizing and hourly dispatch against the "
            "EVN time-of-use tariff over an 8760-hour year.",
            "proforma_vietnam post-processes the REopt run: an hourly ND57/2025 DPPA settlement layer "
            "(when applicable) and a 25-year developer cash flow with Vietnam tax and debt.",
            "This workbook is generated from that engine. The Pro Forma (Audit) sheet re-derives the "
            "full cash flow with live Excel formulas from the named inputs on the Assumptions sheet; "
            "the Checks block ties every metric back to the engine (PASS/REVIEW).",
        ]),
        ("2. Currency & FX", [
            f"All money flows are computed in USD at the fixed contract rate ({fx:,.0f} VND/USD)."
            if fx else
            "All money flows are computed in USD at the fixed contract exchange rate.",
            "Underlying revenue is VND-denominated (EVN tariff, FMP market, DPPA fees). Holding FX flat "
            "for the analysis period is a simplification: the FX Sensitivity sheet quantifies USD-reported "
            "returns under 0-3%/yr VND depreciation.",
            "Sheets that show VND amounts (Hourly/Monthly Settlement, DPPA fee inputs) are VND-native "
            "regulatory quantities, not conversions.",
            "Debt is USD-denominated (international financing), so DSCR is FX-exposed — see the FX "
            "Sensitivity sheet."
            if usd_debt else
            "Debt is assumed VND-denominated (local bank), so DSCR is FX-neutral.",
        ]),
        ("3. Settlement math" + _settlement_title_suffix(is_dppa, is_physical, is_direct),
         _settlement_bullets(is_dppa, is_physical, is_direct, assume_profitable_host)),
        ("4. Multi-year mechanics", [
            "PV degradation compounds on generation-linked terms; energy lost to degradation is repurchased "
            "from EVN at retail (added to the buyer's residual bill / C_BL).",
            "O&M escalates at its own rate; battery replacement is booked in the configured year at REopt "
            "replacement unit costs.",
            *debt_bullets,
            cit_regime_text if is_direct else (
                cit_regime_text + " The 4-year exemption and 9-year 50%-reduction periods count from the "
                "first profitable year, no later than year 4; the 50% reduction applies to the "
                "then-applicable base rate. Tax losses carry forward at most 5 consecutive years, consumed "
                "FIFO — the carryforward schedule is fully visible on the Pro Forma sheet."
            ),
            "Straight-line depreciation: PV over the configured life within the 7-20y band of Circular "
            "45/2013/TT-BTC; BESS over its own life.",
            *replacement_basis_bullets,
            *contract_basis_bullets,
            *vat_basis_bullets,
        ]),
        ("5. Simplifications register (disclosed for audit)", [
            "Fixed FX over the analysis period — quantified on the FX Sensitivity sheet.",
            *construction_register_bullets,
            *usd_debt_register_bullets,
            *debt_sizing_register_bullets,
            *replacement_register_bullets,
            *contract_register_bullets,
            vat_register_line,
            terminal_value_register_line,
            "A single dispatch year (8760 h) is escalated; no re-dispatch in later years.",
            "REopt sizing is optimizer output and can vary slightly between solver versions; the financial "
            "post-processing is deterministic given results.json.",
        ]),
        ("6. How to validate this workbook", [
            "1) Assumptions sheet: confirm every shaded input against the case file / contract term sheet.",
            "2) Pro Forma (Audit): trace any line — formulas reference named cells (e.g. ESC_ENERGY, "
            "DEBT_RATE) and prior rows only.",
            "3) Checks block: every metric shows Excel vs engine with PASS/REVIEW at stated tolerances "
            "(amounts $1, rates 5 bp, DSCR 0.005, payback 0.05 yr).",
            "4) FX Sensitivity: rate cells are editable; Excel recomputes IRR/NPV live.",
            "5) Technical sheets (Technical Results, Dispatch Profile, Load Duration, Settlement) are "
            "engine output for record. The Dispatch sheet's PV production factor (kWh/kW, PVWatts-derived) "
            "is the hourly solar-resource signal — REopt does not persist raw irradiance.",
        ]),
        ("7. Key references", [
            "ND57/2025 (DPPA decree) Art. 14-18 — settlement chain and eligibility.",
            "NSMO/CD7 simulation examples — k price-only conversion (Ví dụ 1), CfD cap on matched volume (Ví dụ 4).",
            "Law 67/2025/QH15 + Decree 320/2025/NĐ-CP — CIT rates, RE-producer preferential incentive, "
            "holiday & loss carryforward (legacy Circular 78/2014 Art. 9, 18 shape kept for the standard regime).",
            "Circular 45/2013/TT-BTC — fixed-asset depreciation bands.",
            "EVN retail tariff (current & QĐ963 TOU structures) as configured in the case file.",
        ]),
    ]

    row = 3
    for heading, bullets in sections:
        cell = worksheet.cell(row=row, column=2, value=heading)
        cell.font = Font(bold=True, size=11, color=NAVY)
        row += 1
        for bullet in bullets:
            bullet_cell = worksheet.cell(row=row, column=2, value="•  " + bullet)
            bullet_cell.alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.row_dimensions[row].height = max(
                14, 13 * (1 + len(bullet) // 105)
            )
            row += 1
        row += 1

    worksheet.column_dimensions["A"].width = 2
    worksheet.column_dimensions["B"].width = 110


# ---------------------------------------------------------------------------
# Cover
# ---------------------------------------------------------------------------

def write_cover_sheet(worksheet, workbook, assumptions, derivation,
                      check_ranges=None):
    derivation = derivation or {}
    is_dppa = derivation.get("structure") == DPPA
    is_physical = derivation.get("structure") == PHYSICAL_DPPA
    is_direct = derivation.get("structure") == DIRECT_OWNERSHIP
    case_name = (assumptions or {}).get("case_name", "Vietnam ESCO / DPPA Case")

    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("B2:E2")
    title = worksheet.cell(row=2, column=2, value=f"Investment Case Workbook — {case_name}")
    title.fill = TITLE_FILL
    title.font = Font(bold=True, color="FFFFFF", size=15)
    worksheet.row_dimensions[2].height = 30

    if is_dppa:
        subtitle = "Grid-connected DPPA with CfD (ND57/2025)"
    elif is_physical:
        subtitle = "Physical (private-wire) DPPA — ND57 Điều 25 (Decree 243/2026)"
    elif is_direct:
        subtitle = "Direct ownership — factory self-invest benchmark (avoided EVN bill)"
    else:
        subtitle = "ESCO discount-to-EVN tariff (behind-the-meter)"
    worksheet.cell(
        row=3, column=2,
        value=f"{subtitle}  ·  prepared {date.today().isoformat()}  ·  "
              "REopt dispatch + proforma_vietnam financial engine",
    ).font = NOTE_FONT
    if (assumptions or {}).get("run_uuid"):
        worksheet.cell(
            row=4, column=2, value=f"REopt run UUID: {assumptions['run_uuid']}"
        ).font = NOTE_FONT

    row = 6
    if check_ranges:
        worksheet.cell(row=row, column=2, value="Model validation status").font = BOLD_FONT
        countifs = "+".join(
            f"COUNTIF('{sheet}'!{rng},\"REVIEW\")" for sheet, rng in check_ranges
        )
        status = worksheet.cell(
            row=row, column=3,
            value=f"=IF({countifs}=0,\"ALL CHECKS PASS\",\"REVIEW REQUIRED\")",
        )
        status.font = Font(bold=True, size=12, color=NAVY)
        worksheet.cell(
            row=row, column=4,
            value="Live tie-out of every Excel-derived metric against the engine",
        ).font = NOTE_FONT
        row += 2

    worksheet.cell(row=row, column=2, value="Colour legend").font = BOLD_FONT
    row += 1
    legend_input = worksheet.cell(row=row, column=2, value="Hardcoded input / engine output")
    legend_input.fill = INPUT_FILL
    worksheet.cell(
        row=row, column=3,
        value="Case file terms, REopt dispatch results, 8760-h settlement totals, tie-out references",
    ).font = NOTE_FONT
    row += 1
    worksheet.cell(row=row, column=2, value="Live Excel formula (white)")
    worksheet.cell(
        row=row, column=3,
        value="Everything derivable is a formula — trace with Excel's formula auditing tools",
    ).font = NOTE_FONT
    row += 1
    legend_check = worksheet.cell(row=row, column=2, value="PASS / REVIEW check cell")
    legend_check.fill = CHECK_FILL
    worksheet.cell(
        row=row, column=3, value="Excel-vs-engine tie-out at stated tolerances"
    ).font = NOTE_FONT
    row += 2

    worksheet.cell(row=row, column=2, value="Contents").font = BOLD_FONT
    row += 1
    guide = [
        ("Executive Summary", "Headline KPIs for both counterparties"),
        ("Assumptions", "Every case input & contract term with unit, source and named cell"),
        ("Model Basis", "Methodology, settlement math, simplifications register"),
        (PRO_FORMA_SHEET, "Full cash flow rebuilt with live formulas + engine tie-out"),
        ("FX Sensitivity", "USD returns under VND depreciation (editable)"),
        ("Buyer Analysis", "Offtaker savings vs business-as-usual"),
        ("Developer Returns", "Sources & uses, coverage, equity cash flow"),
        ("Technical Results", "System sizing, year-1 energy balance, bill comparison"),
        ("Dispatch Profile", "8760-h dispatch incl. original PV generation; chart shows the peak-load week"),
        ("Load Duration", "Load and net-load duration curves"),
    ]
    if is_dppa:
        guide.append(("Year 1 BAU vs DPPA", "Side-by-side year-1 buyer/seller position"))
        guide.append(("Monthly / Hourly Settlement",
                      "ND57 fee chain, VND-native (full configuration on Assumptions)"))
    for sheet_name, description in guide:
        worksheet.cell(row=row, column=2, value=sheet_name).font = BOLD_FONT
        worksheet.cell(row=row, column=3, value=description).font = NOTE_FONT
        row += 1

    row += 1
    for note in (
        "Prepared for investment and lending review. Figures are model outputs, not offers; "
        "REopt sizing is optimizer output and may shift with solver versions.",
        "Regenerate offline anytime: python -m proforma_vietnam.rebuild_report --case-dir <case>.",
    ):
        worksheet.cell(row=row, column=2, value=note).font = NOTE_FONT
        row += 1

    worksheet.column_dimensions["A"].width = 2
    worksheet.column_dimensions["B"].width = 46
    worksheet.column_dimensions["C"].width = 30
    worksheet.column_dimensions["D"].width = 60
    worksheet.column_dimensions["E"].width = 14
