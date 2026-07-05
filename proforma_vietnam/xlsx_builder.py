from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Alignment, Font, PatternFill

from proforma_vietnam import audit_sheets
from proforma_vietnam import proforma_schema as schema
from proforma_vietnam.dppa_settlement import DPPA_TYPE_GRID_CFD

# Proforma line-item columns (Cash Flow, Tax, Debt, DPPA Annual) and the
# Summary / Developer Financials metric rows are derived from proforma_schema at
# render time, filtered by structure. The remaining constants below describe
# DPPA settlement breakouts and REopt-derived report data, which are not
# proforma line-items.

DPPA_HOURLY_COLUMNS = [
    ("Hour", "hour"),
    ("Load (kW)", "load_kw"),
    ("Q_re_meter (kW)", "q_re_meter_kw"),
    ("Q_re_delivered (kW)", "q_re_delivered_kw"),
    ("Q_adj (kW)", "q_adj_kw"),
    ("Q_Khc (kW)", "q_khc_kw"),
    ("FMP (VND/kWh)", "fmp_vnd_per_kwh"),
    ("C_DN (VND)", "c_dn_vnd"),
    ("C_DPPA (VND)", "c_dppa_vnd"),
    ("C_CL (VND)", "c_cl_vnd"),
    ("C_BL (VND)", "c_bl_vnd"),
    ("CfD Payment (VND)", "cfd_payment_vnd"),
]

DPPA_MONTHLY_COLUMNS = [
    ("Month", "month"),
    ("C_DN (VND)", "c_dn_vnd"),
    ("C_DPPA (VND)", "c_dppa_vnd"),
    ("C_CL (VND)", "c_cl_vnd"),
    ("C_BL (VND)", "c_bl_vnd"),
    ("CfD Net (VND)", "cfd_net_vnd"),
    ("Generator Revenue (VND)", "generator_revenue_vnd"),
    ("Customer Total (VND)", "customer_total_vnd"),
]

# Year-1 BAU vs DPPA comparison sheet. Each row pulls from the year-1 cash-flow
# row + assumptions; ("section", None, None) renders as a section header.
BAU_VS_DPPA_ROWS = [
    ("section", "Buyer (Factory) — Year 1 Outflow (USD)", None),
    ("BAU EVN energy + demand bill", "bau_evn_bill_usd", "bau"),
    ("C_DN spot energy (Q_Khc × CFMP × K_pp)", "c_dn_usd", "dppa"),
    ("C_DPPA system service fee", "c_dppa_usd", "dppa"),
    ("C_CL settlement adder", "c_cl_usd", "dppa"),
    ("C_BL retail shortfall (P_evn × shortfall)", "c_bl_usd", "dppa"),
    ("CfD payment to ESCO (positive = pay)", "cfd_net_usd", "dppa"),
    ("Optimized demand charge (kept by EVN)", "optimized_demand_charge_year_usd", "dppa"),
    ("ESCO demand savings share charged back", "esco_demand_revenue_usd", "dppa"),
    ("Total buyer outflow", "total_buyer_outflow_usd", "both"),
    ("Buyer savings vs BAU", "buyer_savings_vs_bau_usd", "delta"),
    ("Buyer savings as % of BAU", "buyer_savings_fraction", "delta_fraction"),
    ("section", "Seller (ESCO / Generator) — Year 1 Revenue (USD)", None),
    ("Discount-to-EVN energy revenue (Phase 2)", "esco_phase2_energy_revenue_usd", "bau_esco"),
    ("FMP market revenue (Q_re_meter × FMP)", "generator_fmp_revenue_usd", "dppa"),
    ("CfD net (from buyer)", "cfd_net_usd", "dppa"),
    ("Demand savings share (80%)", "esco_demand_revenue_usd", "both_esco"),
    ("ESCO grid arbitrage revenue", "esco_grid_arbitrage_revenue_usd", "both_esco"),
    ("Total ESCO / generator revenue", "total_seller_revenue_usd", "both_esco"),
    ("section", "System Energy Flows (kWh / yr)", None),
    ("PV → load", "pv_to_load_kwh", "flow"),
    ("PV → storage", "pv_to_storage_kwh", "flow"),
    ("PV → grid (export at FMP)", "pv_to_grid_effective_kwh", "flow"),
    ("Storage → load", "storage_to_load_kwh", "flow"),
    ("Q_re_meter (total generator injection)", "q_re_meter_kwh", "flow"),
    ("Q_adj (loss-adjusted customer-side)", "q_adj_kwh", "flow"),
    ("Q_Khc (matched consumption, settled)", "q_khc_kwh", "flow"),
    ("EVN retail shortfall", "shortfall_kwh", "flow"),
]

SYSTEM_SIZING_ROWS = [
    ("PV Size (kW)", "pv_kw"),
    ("Battery Power (kW)", "battery_kw"),
    ("Battery Energy (kWh)", "battery_kwh"),
]

RESULTS_COMPARISON_ROWS = [
    ("BAU Utility Bill (USD)", "bau_utility_bill_usd"),
    ("Optimized Utility Bill (USD)", "optimized_utility_bill_usd"),
    ("Utility Bill Savings (USD)", "utility_bill_savings_usd"),
    ("BAU Demand Charge (USD)", "bau_demand_charge_usd"),
    ("Optimized Demand Charge (USD)", "optimized_demand_charge_usd"),
    ("Demand Charge Savings (USD)", "demand_charge_savings_usd"),
]  # keys are genuinely USD (report_data._results_comparison)

ANNUAL_PRODUCTION_ROWS = [
    ("PV to Load (kWh)", "pv_to_load_kwh"),
    ("PV to Storage (kWh)", "pv_to_storage_kwh"),
    ("PV to Grid (kWh)", "pv_to_grid_kwh"),
    ("Storage to Load (kWh)", "storage_to_load_kwh"),
    ("Grid to Load (kWh)", "grid_to_load_kwh"),
    ("PV Curtailed (kWh)", "pv_curtailed_kwh"),
    ("Grid to Storage (kWh)", "grid_to_storage_kwh"),
]

# Hourly dispatch table. "PV Generation Total" is the original PV output before
# the dispatch split; the production factor is the hourly solar-resource signal
# (REopt does not persist raw irradiance — see Model Basis).
DISPATCH_COLUMNS = [
    ("Hour", "hour"),
    ("Load (kW)", "load_kw"),
    ("PV Production Factor (kWh/kW) — irradiation proxy", "pv_production_factor"),
    ("PV Generation Total (kW)", "pv_total_kw"),
    ("PV to Load (kW)", "pv_to_load_kw"),
    ("PV to Storage (kW)", "pv_to_storage_kw"),
    ("PV to Grid (kW)", "pv_to_grid_kw"),
    ("PV Curtailed (kW)", "pv_curtailed_kw"),
    ("Grid to Load (kW)", "grid_to_load_kw"),
    ("Grid to Storage (kW)", "grid_to_storage_kw"),
    ("Storage to Load (kW)", "storage_to_load_kw"),
]

DISPATCH_CHART_SERIES = [
    ("Load (kW)", 2),
    ("PV Generation Total (kW)", 4),
    ("Grid to Load (kW)", 9),
    ("Storage to Load (kW)", 11),
]

HOURS_PER_WEEK = 168

LOAD_DURATION_COLUMNS = [
    ("Rank", "rank"),
    ("Load (kW)", "load_kw"),
    ("Net Load (kW)", "net_load_kw"),
]

NAVY = "1F3864"
ACCENT = "2E74B5"
LIGHT_BAND = "EDF2F9"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
SECTION_FILL = PatternFill("solid", fgColor=ACCENT)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBTITLE_FONT = Font(italic=True, size=9, color="595959")
NOTE_FONT = Font(italic=True, size=9, color="595959")
KPI_FONT = Font(bold=True, size=11)

FORMAT_AMOUNT = "#,##0"
FORMAT_PERCENT = "0.0%"
FORMAT_RATIO = "0.00"
FORMAT_YEARS = "0.0"

# Sheets that carry their own bespoke layout and column widths.
CUSTOM_LAYOUT_SHEETS = {
    "Cover",
    "Executive Summary",
    "Buyer Analysis",
    "Developer Returns",
    "Assumptions",
    "Model Basis",
    audit_sheets.PRO_FORMA_SHEET,
    "FX Sensitivity",
}

BUYER_ANNUAL_COLUMNS = [
    ("Year", "year", None),
    ("BAU Cost (USD)", "bau_evn_bill_usd", FORMAT_AMOUNT),
    ("Cost With Project (USD)", "offtaker_post_project_cost_usd", FORMAT_AMOUNT),
    ("Savings (USD)", "offtaker_savings_usd", FORMAT_AMOUNT),
    ("Savings (% of BAU)", "offtaker_savings_fraction", FORMAT_PERCENT),
]

DEVELOPER_ANNUAL_COLUMNS = [
    ("Year", "year", None),
    ("Revenue (USD)", "esco_revenue_usd", FORMAT_AMOUNT),
    ("O&M + Replacement (USD)", None, FORMAT_AMOUNT),
    ("CIT (USD)", "cit_usd", FORMAT_AMOUNT),
    ("CFADS (USD)", "cash_available_for_debt_service_usd", FORMAT_AMOUNT),
    ("Debt Service (USD)", "debt_service_usd", FORMAT_AMOUNT),
    ("DSCR", "dscr", FORMAT_RATIO),
    ("Equity Cash Flow (USD)", "equity_cash_flow_usd", FORMAT_AMOUNT),
]


def build_vietnam_esco_workbook(cash_flow_result, assumptions=None, report_data=None):
    report_data = report_data or {}
    assumptions = assumptions or {}
    dppa_config = _active_dppa_config(assumptions)
    derivation = cash_flow_result.get("derivation")

    workbook = Workbook()
    cover_sheet = workbook.active
    cover_sheet.title = "Cover"
    _write_executive_summary(
        workbook.create_sheet("Executive Summary"),
        cash_flow_result, assumptions, report_data, dppa_config,
    )
    audit_sheets.write_assumptions_sheet(
        workbook.create_sheet("Assumptions"), workbook, assumptions, derivation
    )
    audit_sheets.write_model_basis_sheet(
        workbook.create_sheet("Model Basis"), assumptions, derivation
    )
    check_ranges = []
    if derivation:
        proforma_refs = audit_sheets.write_pro_forma_audit_sheet(
            workbook.create_sheet(audit_sheets.PRO_FORMA_SHEET),
            cash_flow_result,
            assumptions,
        )
        check_ranges.append((proforma_refs["sheet"], proforma_refs["status_range"]))
        fx_refs = audit_sheets.write_fx_sensitivity_sheet(
            workbook.create_sheet("FX Sensitivity"), cash_flow_result, proforma_refs
        )
        check_ranges.append((fx_refs["sheet"], fx_refs["status_range"]))
    audit_sheets.write_cover_sheet(
        cover_sheet, workbook, assumptions, derivation, check_ranges
    )
    _write_buyer_analysis(
        workbook.create_sheet("Buyer Analysis"), cash_flow_result, dppa_config
    )
    _write_developer_returns(
        workbook.create_sheet("Developer Returns"), cash_flow_result
    )

    # Per-year record tables (Summary, Cash Flow, Tax Schedule, Debt Service,
    # Developer Financials, DPPA Annual Summary, DPPA Configuration) were
    # consolidated: the Pro Forma (Audit) sheet carries every line-item with
    # engine tie-out rows, and Assumptions carries the full DPPA configuration.
    _write_technical_results(
        workbook.create_sheet("Technical Results"), report_data
    )
    _write_dispatch_sheet(
        workbook.create_sheet("Dispatch Profile"),
        report_data.get("dispatch_profile", []),
    )
    _write_table_sheet(
        workbook.create_sheet("Load Duration"),
        LOAD_DURATION_COLUMNS,
        report_data.get("load_duration", []),
        chart_title="Load Duration Curve",
    )

    if dppa_config is not None:
        _write_bau_vs_dppa_sheet(
            workbook.create_sheet("Year 1 BAU vs DPPA"),
            cash_flow_result,
            report_data,
            assumptions,
        )
        _write_table_sheet(
            workbook.create_sheet("Monthly Settlement"),
            DPPA_MONTHLY_COLUMNS,
            report_data.get("dppa_monthly_breakout", []),
        )
        _write_table_sheet(
            workbook.create_sheet("Hourly Settlement"),
            DPPA_HOURLY_COLUMNS,
            report_data.get("dppa_hourly_breakout", []),
        )

    for worksheet in workbook.worksheets:
        if worksheet.title not in CUSTOM_LAYOUT_SHEETS:
            _autosize_columns(worksheet)

    for title in CUSTOM_LAYOUT_SHEETS:
        if title in workbook.sheetnames:
            workbook[title].sheet_properties.tabColor = NAVY

    return workbook


def _minimum_debt_service_dscr(cash_flow_result):
    dscr_values = [
        row["dscr"]
        for row in cash_flow_result.get("annual_cash_flows", [])
        if row.get("debt_service_vnd") and row.get("dscr") is not None
    ]
    return min(dscr_values) if dscr_values else None


def _write_title(worksheet, title, subtitle, last_column):
    worksheet.merge_cells(
        start_row=1, start_column=2, end_row=1, end_column=last_column
    )
    cell = worksheet.cell(row=1, column=2, value=title)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 26
    worksheet.cell(row=2, column=2, value=subtitle).font = SUBTITLE_FONT
    worksheet.sheet_view.showGridLines = False


def _write_section_header(worksheet, row, label, last_column):
    for column in range(2, last_column + 1):
        worksheet.cell(row=row, column=column).fill = SECTION_FILL
    cell = worksheet.cell(row=row, column=2, value=label)
    cell.font = SECTION_FONT


def _write_kpi_rows(worksheet, start_row, rows):
    """Write (label, value, number_format, note) tuples as label/value lines."""
    row = start_row
    for label, value, number_format, note in rows:
        worksheet.cell(row=row, column=2, value=label)
        value_cell = worksheet.cell(row=row, column=3, value=value)
        value_cell.font = KPI_FONT
        if number_format:
            value_cell.number_format = number_format
        if note:
            worksheet.cell(row=row, column=4, value=note).font = NOTE_FONT
        if row % 2 == 0:
            for column in (2, 3, 4):
                worksheet.cell(row=row, column=column).fill = PatternFill(
                    "solid", fgColor=LIGHT_BAND
                )
        row += 1
    return row


def _write_executive_summary(worksheet, cash_flow_result, assumptions, report_data, dppa_config):
    summary = cash_flow_result.get("summary", {})
    annual_rows = cash_flow_result.get("annual_cash_flows", []) or [{}]
    year_one = annual_rows[0]
    sizing = report_data.get("system_sizing", {}) or {}

    case_name = assumptions.get("case_name") or "Vietnam ESCO / DPPA Case"
    contract_label = (
        "Grid-connected DPPA with CfD (ND57/2025)" if dppa_config else
        "ESCO discount-to-EVN tariff (behind-the-meter)"
    )
    _write_title(
        worksheet,
        f"Investment & PPA Negotiation Summary — {case_name}",
        f"{contract_label}  ·  prepared {date.today().isoformat()}  ·  "
        "REopt dispatch + proforma_vietnam financial model",
        last_column=4,
    )

    lifetime_bau = sum(_lookup(row, "bau_evn_bill_usd") or 0.0 for row in annual_rows)
    lifetime_savings = sum(_lookup(row, "offtaker_savings_usd") or 0.0 for row in annual_rows)
    minimum_dscr = _minimum_debt_service_dscr(cash_flow_result)

    row = 4
    _write_section_header(worksheet, row, "Project & System", 4)
    row = _write_kpi_rows(worksheet, row + 1, [
        ("PV Size (kW)", sizing.get("pv_kw"), FORMAT_AMOUNT, None),
        ("Battery Power (kW)", sizing.get("battery_kw"), FORMAT_AMOUNT, None),
        ("Battery Energy (kWh)", sizing.get("battery_kwh"), FORMAT_AMOUNT, None),
        ("Total Investment (USD)", _lookup(summary, "total_capex_usd"), FORMAT_AMOUNT, None),
    ])

    row += 1
    _write_section_header(worksheet, row, "Contract Terms", 4)
    terms = []
    if dppa_config is not None:
        volume = dppa_config.get("cfd_contract_volume_kwh_per_hour")
        annual_volume = (
            sum(volume) if isinstance(volume, list)
            else (volume or 0.0) * 8760
        )
        terms.extend([
            ("CfD Strike Price (VND/kWh)", dppa_config.get("cfd_strike_per_kwh_vnd"), FORMAT_AMOUNT, None),
            ("CfD Strike Escalation (per year)", dppa_config.get("cfd_strike_escalation_rate"), FORMAT_PERCENT, None),
            ("Annual Contract Volume (kWh)", annual_volume, FORMAT_AMOUNT, None),
            ("Transmission Loss Factor k", dppa_config.get("transmission_loss_factor_k"), FORMAT_RATIO, None),
            ("Distribution Loss Factor K_pp", dppa_config.get("distribution_loss_factor_kpp"), "0.000000", None),
            ("DPPA Service Fee (VND/kWh)", dppa_config.get("c_dppa_service_fee_vnd_per_kwh"), FORMAT_AMOUNT, None),
            ("Settlement Adder C_CL (VND/kWh)", dppa_config.get("c_cl_settlement_adder_vnd_per_kwh"), FORMAT_AMOUNT, None),
        ])
    else:
        terms.extend([
            ("ESCO Energy Price (fraction of EVN tariff)", assumptions.get("esco_energy_discount_fraction"), FORMAT_PERCENT, None),
            ("Demand Savings Share to ESCO", assumptions.get("demand_savings_esco_share"), FORMAT_PERCENT, None),
        ])
    terms.append(("Analysis Period (Years)", len(annual_rows), None, None))
    row = _write_kpi_rows(worksheet, row + 1, terms)

    row += 1
    _write_section_header(worksheet, row, "Developer (Seller) Returns", 4)
    row = _write_kpi_rows(worksheet, row + 1, [
        ("Equity IRR", summary.get("equity_irr_fraction"), FORMAT_PERCENT, None),
        ("Project IRR", summary.get("project_irr_fraction"), FORMAT_PERCENT, None),
        ("NPV (USD)", _lookup(summary, "npv_usd"), FORMAT_AMOUNT, "At owner discount rate"),
        ("Minimum DSCR (debt years)", minimum_dscr, FORMAT_RATIO, "Lender covenant view"),
        ("Average DSCR", summary.get("average_dscr"), FORMAT_RATIO, None),
        ("Simple Payback (Years)", summary.get("simple_payback_years"), FORMAT_YEARS, None),
        ("Equity Investment (USD)", _lookup(summary, "equity_investment_usd"), FORMAT_AMOUNT, None),
        ("Debt Principal (USD)", _lookup(summary, "debt_principal_usd"), FORMAT_AMOUNT, None),
    ])

    row += 1
    _write_section_header(worksheet, row, "Buyer (Offtaker) Outcome", 4)
    row = _write_kpi_rows(worksheet, row + 1, [
        ("Year 1 BAU Cost (USD)", _lookup(year_one, "bau_evn_bill_usd"), FORMAT_AMOUNT, None),
        ("Year 1 Cost With Project (USD)", _lookup(year_one, "offtaker_post_project_cost_usd"), FORMAT_AMOUNT, None),
        ("Year 1 Buyer Savings (USD)", _lookup(year_one, "offtaker_savings_usd"), FORMAT_AMOUNT, None),
        ("Year 1 Savings (% of BAU)", year_one.get("offtaker_savings_fraction"), FORMAT_PERCENT, None),
        ("First 10 Years Buyer Savings (USD)", _lookup(summary, "buyer_savings_10yr_usd"), FORMAT_AMOUNT, "Cumulative, years 1-10"),
        ("First 10 Years Savings (% of BAU)", summary.get("buyer_savings_10yr_fraction"), FORMAT_PERCENT, None),
        ("Lifetime Buyer Savings (USD)", lifetime_savings, FORMAT_AMOUNT, "Sum over analysis period"),
        (
            "Lifetime Savings (% of BAU)",
            (lifetime_savings / lifetime_bau) if lifetime_bau else None,
            FORMAT_PERCENT,
            None,
        ),
    ])

    row += 1
    _write_section_header(worksheet, row, "Model Basis & Conventions", 4)
    notes = [
        "Buyer settlement quantity Q_Khc = min(hourly load, loss-adjusted generation Q_adj); "
        "excess generation is sold by the generator at FMP and never billed to the buyer."
        if dppa_config else
        "ESCO is paid a discount to the time-specific EVN tariff for project-served energy.",
        "PV degradation, O&M escalation and battery replacement (REopt schedule) are applied across the analysis period.",
        (
            "PV straight-line depreciation over "
            f"{assumptions.get('pv_depreciation_years', 20)} years "
            "(Circular 45/2013/TT-BTC permits 7-20 years for generating equipment)."
        ),
        "Vietnam CIT: 4-year exemption + 9-year 50% reduction from first profitable year; "
        "5-year tax-loss carryforward.",
        "All USD figures at the fixed contract exchange rate (see Assumptions); FX drift "
        "between VND revenue and USD reporting is quantified on the FX Sensitivity sheet.",
    ]
    for note in notes:
        row += 1
        worksheet.cell(row=row, column=2, value="•  " + note).font = NOTE_FONT

    worksheet.column_dimensions["A"].width = 2
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 36


def _write_buyer_analysis(worksheet, cash_flow_result, dppa_config):
    annual_rows = cash_flow_result.get("annual_cash_flows", []) or [{}]
    year_one = annual_rows[0]
    summary = cash_flow_result.get("summary", {})

    _write_title(
        worksheet,
        "Buyer Analysis — Cost vs Business-As-Usual",
        "Annual electricity cost with the project compared against the EVN-only baseline (USD)",
        last_column=6,
    )

    row = 4
    _write_section_header(worksheet, row, "Year 1 Snapshot", 6)
    snapshot = [
        ("Year 1 BAU Cost (USD)", _lookup(year_one, "bau_evn_bill_usd"), FORMAT_AMOUNT, None),
        ("Year 1 Cost With Project (USD)", _lookup(year_one, "offtaker_post_project_cost_usd"), FORMAT_AMOUNT, None),
        ("Year 1 Savings (USD)", _lookup(year_one, "offtaker_savings_usd"), FORMAT_AMOUNT, None),
        ("Year 1 Savings (% of BAU)", year_one.get("offtaker_savings_fraction"), FORMAT_PERCENT, None),
    ]
    if dppa_config is not None:
        snapshot.extend([
            ("of which C_DN spot energy (USD)", _lookup(year_one, "c_dn_usd"), FORMAT_AMOUNT, None),
            ("of which C_DPPA + C_CL fees (USD)",
             (_lookup(year_one, "c_dppa_usd") or 0.0) + (_lookup(year_one, "c_cl_usd") or 0.0),
             FORMAT_AMOUNT, None),
            ("of which C_BL retail shortfall (USD)", _lookup(year_one, "c_bl_usd"), FORMAT_AMOUNT, None),
            ("of which CfD payment, net (USD)", _lookup(year_one, "cfd_net_usd"), FORMAT_AMOUNT,
             "Positive = buyer pays generator"),
        ])
    snapshot.extend([
        ("First 10 Years Savings (USD)", _lookup(summary, "buyer_savings_10yr_usd"), FORMAT_AMOUNT,
         "Cumulative, years 1-10"),
        ("First 10 Years Savings (% of BAU)", summary.get("buyer_savings_10yr_fraction"), FORMAT_PERCENT, None),
        ("Lifetime Savings (USD)", _lookup(summary, "buyer_savings_lifetime_usd"), FORMAT_AMOUNT,
         "Cumulative over analysis period"),
        ("Lifetime Savings (% of BAU)", summary.get("buyer_savings_lifetime_fraction"), FORMAT_PERCENT, None),
    ])
    row = _write_kpi_rows(worksheet, row + 1, snapshot)

    row += 1
    header_row = row
    headers = [label for label, _key, _fmt in BUYER_ANNUAL_COLUMNS] + ["Cumulative Savings (USD)"]
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    cumulative = 0.0
    for offset, annual in enumerate(annual_rows, start=1):
        data_row = header_row + offset
        for column_index, (_label, key, number_format) in enumerate(BUYER_ANNUAL_COLUMNS, start=1):
            cell = worksheet.cell(row=data_row, column=column_index, value=_lookup(annual, key))
            if number_format:
                cell.number_format = number_format
        cumulative += _lookup(annual, "offtaker_savings_usd") or 0.0
        cumulative_cell = worksheet.cell(
            row=data_row, column=len(headers), value=cumulative
        )
        cumulative_cell.number_format = FORMAT_AMOUNT

    if len(annual_rows) > 1:
        chart = LineChart()
        chart.title = "Cumulative Buyer Savings (USD)"
        data = Reference(
            worksheet,
            min_col=len(headers),
            min_row=header_row,
            max_row=header_row + len(annual_rows),
        )
        categories = Reference(
            worksheet,
            min_col=1,
            min_row=header_row + 1,
            max_row=header_row + len(annual_rows),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 18
        worksheet.add_chart(chart, "H4")

    worksheet.column_dimensions["A"].width = 10
    for letter in ("B", "C", "D", "E", "F"):
        worksheet.column_dimensions[letter].width = 22


def _write_developer_returns(worksheet, cash_flow_result):
    summary = cash_flow_result.get("summary", {})
    annual_rows = cash_flow_result.get("annual_cash_flows", []) or [{}]
    minimum_dscr = _minimum_debt_service_dscr(cash_flow_result)

    _write_title(
        worksheet,
        "Developer Returns — Financing & Cash Flow",
        "Seller/ESCO investment case: sources & uses, return metrics and annual equity cash flow (USD)",
        last_column=8,
    )

    row = 4
    _write_section_header(worksheet, row, "Sources & Uses", 8)
    row = _write_kpi_rows(worksheet, row + 1, [
        ("Total Investment (USD)", _lookup(summary, "total_capex_usd"), FORMAT_AMOUNT, None),
        ("Debt Principal (USD)", _lookup(summary, "debt_principal_usd"), FORMAT_AMOUNT, None),
        ("Equity Investment (USD)", _lookup(summary, "equity_investment_usd"), FORMAT_AMOUNT, None),
    ])

    row += 1
    _write_section_header(worksheet, row, "Return Metrics", 8)
    row = _write_kpi_rows(worksheet, row + 1, [
        ("Equity IRR", summary.get("equity_irr_fraction"), FORMAT_PERCENT, None),
        ("Project IRR", summary.get("project_irr_fraction"), FORMAT_PERCENT, None),
        ("NPV (USD)", _lookup(summary, "npv_usd"), FORMAT_AMOUNT, "At owner discount rate"),
        ("Minimum DSCR (debt years)", minimum_dscr, FORMAT_RATIO, None),
        ("Average DSCR", summary.get("average_dscr"), FORMAT_RATIO, None),
        ("Simple Payback (Years)", summary.get("simple_payback_years"), FORMAT_YEARS, None),
        ("ROI (cumulative equity CF / equity)", summary.get("roi_fraction"), FORMAT_PERCENT, None),
    ])

    row += 1
    header_row = row
    headers = [label for label, _key, _fmt in DEVELOPER_ANNUAL_COLUMNS] + ["Cumulative Equity CF (USD)"]
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    cumulative = -(_lookup(summary, "equity_investment_usd") or 0.0)
    for offset, annual in enumerate(annual_rows, start=1):
        data_row = header_row + offset
        for column_index, (_label, key, number_format) in enumerate(DEVELOPER_ANNUAL_COLUMNS, start=1):
            if key is None:
                value = (
                    (_lookup(annual, "annual_om_usd") or 0.0)
                    + (_lookup(annual, "replacement_cost_usd") or 0.0)
                )
            else:
                value = _lookup(annual, key)
            cell = worksheet.cell(row=data_row, column=column_index, value=value)
            if number_format:
                cell.number_format = number_format
        cumulative += _lookup(annual, "equity_cash_flow_usd") or 0.0
        cumulative_cell = worksheet.cell(
            row=data_row, column=len(headers), value=cumulative
        )
        cumulative_cell.number_format = FORMAT_AMOUNT

    if len(annual_rows) > 1:
        chart = LineChart()
        chart.title = "Cumulative Equity Cash Flow (USD)"
        data = Reference(
            worksheet,
            min_col=len(headers),
            min_row=header_row,
            max_row=header_row + len(annual_rows),
        )
        categories = Reference(
            worksheet,
            min_col=1,
            min_row=header_row + 1,
            max_row=header_row + len(annual_rows),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 18
        worksheet.add_chart(chart, "K4")

    worksheet.column_dimensions["A"].width = 10
    for letter in ("B", "C", "D", "E", "F", "G", "H", "I"):
        worksheet.column_dimensions[letter].width = 20


def _write_bau_vs_dppa_sheet(worksheet, cash_flow_result, report_data, assumptions):
    annual_rows = cash_flow_result.get("annual_cash_flows", []) or [{}]
    year_one = annual_rows[0]
    hourly = report_data.get("dppa_hourly_breakout", []) or []
    annual_production = report_data.get("annual_production") or {}

    def g(key):
        return year_one.get(key, 0.0) or 0.0

    bau_evn_bill_usd = g("bau_evn_bill_usd")
    c_dn_usd = g("c_dn_usd")
    c_dppa_usd = g("c_dppa_usd")
    c_cl_usd = g("c_cl_usd")
    c_bl_usd = g("c_bl_usd")
    cfd_net_usd = g("cfd_net_usd")
    esco_demand_revenue_usd = g("esco_demand_revenue_usd")
    esco_grid_arb_usd = g("esco_grid_arbitrage_revenue_usd")
    gen_fmp_usd = g("generator_fmp_revenue_usd")
    offtaker_post_cost_usd = g("offtaker_post_project_cost_usd")
    optimized_demand_charge_year_usd = g("optimized_demand_charge_usd")

    buyer_savings_usd = bau_evn_bill_usd - offtaker_post_cost_usd
    buyer_savings_fraction = (
        buyer_savings_usd / bau_evn_bill_usd if bau_evn_bill_usd else None
    )

    total_seller_revenue_usd = gen_fmp_usd + cfd_net_usd + esco_demand_revenue_usd + esco_grid_arb_usd

    # Energy flows from annual_production (if present) else sum from hourly.
    def _sum_hourly(key):
        return sum(row.get(key, 0.0) or 0.0 for row in hourly)

    pv_to_load_kwh = annual_production.get("pv_to_load_kwh") or 0.0
    pv_to_storage_kwh = annual_production.get("pv_to_storage_kwh") or 0.0
    storage_to_load_kwh = annual_production.get("storage_to_load_kwh") or 0.0
    pv_to_grid_effective_kwh = annual_production.get("pv_to_grid_effective_kwh") or 0.0
    # Quantity conversion generator meter -> customer side uses K_pp only
    # (k is price-only, CD7 Ví dụ 1); settlement always emits q_adj_kw, the
    # fallback recomputes it from the configured K_pp for older breakouts.
    kpp = (assumptions.get("dppa") or {}).get("distribution_loss_factor_kpp") or 1.0
    q_re_meter_kwh = _sum_hourly("q_re_meter_kw")
    q_adj_kwh = sum(
        (row.get("q_adj_kw") if row.get("q_adj_kw") is not None
         else (row.get("q_re_meter_kw", 0.0) or 0.0) / kpp)
        for row in hourly
    )
    q_khc_kwh = sum(
        (row.get("q_khc_kw") if row.get("q_khc_kw") is not None
         else min(row.get("load_kw", 0.0) or 0.0, row.get("q_adj_kw", 0.0) or 0.0))
        for row in hourly
    )
    shortfall_kwh = sum(
        max(
            (row.get("load_kw", 0.0) or 0.0)
            - (row.get("q_khc_kw") if row.get("q_khc_kw") is not None else (row.get("q_adj_kw", 0.0) or 0.0)),
            0.0,
        )
        for row in hourly
    )

    values = {
        "bau_evn_bill_usd": bau_evn_bill_usd,
        "c_dn_usd": c_dn_usd,
        "c_dppa_usd": c_dppa_usd,
        "c_cl_usd": c_cl_usd,
        "c_bl_usd": c_bl_usd,
        "cfd_net_usd": cfd_net_usd,
        "optimized_demand_charge_year_usd": optimized_demand_charge_year_usd,
        "esco_demand_revenue_usd": esco_demand_revenue_usd,
        "total_buyer_outflow_usd": offtaker_post_cost_usd,
        "buyer_savings_vs_bau_usd": buyer_savings_usd,
        "buyer_savings_fraction": buyer_savings_fraction,
        "esco_phase2_energy_revenue_usd": 0.0,  # zeroed under DPPA per design doc
        "generator_fmp_revenue_usd": gen_fmp_usd,
        "esco_grid_arbitrage_revenue_usd": esco_grid_arb_usd,
        "total_seller_revenue_usd": total_seller_revenue_usd,
        "pv_to_load_kwh": pv_to_load_kwh,
        "pv_to_storage_kwh": pv_to_storage_kwh,
        "pv_to_grid_effective_kwh": pv_to_grid_effective_kwh,
        "storage_to_load_kwh": storage_to_load_kwh,
        "q_re_meter_kwh": q_re_meter_kwh,
        "q_adj_kwh": q_adj_kwh,
        "q_khc_kwh": q_khc_kwh,
        "shortfall_kwh": shortfall_kwh,
    }

    worksheet.cell(row=1, column=1, value="Line Item").font = HEADER_FONT
    worksheet.cell(row=1, column=2, value="Year 1 (USD)").font = HEADER_FONT
    worksheet.cell(row=1, column=3, value="Side").font = HEADER_FONT

    row_index = 2
    for label, key, side in BAU_VS_DPPA_ROWS:
        if label == "section":
            worksheet.cell(row=row_index, column=1, value=key).font = HEADER_FONT
            worksheet.cell(row=row_index, column=1).fill = HEADER_FILL
            row_index += 1
            continue
        worksheet.cell(row=row_index, column=1, value=label)
        worksheet.cell(row=row_index, column=2, value=values.get(key))
        worksheet.cell(row=row_index, column=3, value=side)
        row_index += 1

    worksheet.column_dimensions["A"].width = 48
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 14


def _active_dppa_config(assumptions):
    # Only the grid-CfD settlement drives the CfD-specific sheets (Year 1 BAU vs
    # DPPA, Monthly/Hourly Settlement) and the CfD summary rows. The physical
    # private wire has no settlement chain — the standard ESCO-style Buyer
    # Analysis / Executive Summary cover it — so it is not an "active" DPPA here.
    dppa = assumptions.get("dppa") if assumptions else None
    if not dppa or dppa.get("type", "none") != DPPA_TYPE_GRID_CFD:
        return None
    return dppa


def _number_format_for(key):
    return schema.number_format(key)


def _write_technical_results(worksheet, report_data):
    """System sizing, year-1 energy balance and bill comparison in one sheet."""
    sections = [
        ("System Sizing", SYSTEM_SIZING_ROWS, report_data.get("system_sizing", {})),
        ("Annual Energy Balance (Year 1)", ANNUAL_PRODUCTION_ROWS,
         report_data.get("annual_production", {})),
        ("Year-1 Utility Bill Comparison", RESULTS_COMPARISON_ROWS,
         report_data.get("results_comparison", {})),
    ]

    worksheet.cell(row=1, column=1, value="Metric").font = HEADER_FONT
    worksheet.cell(row=1, column=2, value="Value").font = HEADER_FONT
    worksheet.cell(row=1, column=1).fill = HEADER_FILL
    worksheet.cell(row=1, column=2).fill = HEADER_FILL

    row_index = 2
    production_bounds = None
    for title, rows, values in sections:
        for column in (1, 2):
            worksheet.cell(row=row_index, column=column).fill = SECTION_FILL
        worksheet.cell(row=row_index, column=1, value=title).font = SECTION_FONT
        row_index += 1
        first_data_row = row_index
        for label, key in rows:
            worksheet.cell(row=row_index, column=1, value=label)
            value_cell = worksheet.cell(row=row_index, column=2, value=_lookup(values, key))
            number_format = _number_format_for(key)
            if number_format:
                value_cell.number_format = number_format
            row_index += 1
        if title.startswith("Annual Energy Balance"):
            production_bounds = (first_data_row, row_index - 1)
        row_index += 1

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 18

    if production_bounds:
        first_row, last_row = production_bounds
        chart = BarChart()
        chart.title = "Annual Electricity Production Breakdown (kWh)"
        data = Reference(worksheet, min_col=2, min_row=first_row, max_row=last_row)
        categories = Reference(worksheet, min_col=1, min_row=first_row, max_row=last_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        worksheet.add_chart(chart, "D2")


def _write_dispatch_sheet(worksheet, rows):
    """Hourly dispatch with the original PV generation split and a peak-week chart."""
    for column_index, (header, _key) in enumerate(DISPATCH_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_index, row in enumerate(rows, start=2):
        for column_index, (_header, key) in enumerate(DISPATCH_COLUMNS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=_lookup(row, key))
            if key == "pv_production_factor":
                cell.number_format = "0.000"
            elif key != "hour":
                cell.number_format = FORMAT_AMOUNT

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["C"].width = 22

    if not rows:
        return
    # Chart one representative week only — the week containing the annual peak
    # load — instead of all 8760 hours.
    peak_index = max(range(len(rows)), key=lambda i: rows[i].get("load_kw") or 0)
    week_start = (peak_index // HOURS_PER_WEEK) * HOURS_PER_WEEK
    week_end = min(week_start + HOURS_PER_WEEK, len(rows))
    first_row = 2 + week_start
    last_row = 1 + week_end

    chart = LineChart()
    chart.title = (
        f"Dispatch — Peak-Load Week (hours {week_start + 1}-{week_end})"
    )
    for series_title, column in DISPATCH_CHART_SERIES:
        values = Reference(
            worksheet, min_col=column, min_row=first_row, max_row=last_row
        )
        series = Series(values, title=series_title)
        chart.series.append(series)
    categories = Reference(
        worksheet, min_col=1, min_row=first_row, max_row=last_row
    )
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 24
    worksheet.add_chart(chart, "M2")


def _write_table_sheet(worksheet, columns, rows, chart_title=None):
    for column_index, (header, _key) in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    column_formats = [_number_format_for(key) for _header, key in columns]
    for row_index, row in enumerate(rows, start=2):
        for column_index, (_header, key) in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=_lookup(row, key))
            number_format = column_formats[column_index - 1]
            if number_format:
                cell.number_format = number_format

    worksheet.freeze_panes = "A2"
    if chart_title and rows:
        _add_line_chart(worksheet, chart_title, len(columns), len(rows) + 1)


def _autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        lengths = [
            len(str(cell.value))
            for cell in column_cells
            if cell.value is not None
        ]
        if not lengths:
            continue
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(lengths) + 2, 40)


def _lookup(values, key):
    if key in values:
        return values.get(key)
    if key.endswith("_usd"):
        return values.get(f"{key[:-4]}_vnd")
    return values.get(key)


def _add_line_chart(worksheet, title, max_col, last_row):
    chart = LineChart()
    chart.title = title
    data = Reference(worksheet, min_col=2, max_col=max_col, min_row=1, max_row=last_row)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 16
    worksheet.add_chart(chart, "H2")
