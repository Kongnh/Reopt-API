from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill


SCENARIO_COLUMNS = [
    ("Scenario", "scenario_id"),
    ("Strike (VND/kWh)", "strike_vnd_per_kwh"),
    ("Contract Volume", "volume_fraction"),
    ("Annual Contracted Energy (kWh)", "annual_contracted_kwh"),
    ("Buyer Year 1 Savings vs BAU (USD)", "factory_savings_vs_bau_usd"),
    ("Buyer Year 1 Savings vs BAU", "factory_savings_vs_bau_fraction"),
    ("Buyer 10yr Savings vs BAU (USD)", "buyer_savings_10yr_usd"),
    ("Buyer 10yr Savings vs BAU", "buyer_savings_10yr_fraction"),
    ("Buyer Lifetime Savings vs BAU (USD)", "buyer_savings_lifetime_usd"),
    ("Buyer Lifetime Savings vs BAU", "buyer_savings_lifetime_fraction"),
    ("Buyer Savings vs non-DPPA case_2 (USD)", "factory_savings_vs_case_2_usd"),
    ("Buyer Savings vs non-DPPA case_2", "factory_savings_vs_case_2_fraction"),
    ("Buyer Year 1 Outflow (USD)", "factory_year_one_outflow_usd"),
    ("Seller Equity IRR", "equity_irr_fraction"),
    ("Seller NPV (USD)", "npv_usd"),
    ("Lender Minimum DSCR", "minimum_dscr"),
    ("Lender Average DSCR", "average_dscr"),
    ("Seller Year 1 Revenue (USD)", "generator_year_one_revenue_usd"),
    ("CfD Net Transfer (USD)", "cfd_year_one_net_usd"),
    ("CfD Direction", "cfd_transfer_direction"),
    ("Balanced Deal", "balanced_deal"),
    ("Failed Gates", "failed_qualification_reasons"),
    ("Negotiation Frontier", "pareto_frontier"),
]

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def build_dppa_negotiation_workbook(study):
    workbook = Workbook()
    executive = workbook.active
    executive.title = "Executive Summary"
    _write_executive_summary(executive, study)
    _write_scenarios(workbook.create_sheet("Scenario Matrix"), study["scenarios"])
    _write_scenarios(workbook.create_sheet("Balanced Deals"), study["balanced_deals"])
    _write_scenarios(
        workbook.create_sheet("Negotiation Frontier"),
        study["negotiation_frontier"],
    )
    _write_matrix(workbook.create_sheet("Strike-Volume Matrix"), study, "balanced_deal")
    _write_matrix(
        workbook.create_sheet("Buyer 10yr Savings Heatmap"),
        study,
        "buyer_savings_10yr_fraction",
        heatmap=True,
    )
    _write_matrix(
        workbook.create_sheet("Lifetime Savings Heatmap"),
        study,
        "buyer_savings_lifetime_fraction",
        heatmap=True,
    )
    _write_matrix(
        workbook.create_sheet("ESCO IRR Heatmap"),
        study,
        "equity_irr_fraction",
        heatmap=True,
    )
    _write_matrix(
        workbook.create_sheet("Minimum DSCR Heatmap"),
        study,
        "minimum_dscr",
        heatmap=True,
    )
    _write_methodology(workbook.create_sheet("Methodology"), study)
    return workbook


def build_dppa_negotiation_summary(study):
    balanced = study["balanced_deals"]
    frontier = study["negotiation_frontier"]
    recommended = study["recommended_negotiation_range"]
    buyer_best = max(
        study["scenarios"],
        key=lambda row: row["buyer_savings_lifetime_fraction"],
    )
    lender_best = max(study["scenarios"], key=lambda row: row["minimum_dscr"])
    balanced_statement = (
        f"The balanced-deal range is {_range_text(balanced)}."
        if balanced
        else "No scenario passes all buyer, seller, and lender gates."
    )
    return "\n".join([
        "# Factory A DPPA Negotiation Summary",
        "",
        f"- Balanced deals: {len(balanced)} of {len(study['scenarios'])} scenarios.",
        f"- {balanced_statement}",
        f"- Negotiation frontier: {_range_text(frontier)}.",
        f"- Recommended starting range: {_range_text(recommended)}.",
        (
            f"- Buyer-best term: {buyer_best['scenario_id']} "
            f"{_buyer_comparison_text(buyer_best)}."
        ),
        (
            f"- Lender-best term: {lender_best['scenario_id']} reaches "
            f"{lender_best['minimum_dscr']:.2f}x minimum DSCR and "
            f"{lender_best['equity_irr_fraction']:.1%} seller equity IRR, and "
            f"{_buyer_bau_text(lender_best)}."
        ),
        "- Buyer qualification is measured against EVN-only BAU over the first 10 years AND the full project lifetime (Year 1 is reported as context).",
        "- The separate comparison against non-DPPA case_2 is disclosed and is not a qualification gate.",
        "- Higher strike generally improves seller equity IRR and lender DSCR while reducing buyer savings.",
        "- Fixed technical sizing, dispatch, market-price series, financing, and settlement formulas; no REopt rerun.",
    ]) + "\n"


def _write_executive_summary(sheet, study):
    thresholds = study["configuration"]["thresholds"]
    buyer_best = max(
        study["scenarios"],
        key=lambda row: row["buyer_savings_lifetime_fraction"],
    )
    lender_best = max(study["scenarios"], key=lambda row: row["minimum_dscr"])
    rows = [
        ("Factory A DPPA Commercial Negotiation Sweep", None),
        ("Scenarios evaluated", len(study["scenarios"])),
        ("Balanced deals", len(study["balanced_deals"])),
        ("Negotiation frontier terms", len(study["negotiation_frontier"])),
        ("Buyer 10yr savings vs BAU threshold", thresholds["buyer_savings_10yr_fraction"]),
        ("Buyer lifetime savings vs BAU threshold", thresholds["buyer_savings_lifetime_fraction"]),
        ("Seller equity IRR threshold", thresholds["equity_irr_fraction"]),
        ("Lender minimum DSCR threshold", thresholds["minimum_dscr"]),
        ("Recommended starting range", _range_text(study["recommended_negotiation_range"])),
        ("Buyer-best term", buyer_best["scenario_id"]),
        ("Buyer-best 10yr savings vs BAU", buyer_best["buyer_savings_10yr_fraction"]),
        ("Buyer-best lifetime savings vs BAU", buyer_best["buyer_savings_lifetime_fraction"]),
        ("Lender-best term", lender_best["scenario_id"]),
        ("Lender-best minimum DSCR", lender_best["minimum_dscr"]),
        ("Disclosure", "Passing BAU does not mean beating non-DPPA case_2."),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 85


def _write_scenarios(sheet, rows):
    for column, (label, _) in enumerate(SCENARIO_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row_index, row in enumerate(rows, start=2):
        for column, (_, key) in enumerate(SCENARIO_COLUMNS, start=1):
            value = row.get(key)
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
            sheet.cell(row=row_index, column=column, value=value)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, len(SCENARIO_COLUMNS) + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = 20


def _write_matrix(sheet, study, key, heatmap=False):
    strikes = study["configuration"]["strikes_vnd_per_kwh"]
    volumes = study["configuration"]["volume_fractions"]
    by_term = {
        (row["strike_vnd_per_kwh"], row["volume_fraction"]): row
        for row in study["scenarios"]
    }
    sheet.cell(row=1, column=1, value="Strike / Contract Volume")
    for column, volume in enumerate(volumes, start=2):
        sheet.cell(row=1, column=column, value=volume)
    for row_index, strike in enumerate(strikes, start=2):
        sheet.cell(row=row_index, column=1, value=strike)
        for column, volume in enumerate(volumes, start=2):
            value = by_term[(float(strike), float(volume))][key]
            sheet.cell(row=row_index, column=column, value=value)
    if heatmap and strikes and volumes:
        end = sheet.cell(row=len(strikes) + 1, column=len(volumes) + 1).coordinate
        sheet.conditional_formatting.add(
            f"B2:{end}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )
    sheet.freeze_panes = "B2"


def _write_methodology(sheet, study):
    rows = [
        ("Fixed technical baseline", str(study.get("baseline_sources", {}))),
        ("Scenario grid", "CfD strike 1400-2200 VND/kWh x contract volume 70%-100%."),
        ("Buyer gate", "Cumulative savings versus EVN-only BAU >= 0% over both the first 10 years and the full project lifetime."),
        ("Seller gate", "Equity IRR >= 12%."),
        ("Lender gate", "Minimum DSCR in debt-service years >= 1.20x."),
        ("Frontier", "Non-dominated balanced terms on buyer lifetime savings vs BAU and seller equity IRR."),
        ("Important disclosure", "non-DPPA case_2 comparison is reported but is not a qualification gate."),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=value)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 100


def _range_text(rows):
    if not rows:
        return "none"
    strikes = sorted({int(row["strike_vnd_per_kwh"]) for row in rows})
    volumes = sorted({int(round(row["volume_fraction"] * 100)) for row in rows})
    return (
        f"{strikes[0]}-{strikes[-1]} VND/kWh at "
        f"{volumes[0]}%-{volumes[-1]}% contract volume"
    )


def _buyer_comparison_text(row):
    return (
        f"{_buyer_bau_text(row)} and "
        f"{_buyer_case_2_text(row)}"
    )


def _buyer_bau_text(row):
    savings_10yr = row["buyer_savings_10yr_fraction"]
    savings_lifetime = row["buyer_savings_lifetime_fraction"]
    direction_10yr = "saves" if savings_10yr >= 0 else "costs"
    direction_lifetime = "saves" if savings_lifetime >= 0 else "costs"
    return (
        f"{direction_10yr} the buyer {abs(savings_10yr):.1%} over the first 10 years "
        f"and {direction_lifetime} {abs(savings_lifetime):.1%} over the project lifetime versus BAU"
    )


def _buyer_case_2_text(row):
    savings = row["factory_savings_vs_case_2_fraction"]
    if savings >= 0:
        return f"saves the buyer {savings:.1%} versus non-DPPA case_2"
    return f"increases buyer Year 1 cost {abs(savings):.1%} versus non-DPPA case_2"
