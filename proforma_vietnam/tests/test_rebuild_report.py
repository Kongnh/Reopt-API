import json
import tempfile
from pathlib import Path
from unittest import TestCase

from openpyxl import load_workbook

from proforma_vietnam.rebuild_report import rebuild_report


class RebuildReportTests(TestCase):

    def test_rebuilds_workbook_from_saved_results_and_assumptions(self):
        case_dir = Path(tempfile.mkdtemp())
        results = {
            "run_uuid": "test-uuid-123",
            "inputs": {
                "ElectricTariff": {"tou_energy_rates_per_kwh": [1000, 2000]},
                "ElectricStorage": {"can_grid_charge": False},
                "Financial": {"owner_discount_rate_fraction": 0.10},
            },
            "outputs": {
                "PV": {
                    "size_kw": 100,
                    "installed_cost_per_kw": 1000,
                    "electric_to_load_series_kw": [1, 2],
                },
                "ElectricStorage": {
                    "initial_capital_cost": 10000,
                    "storage_to_load_series_kw": [3, 4],
                },
                "ElectricTariff": {
                    "year_one_bill_before_tax_bau": 50000,
                    "year_one_bill_before_tax": 30000,
                    "year_one_demand_cost_before_tax_bau": 8000,
                    "year_one_demand_cost_before_tax": 3000,
                },
                "Financial": {"year_one_om_costs_before_tax": 1000},
            },
        }
        assumptions = {
            "case_name": "Rebuild Test Case",
            "esco_energy_discount_fraction": 0.9,
            "debt_fraction": 0.7,
        }
        (case_dir / "results.json").write_text(json.dumps(results), encoding="utf-8")
        (case_dir / "assumptions.json").write_text(json.dumps(assumptions), encoding="utf-8")

        out_path = rebuild_report(case_dir)

        self.assertEqual(out_path.name, "vietnam_report_test-uuid-123.xlsx")
        workbook = load_workbook(out_path)
        self.assertIn("Executive Summary", workbook.sheetnames)
        self.assertIn("Cash Flow", workbook.sheetnames)
