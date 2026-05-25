from unittest import TestCase

from openpyxl import Workbook

from proforma_vietnam.reference_workbook_comparison import (
    compare_reference_workbook,
)


class ReferenceEscoWorkbookTests(TestCase):

    def test_reference_esco_project_matches_hand_built_workbook_within_one_percent(self):
        discrepancies = compare_reference_workbook(_reference_workbook())

        self.assertEqual(discrepancies, [])


def _reference_workbook():
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs.append(["Input", "Value"])
    for key, value in REFERENCE_INPUTS.items():
        inputs.append([key, value])

    summary = workbook.create_sheet("Reference Summary")
    summary.append(["Metric", "Reference Value"])
    for key, value in REFERENCE_SUMMARY.items():
        summary.append([key, value])

    annual = workbook.create_sheet("Reference Annual Cash Flow")
    annual_headers = list(REFERENCE_ANNUAL_ROWS[0].keys())
    annual.append(annual_headers)
    for row in REFERENCE_ANNUAL_ROWS:
        annual.append([row[header] for header in annual_headers])

    return workbook


REFERENCE_INPUTS = {
    "project_served_pv_kwh": "100000,120000,80000",
    "evn_energy_rates_vnd_per_kwh": "1800,2500,1100",
    "bau_evn_bill_vnd": 900000000,
    "optimized_evn_bill_vnd": 630000000,
    "bau_demand_charge_vnd": 180000000,
    "optimized_demand_charge_vnd": 120000000,
    "pv_capex_vnd": 2100000000,
    "bess_capex_vnd": 900000000,
    "other_capex_vnd": 150000000,
    "annual_om_vnd": 45000000,
    "esco_energy_discount_fraction": 0.9,
    "owner_discount_rate_fraction": 0.10,
    "evn_energy_escalation_rate": 0.04,
    "evn_capacity_escalation_rate": 0.04,
    "esco_demand_savings_share": 0.80,
    "debt_fraction": 0.65,
    "debt_interest_rate_fraction": 0.085,
    "debt_term_years": 10,
    "project_years": 25,
}

REFERENCE_SUMMARY = {
    "total_capex_vnd": 3150000000,
    "debt_principal_vnd": 2047500000.0,
    "equity_investment_vnd": 1102500000.0,
    "project_irr_fraction": 0.19179371,
    "equity_irr_fraction": 0.27255769,
    "npv_vnd": 3017552407.85,
    "average_dscr": 1.92595413,
    "simple_payback_years": 4.57805,
    "roi_fraction": 14.54198349,
}

REFERENCE_ANNUAL_ROWS = [
    {
        "year": 1,
        "esco_revenue_vnd": 559200000.0,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 0.0,
        "cash_available_for_debt_service_vnd": 514200000.0,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 202145223.79,
    },
    {
        "year": 2,
        "esco_revenue_vnd": 581568000.0,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 0.0,
        "cash_available_for_debt_service_vnd": 536568000.0,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 224513223.79,
    },
    {
        "year": 3,
        "esco_revenue_vnd": 604830720.0,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 0.0,
        "cash_available_for_debt_service_vnd": 559830720.0,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 247775943.79,
    },
    {
        "year": 4,
        "esco_revenue_vnd": 629023948.8,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 0.0,
        "cash_available_for_debt_service_vnd": 584023948.8,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 271969172.59,
    },
    {
        "year": 5,
        "esco_revenue_vnd": 654184906.75,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 27090257.36,
        "cash_available_for_debt_service_vnd": 582094649.39,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 270039873.18,
    },
    {
        "year": 6,
        "esco_revenue_vnd": 680352303.02,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 31332812.75,
        "cash_available_for_debt_service_vnd": 604019490.27,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 291964714.06,
    },
    {
        "year": 7,
        "esco_revenue_vnd": 707566395.14,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 35818232.07,
        "cash_available_for_debt_service_vnd": 626748163.07,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 314693386.86,
    },
    {
        "year": 8,
        "esco_revenue_vnd": 735869050.95,
        "depreciation_vnd": 217500000.0,
        "cit_vnd": 40562448.62,
        "cash_available_for_debt_service_vnd": 650306602.33,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 338251826.12,
    },
    {
        "year": 9,
        "esco_revenue_vnd": 765303812.99,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 56832561.62,
        "cash_available_for_debt_service_vnd": 663471251.37,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 351416475.16,
    },
    {
        "year": 10,
        "esco_revenue_vnd": 795915965.51,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 62146927.8,
        "cash_available_for_debt_service_vnd": 688769037.71,
        "debt_service_vnd": 312054776.21,
        "equity_cash_flow_vnd": 376714261.5,
    },
    {
        "year": 11,
        "esco_revenue_vnd": 827752604.13,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 67775260.41,
        "cash_available_for_debt_service_vnd": 714977343.71,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 714977343.71,
    },
    {
        "year": 12,
        "esco_revenue_vnd": 860862708.29,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 71086270.83,
        "cash_available_for_debt_service_vnd": 744776437.46,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 744776437.46,
    },
    {
        "year": 13,
        "esco_revenue_vnd": 895297216.62,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 74529721.66,
        "cash_available_for_debt_service_vnd": 775767494.96,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 775767494.96,
    },
    {
        "year": 14,
        "esco_revenue_vnd": 931109105.29,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 156221821.06,
        "cash_available_for_debt_service_vnd": 729887284.23,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 729887284.23,
    },
    {
        "year": 15,
        "esco_revenue_vnd": 968353469.5,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 163670693.9,
        "cash_available_for_debt_service_vnd": 759682775.6,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 759682775.6,
    },
    {
        "year": 16,
        "esco_revenue_vnd": 1007087608.28,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 171417521.66,
        "cash_available_for_debt_service_vnd": 790670086.62,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 790670086.62,
    },
    {
        "year": 17,
        "esco_revenue_vnd": 1047371112.61,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 179474222.52,
        "cash_available_for_debt_service_vnd": 822896890.09,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 822896890.09,
    },
    {
        "year": 18,
        "esco_revenue_vnd": 1089265957.12,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 187853191.42,
        "cash_available_for_debt_service_vnd": 856412765.69,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 856412765.69,
    },
    {
        "year": 19,
        "esco_revenue_vnd": 1132836595.4,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 196567319.08,
        "cash_available_for_debt_service_vnd": 891269276.32,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 891269276.32,
    },
    {
        "year": 20,
        "esco_revenue_vnd": 1178150059.22,
        "depreciation_vnd": 105000000.0,
        "cit_vnd": 205630011.84,
        "cash_available_for_debt_service_vnd": 927520047.37,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 927520047.37,
    },
    {
        "year": 21,
        "esco_revenue_vnd": 1225276061.58,
        "depreciation_vnd": 0,
        "cit_vnd": 236055212.32,
        "cash_available_for_debt_service_vnd": 944220849.27,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 944220849.27,
    },
    {
        "year": 22,
        "esco_revenue_vnd": 1274287104.05,
        "depreciation_vnd": 0,
        "cit_vnd": 245857420.81,
        "cash_available_for_debt_service_vnd": 983429683.24,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 983429683.24,
    },
    {
        "year": 23,
        "esco_revenue_vnd": 1325258588.21,
        "depreciation_vnd": 0,
        "cit_vnd": 256051717.64,
        "cash_available_for_debt_service_vnd": 1024206870.57,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 1024206870.57,
    },
    {
        "year": 24,
        "esco_revenue_vnd": 1378268931.74,
        "depreciation_vnd": 0,
        "cit_vnd": 266653786.35,
        "cash_available_for_debt_service_vnd": 1066615145.39,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 1066615145.39,
    },
    {
        "year": 25,
        "esco_revenue_vnd": 1433399689.01,
        "depreciation_vnd": 0,
        "cit_vnd": 277679937.8,
        "cash_available_for_debt_service_vnd": 1110719751.21,
        "debt_service_vnd": 0,
        "equity_cash_flow_vnd": 1110719751.21,
    },
]
