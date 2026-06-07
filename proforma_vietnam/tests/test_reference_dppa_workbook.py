from unittest import TestCase

from openpyxl import Workbook

from proforma_vietnam.reference_workbook_comparison import (
    compare_dppa_reference_workbook,
)


class ReferenceDppaWorkbookTests(TestCase):

    def test_grid_dppa_cfd_reference_workbook_matches_within_one_percent(self):
        discrepancies = compare_dppa_reference_workbook(_reference_workbook())

        self.assertEqual(discrepancies, [])


def _reference_workbook():
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs.append(["Input", "Value"])
    for key, value in REFERENCE_INPUTS.items():
        inputs.append([key, value])

    dppa = workbook.create_sheet("DPPA Year One")
    dppa.append(["Input", "Value"])
    for key, value in REFERENCE_DPPA_YEAR_ONE.items():
        dppa.append([key, value])

    summary = workbook.create_sheet("Reference Summary")
    summary.append(["Metric", "Reference Value"])
    for key, value in REFERENCE_SUMMARY.items():
        summary.append([key, value])

    annual = workbook.create_sheet("Reference Annual Cash Flow")
    annual_headers = list(REFERENCE_ANNUAL_ROWS[0].keys())
    annual.append(annual_headers)
    for row in REFERENCE_ANNUAL_ROWS:
        annual.append([row[header] for header in annual_headers])

    return workbook


# Minimal Phase 2 inputs (used only for capex/debt/tax — energy/demand are
# overridden by the DPPA settlement primitives).
REFERENCE_INPUTS = {
    "project_served_pv_kwh": "0,0,0",
    "evn_energy_rates_vnd_per_kwh": "0,0,0",
    "bau_evn_bill_vnd": 0,
    "optimized_evn_bill_vnd": 0,
    "bau_demand_charge_vnd": 0,
    "optimized_demand_charge_vnd": 0,
    "pv_capex_vnd": 0,
    "bess_capex_vnd": 0,
    "annual_om_vnd": 0,
    "esco_energy_discount_fraction": 0.9,
    "owner_discount_rate_fraction": 0.10,
    "evn_energy_escalation_rate": 0.0,
    "evn_capacity_escalation_rate": 0.0,
    "esco_demand_savings_share": 0.80,
    "debt_fraction": 0.0,
    "debt_interest_rate_fraction": 0.0,
    "debt_term_years": 1,
    "project_years": 2,
}

# Year-one DPPA primitives: simple round numbers so the assertion math is obvious.
REFERENCE_DPPA_YEAR_ONE = {
    "c_dn_vnd": 1000.0,
    "c_dppa_vnd": 200.0,
    "c_cl_vnd": 100.0,
    "c_bl_vnd": 400.0,
    "cfd_strike_revenue_vnd": 1700.0,
    "cfd_fmp_offset_vnd": 1500.0,
    "generator_fmp_revenue_vnd": 2000.0,
    "fee_escalation_rate": 0.0,
    "cfd_strike_escalation_rate": 0.0,
}

# Generator revenue = generator_fmp + (cfd_strike - cfd_fmp) = 2000 + 200 = 2200
# ESCO revenue = generator_revenue + 0 demand + 0 grid = 2200
# DPPA offtaker cost = 1000 + 200 + 100 + 400 + 200 = 1900
# bau_evn_bill = 0; offtaker_savings = 0 - (1900 + 0 + 0) = -1900
REFERENCE_SUMMARY = {
    "total_capex_vnd": 0.0,
}

REFERENCE_ANNUAL_ROWS = [
    {
        "year": 1,
        "esco_energy_revenue_vnd": 2200.0,
        "esco_revenue_vnd": 2200.0,
        "c_dn_vnd": 1000.0,
        "c_dppa_vnd": 200.0,
        "c_cl_vnd": 100.0,
        "c_bl_vnd": 400.0,
        "cfd_net_vnd": 200.0,
        "generator_revenue_vnd": 2200.0,
        "dppa_offtaker_cost_vnd": 1900.0,
    },
    {
        "year": 2,
        "esco_energy_revenue_vnd": 2200.0,
        "esco_revenue_vnd": 2200.0,
        "c_dn_vnd": 1000.0,
        "c_dppa_vnd": 200.0,
        "c_cl_vnd": 100.0,
        "c_bl_vnd": 400.0,
        "cfd_net_vnd": 200.0,
        "generator_revenue_vnd": 2200.0,
        "dppa_offtaker_cost_vnd": 1900.0,
    },
]
