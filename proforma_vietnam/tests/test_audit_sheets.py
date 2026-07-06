from unittest import TestCase

from openpyxl import Workbook

from proforma_vietnam import audit_sheets
from proforma_vietnam.cash_flow import (
    calculate_fx_sensitivity,
    calculate_vietnam_esco_cash_flow,
)
from proforma_vietnam.dppa_settlement import settle_dppa_year_one
from proforma_vietnam.xlsx_builder import build_vietnam_esco_workbook


def _esco_result(**overrides):
    inputs = dict(
        project_served_pv_kwh=[1000.0] * 8760,
        evn_energy_rates_vnd_per_kwh=[0.08] * 8760,
        bau_evn_bill_vnd=900000,
        optimized_evn_bill_vnd=630000,
        bau_demand_charge_vnd=180000,
        optimized_demand_charge_vnd=120000,
        pv_capex_vnd=2100000,
        bess_capex_vnd=900000,
        annual_om_vnd=45000,
        esco_energy_discount_fraction=0.9,
        pv_degradation_rate=0.005,
        om_escalation_rate=0.02,
        replacement_costs_by_year=[0.0] * 10 + [250000.0],
        exchange_rate_vnd_per_usd=25000,
    )
    inputs.update(overrides)
    return calculate_vietnam_esco_cash_flow(**inputs)


def _dppa_result():
    dppa_inputs = {
        "type": "grid_dppa_cfd",
        "fmp_series_vnd_per_kwh": [1500.0] * 24,
        "cfmp_series_vnd_per_kwh": [1539.0] * 24,
        "cfd_strike_per_kwh_vnd": 1700.0,
        "cfd_strike_escalation_rate": 0.02,
        "cfd_contract_volume_kwh_per_hour": 500.0,
        "transmission_loss_factor_k": 1.026,
        "distribution_loss_factor_kpp": 1.027263,
        "allocation_fraction_delta": 1.0,
        "c_dppa_service_fee_vnd_per_kwh": 360.0,
        "c_cl_settlement_adder_vnd_per_kwh": 163.3,
        "fee_escalation_rate": 0.04,
    }
    dispatch = {
        "load_kw": [800.0] * 24,
        "pv_to_load_kw": [600.0] * 24,
        "pv_to_grid_kw": [100.0] * 24,
        "pv_curtailed_kw": [50.0] * 24,
        "storage_to_load_kw": [0.0] * 24,
        "storage_to_grid_kw": [0.0] * 24,
    }
    settlement = settle_dppa_year_one(
        dppa_inputs=dppa_inputs,
        dispatch=dispatch,
        evn_energy_rates_vnd_per_kwh=[2000.0] * 24,
    )
    # convert year-one primitives to USD as esco_pro_forma does
    year_one = {
        key: (value / 25000 if key.endswith("_vnd") else value)
        for key, value in settlement["year_one"].items()
    }
    settlement = {**settlement, "year_one": year_one}
    return calculate_vietnam_esco_cash_flow(
        project_served_pv_kwh=[600.0] * 24,
        evn_energy_rates_vnd_per_kwh=[0.08] * 24,
        bau_evn_bill_vnd=900000,
        optimized_evn_bill_vnd=630000,
        bau_demand_charge_vnd=180000,
        optimized_demand_charge_vnd=120000,
        pv_capex_vnd=2100000,
        bess_capex_vnd=900000,
        annual_om_vnd=45000,
        esco_energy_discount_fraction=0.9,
        dppa_settlement=settlement,
        exchange_rate_vnd_per_usd=25000,
    ), dppa_inputs


def _esco_surplus_result():
    return _esco_result(
        surplus_export_kwh_year1=500000.0,
        surplus_export_price_usd_per_kwh=0.04,
        surplus_price_escalation_rate=0.04,
        surplus_cap_fraction=0.5,
    )


def _physical_result():
    # ND57 Điều 25 private-wire DPPA with the surplus leg enabled, so the audit
    # sheet exercises both the live PPA revenue formula and the nested surplus
    # cells. matched kWh = 1000 kW × 8760 h to match the project-served basis.
    return _esco_result(
        physical_dppa={
            "matched_kwh_year1": 8_760_000.0,
            "ppa_price_usd_per_kwh": 0.07,
            "ppa_price_escalation_rate": 0.02,
        },
        surplus_export_kwh_year1=500000.0,
        surplus_export_price_usd_per_kwh=0.04,
        surplus_price_escalation_rate=0.04,
        surplus_cap_fraction=0.5,
    )


PHYSICAL_ASSUMPTIONS = {
    **{
        "case_name": "Audit Test",
        "exchange_rate_vnd_per_usd": 25000,
        "run_uuid": "test-uuid",
    },
    "dppa": {"type": "physical_private_wire", "ppa_price_vnd_per_kwh": 1750.0},
}


def _direct_result():
    # Factory self-invest (DIRECT_OWNERSHIP) with the surplus leg enabled, so the
    # audit sheet exercises the live bill-savings formula, the shared surplus
    # cells, and the flat-CIT (profitable-host) row. The year-11 replacement
    # (inherited from _esco_result) also drives an EBT sign flip, so the negative
    # CIT shield is covered by the Excel tie-out.
    return _esco_result(
        direct_ownership={},
        surplus_export_kwh_year1=500000.0,
        surplus_export_price_usd_per_kwh=0.04,
        surplus_price_escalation_rate=0.04,
        surplus_cap_fraction=0.5,
    )


DIRECT_ASSUMPTIONS = {
    "case_name": "Audit Test",
    "exchange_rate_vnd_per_usd": 25000,
    "run_uuid": "test-uuid",
    "direct_ownership": {"enabled": True},
}


def _construction_result():
    # Construction + grace financing on the ESCO base: 12 months of capitalized
    # IDC rolled into the COD debt balance and a 2-year principal grace, so the
    # audit sheet exercises the gated named cells, the interest-only grace rows
    # and the IDC-inclusive depreciation formulas.
    return _esco_result(construction_months=12, principal_grace_years=2)


def _usd_debt_result():
    # USD-denominated debt on the ESCO base: the base case is mechanically
    # identical to VND except the default 5% rate, so the audit sheet exercises
    # the gated DEBT_CURRENCY cell and the FX-Sensitivity DSCR-vs-depreciation
    # column (fixed USD debt service against deflating VND revenue).
    return _esco_result(debt_currency="USD")


def _dscr_sized_result():
    # DSCR-sized debt on the ESCO base: the fraction-based loan yields a minimum
    # DSCR of ~1.98x, so a 2.5x covenant BINDS — the loan is sized down until the
    # minimum DSCR equals 2.5x. Exercises the gated TARGET_MIN_DSCR / FRACTION_DEBT
    # / SUPPORTED_DEBT cells, the live DEBT_PRINCIPAL = MIN(...) rule and the
    # fixed-point tie-out row.
    return _esco_result(target_min_dscr=2.5)


def _esco_expense_result():
    # Legacy expense treatment on the ESCO base (year-11 replacement): the audit
    # workbook must keep the two-class depreciation / EBT formulas and the
    # expensing register line, and still tie out under Excel recalc.
    return _esco_result(battery_replacement_treatment="expense")


ESCO_ASSUMPTIONS = {
    "case_name": "Audit Test",
    "exchange_rate_vnd_per_usd": 25000,
    "run_uuid": "test-uuid",
}


class AssumptionsSheetTests(TestCase):

    def test_defines_named_cells_for_every_formula_input(self):
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)

        for name in (
            "PROJECT_YEARS", "FX_VND_PER_USD", "PV_CAPEX", "BESS_CAPEX",
            "TOTAL_CAPEX", "OM_YEAR1", "ESC_ENERGY", "ESC_CAPACITY",
            "PV_DEGRADATION", "DEBT_FRACTION", "DEBT_RATE", "DEBT_TERM_YEARS",
            "DEBT_PRINCIPAL", "EQUITY_INVESTMENT", "DEBT_PAYMENT", "DISC_RATE",
            "CIT_RATE", "CIT_HOLIDAY_YEARS", "PV_DEP_YEARS", "BESS_DEP_YEARS",
            "BASE_ENERGY_REV", "BASE_DEMAND_SAVINGS", "BASE_SERVED_RETAIL",
            "BAU_BILL_Y1", "OPT_BILL_Y1",
        ):
            self.assertIn(name, workbook.defined_names, name)

    def test_no_cell_holds_an_empty_string_or_accidental_formula(self):
        # Empty-string cells serialise as corrupt inlineStr XML; label/source
        # strings starting with "=" become invalid formulas. Both have broken
        # Excel loads before — guard every sheet.
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    self.assertNotEqual(
                        cell.value, "",
                        f"{worksheet.title}!{cell.coordinate} is an empty string",
                    )
                    if cell.value.startswith("="):
                        self.assertNotIn(
                            " − ", cell.value,
                            f"{worksheet.title}!{cell.coordinate} looks like prose "
                            "stored as a formula",
                        )


class ProFormaAuditSheetTests(TestCase):

    def test_white_cells_are_formulas_and_engine_rows_are_hardcoded(self):
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Pro Forma (Audit)"]

        labels = {
            sheet.cell(row=row, column=1).value: row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value
        }
        # formula rows reference named assumption cells
        revenue_row = labels["ESCO energy revenue (discount-to-EVN)"]
        year1 = sheet.cell(row=revenue_row, column=4).value
        self.assertTrue(str(year1).startswith("=BASE_ENERGY_REV"))
        # the engine tie-out row is hardcoded to the engine's numbers
        engine_row = labels["Equity cash flow (engine)"]
        engine_year1 = sheet.cell(row=engine_row, column=4).value
        self.assertAlmostEqual(
            engine_year1,
            result["annual_cash_flows"][0]["equity_cash_flow_usd"],
        )
        # year 0 equity = -equity investment
        engine_year0 = sheet.cell(row=engine_row, column=3).value
        self.assertAlmostEqual(
            engine_year0, -result["summary"]["equity_investment_usd"]
        )

    def test_check_block_carries_engine_metrics_and_pass_formulas(self):
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Pro Forma (Audit)"]

        checks = {}
        for row in range(1, sheet.max_row + 1):
            status = sheet.cell(row=row, column=6).value
            if isinstance(status, str) and status.startswith("=IF(ABS"):
                checks[sheet.cell(row=row, column=1).value] = row

        self.assertIn("Equity IRR", checks)
        self.assertIn("Equity NPV", checks)
        self.assertIn("Simple equity payback", checks)
        irr_row = checks["Equity IRR"]
        self.assertAlmostEqual(
            sheet.cell(row=irr_row, column=4).value,
            result["summary"]["equity_irr_fraction"],
        )
        self.assertTrue(
            str(sheet.cell(row=irr_row, column=3).value).startswith("=")
        )

    def test_dppa_structure_gets_settlement_rows_and_names(self):
        result, dppa_inputs = _dppa_result()
        assumptions = {**ESCO_ASSUMPTIONS, "dppa": dppa_inputs}
        workbook = build_vietnam_esco_workbook(result, assumptions=assumptions)

        for name in ("DPPA_C_DN_Y1", "DPPA_C_BL_Y1", "DPPA_CFD_STRIKE_Y1",
                     "DPPA_FMP_REV_Y1", "ESC_FEE", "ESC_STRIKE"):
            self.assertIn(name, workbook.defined_names, name)

        sheet = workbook["Pro Forma (Audit)"]
        labels = [
            sheet.cell(row=row, column=1).value
            for row in range(1, sheet.max_row + 1)
        ]
        self.assertIn("Generator FMP market revenue", labels)
        self.assertIn("CfD net settlement (to generator)", labels)
        self.assertIn("C_BL retail shortfall (incl. degradation repurchase)", labels)
        self.assertNotIn("ESCO energy revenue (discount-to-EVN)", labels)


class SurplusExportAuditTests(TestCase):

    def test_disabled_surplus_leaves_no_named_cells_or_rows(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)

        for name in ("SURPLUS_KWH_Y1", "SURPLUS_PRICE", "SURPLUS_ESC", "SURPLUS_CAP"):
            self.assertNotIn(name, workbook.defined_names, name)
        sheet = workbook["Pro Forma (Audit)"]
        labels = [
            sheet.cell(row=row, column=1).value for row in range(1, sheet.max_row + 1)
        ]
        self.assertNotIn("Surplus export revenue (Decree 243)", labels)

    def test_enabled_surplus_defines_named_cells(self):
        workbook = build_vietnam_esco_workbook(
            _esco_surplus_result(), assumptions=ESCO_ASSUMPTIONS
        )

        for name in ("SURPLUS_KWH_Y1", "SURPLUS_PRICE", "SURPLUS_ESC", "SURPLUS_CAP"):
            self.assertIn(name, workbook.defined_names, name)

    def test_enabled_surplus_adds_live_formula_revenue_row(self):
        result = _esco_surplus_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Pro Forma (Audit)"]

        labels = {
            sheet.cell(row=row, column=1).value: row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value
        }
        surplus_row = labels["Surplus export revenue (Decree 243)"]
        year1_formula = sheet.cell(row=surplus_row, column=4).value
        self.assertIn("SURPLUS_KWH_Y1", year1_formula)
        self.assertIn("SURPLUS_PRICE", year1_formula)
        self.assertIn("SURPLUS_ESC", year1_formula)

        # The total developer revenue row must fold in the surplus line so it
        # propagates to EBITDA/CFADS/tax. Year-1 formulas reference column D.
        revenue_row = labels["Total developer revenue"]
        revenue_formula = sheet.cell(row=revenue_row, column=4).value
        self.assertIn(f"D{surplus_row}", revenue_formula)


class PhysicalDppaAuditTests(TestCase):

    def test_esco_workbook_has_no_ppa_named_cells_or_row(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)

        for name in ("PPA_PRICE", "PPA_ESC", "PPA_MATCHED_KWH_Y1"):
            self.assertNotIn(name, workbook.defined_names, name)
        sheet = workbook["Pro Forma (Audit)"]
        labels = [sheet.cell(row=row, column=1).value for row in range(1, sheet.max_row + 1)]
        self.assertNotIn("PPA energy revenue (matched × price)", labels)

    def test_physical_workbook_defines_ppa_named_cells(self):
        workbook = build_vietnam_esco_workbook(
            _physical_result(), assumptions=PHYSICAL_ASSUMPTIONS
        )

        for name in ("PPA_PRICE", "PPA_ESC", "PPA_MATCHED_KWH_Y1"):
            self.assertIn(name, workbook.defined_names, name)
        # Nested surplus cells are defined by the shared 3a machinery.
        for name in ("SURPLUS_KWH_Y1", "SURPLUS_PRICE", "SURPLUS_ESC", "SURPLUS_CAP"):
            self.assertIn(name, workbook.defined_names, name)

    def test_physical_workbook_adds_live_ppa_revenue_row(self):
        workbook = build_vietnam_esco_workbook(
            _physical_result(), assumptions=PHYSICAL_ASSUMPTIONS
        )
        sheet = workbook["Pro Forma (Audit)"]

        labels = {
            sheet.cell(row=row, column=1).value: row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value
        }
        # The private wire replaces the ESCO energy line with the PPA line.
        self.assertNotIn("ESCO energy revenue (discount-to-EVN)", labels)
        ppa_row = labels["PPA energy revenue (matched × price)"]
        year1_formula = sheet.cell(row=ppa_row, column=4).value
        self.assertIn("PPA_MATCHED_KWH_Y1", year1_formula)
        self.assertIn("PPA_PRICE", year1_formula)
        self.assertIn("PPA_ESC", year1_formula)

        # Total developer revenue must fold in the PPA energy line.
        revenue_formula = sheet.cell(row=labels["Total developer revenue"], column=4).value
        self.assertIn(f"D{ppa_row}", revenue_formula)

    def test_physical_workbook_omits_cfd_settlement_sheets(self):
        workbook = build_vietnam_esco_workbook(
            _physical_result(), assumptions=PHYSICAL_ASSUMPTIONS
        )
        # The Year 1 BAU vs DPPA / Settlement sheets are grid-CfD only.
        for sheet_name in ("Year 1 BAU vs DPPA", "Monthly Settlement", "Hourly Settlement"):
            self.assertNotIn(sheet_name, workbook.sheetnames)


class DirectOwnershipAuditTests(TestCase):

    BILL_SAVINGS_LABEL = "Bill savings (avoided EVN bill: BAU − optimized)"
    FLAT_CIT_LABEL = "CIT payable (flat 20%; profitable-host shield)"

    def _labels(self, workbook):
        sheet = workbook["Pro Forma (Audit)"]
        return {
            sheet.cell(row=row, column=1).value: row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value
        }

    def test_esco_workbook_has_no_direct_ownership_rows(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)

        labels = self._labels(workbook)
        self.assertNotIn(self.BILL_SAVINGS_LABEL, labels)
        self.assertNotIn(self.FLAT_CIT_LABEL, labels)
        # The ESCO workbook keeps its discount-to-EVN energy line and the full
        # loss-carryforward CIT schedule.
        self.assertIn("ESCO energy revenue (discount-to-EVN)", labels)
        self.assertIn("CIT payable", labels)

    def test_direct_workbook_adds_live_bill_savings_row(self):
        workbook = build_vietnam_esco_workbook(
            _direct_result(), assumptions=DIRECT_ASSUMPTIONS
        )
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        # The self-invest factory replaces the ESCO energy line with the single
        # avoided-bill line, rebuilt from the reused BAU/optimized named cells.
        self.assertNotIn("ESCO energy revenue (discount-to-EVN)", labels)
        savings_row = labels[self.BILL_SAVINGS_LABEL]
        year1_formula = sheet.cell(row=savings_row, column=4).value
        self.assertIn("BAU_BILL_Y1", year1_formula)
        self.assertIn("OPT_BILL_Y1", year1_formula)
        self.assertIn("BASE_SERVED_RETAIL", year1_formula)

        # Total developer revenue folds in the bill-savings line (and surplus).
        revenue_formula = sheet.cell(row=labels["Total developer revenue"], column=4).value
        self.assertIn(f"D{savings_row}", revenue_formula)

    def test_direct_workbook_uses_flat_cit_row_and_no_carryforward(self):
        workbook = build_vietnam_esco_workbook(
            _direct_result(), assumptions=DIRECT_ASSUMPTIONS
        )
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        cit_row = labels[self.FLAT_CIT_LABEL]
        cit_formula = sheet.cell(row=cit_row, column=4).value
        # Flat 20% on EBT every year (negative EBT → negative CIT shield).
        self.assertIn("CIT_RATE", cit_formula)
        # Under the profitable-host convention the FIFO carryforward schedule and
        # the holiday-clock CIT-rate row are dropped entirely.
        self.assertNotIn("Loss vintage aged 1y — available", labels)
        self.assertNotIn("Applicable CIT rate", labels)
        self.assertNotIn("CIT payable", labels)

    def test_direct_workbook_reuses_shared_surplus_named_cells(self):
        workbook = build_vietnam_esco_workbook(
            _direct_result(), assumptions=DIRECT_ASSUMPTIONS
        )
        # The Decree 243 surplus leg rides on the shared 3a machinery; no new
        # direct-ownership named cells are minted (BAU/OPT/BASE are reused).
        for name in ("SURPLUS_KWH_Y1", "SURPLUS_PRICE", "SURPLUS_ESC", "SURPLUS_CAP"):
            self.assertIn(name, workbook.defined_names, name)
        for name in ("BAU_BILL_Y1", "OPT_BILL_Y1", "BASE_SERVED_RETAIL"):
            self.assertIn(name, workbook.defined_names, name)

    def test_direct_workbook_buyer_cost_is_residual_optimized_bill(self):
        workbook = build_vietnam_esco_workbook(
            _direct_result(), assumptions=DIRECT_ASSUMPTIONS
        )
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        # Buyer view equals developer view: residual cost is just the optimized
        # bill (no ESCO fee), so the row rebuilds from OPT_BILL_Y1.
        post_row = labels["Buyer cost with project (residual EVN bill)"]
        formula = sheet.cell(row=post_row, column=4).value
        self.assertIn("OPT_BILL_Y1", formula)

    def test_preferential_regime_with_host_disabled_has_no_shield_row(self):
        # Task 2a's carryforward machinery is already regime-aware; confirm
        # the flat profitable-host shield row is absent and the workbook
        # still builds cleanly for a host=False + re_producer DIRECT result.
        result = _esco_result(
            direct_ownership={"assume_profitable_host": False, "cit_regime": "re_producer"},
        )
        workbook = build_vietnam_esco_workbook(
            result,
            assumptions={
                **DIRECT_ASSUMPTIONS,
                "direct_ownership": {
                    "enabled": True,
                    "assume_profitable_host": False,
                    "cit_regime": "re_producer",
                },
            },
        )
        labels = self._labels(workbook)
        self.assertNotIn(self.FLAT_CIT_LABEL, labels)


class ConstructionGraceAuditTests(TestCase):

    NAMES = ("CONSTRUCTION_MONTHS", "IDC", "PRINCIPAL_GRACE_YEARS",
             "COD_DEBT_BALANCE")

    def _labels(self, workbook):
        sheet = workbook["Pro Forma (Audit)"]
        return {
            sheet.cell(row=row, column=1).value: row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value
        }

    def test_disabled_construction_leaves_no_named_cells_or_formulas(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)

        for name in self.NAMES:
            self.assertNotIn(name, workbook.defined_names, name)
        # Debt and depreciation formulas keep the overnight-build shape.
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]
        self.assertEqual(
            sheet.cell(row=labels["Opening debt balance"], column=4).value,
            "=DEBT_PRINCIPAL",
        )
        self.assertNotIn(
            "PRINCIPAL_GRACE_YEARS",
            sheet.cell(row=labels["Principal repayment"], column=4).value,
        )
        self.assertNotIn(
            "IDC",
            sheet.cell(row=labels["PV depreciation (straight-line)"], column=4).value,
        )

    def test_enabled_construction_defines_named_cells(self):
        workbook = build_vietnam_esco_workbook(
            _construction_result(), assumptions=ESCO_ASSUMPTIONS
        )

        for name in self.NAMES:
            self.assertIn(name, workbook.defined_names, name)

    def test_debt_schedule_reproduces_grace_pattern(self):
        workbook = build_vietnam_esco_workbook(
            _construction_result(), assumptions=ESCO_ASSUMPTIONS
        )
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        # The year-1 opening balance is the rolled-up COD balance, and the
        # principal row stays 0 through the grace window.
        self.assertEqual(
            sheet.cell(row=labels["Opening debt balance"], column=4).value,
            "=COD_DEBT_BALANCE",
        )
        principal_formula = sheet.cell(
            row=labels["Principal repayment"], column=4
        ).value
        self.assertIn("PRINCIPAL_GRACE_YEARS", principal_formula)

    def test_depreciable_base_tie_out_includes_idc(self):
        workbook = build_vietnam_esco_workbook(
            _construction_result(), assumptions=ESCO_ASSUMPTIONS
        )
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        for label in ("PV depreciation (straight-line)",
                      "BESS depreciation (straight-line)"):
            formula = sheet.cell(row=labels[label], column=4).value
            self.assertIn("IDC", formula, label)


class CitRegimeAuditTests(TestCase):

    def test_re_producer_case_defines_preferential_named_cells(self):
        result, dppa_inputs = _dppa_result()
        workbook = build_vietnam_esco_workbook(
            result, assumptions={**ESCO_ASSUMPTIONS, "dppa": dppa_inputs}
        )

        self.assertEqual(result["derivation"]["cit"]["regime"], "re_producer")
        for name in ("CIT_PREF_RATE", "CIT_PREF_YEARS"):
            self.assertIn(name, workbook.defined_names, name)

    def test_re_producer_rate_formula_switches_at_preferential_window(self):
        result, dppa_inputs = _dppa_result()
        workbook = build_vietnam_esco_workbook(
            result, assumptions={**ESCO_ASSUMPTIONS, "dppa": dppa_inputs}
        )
        sheet = workbook["Pro Forma (Audit)"]

        year1_formulas = [
            sheet.cell(row=row, column=4).value
            for row in range(1, sheet.max_row + 1)
        ]
        self.assertTrue(
            any(
                isinstance(formula, str)
                and "CIT_PREF_YEARS" in formula
                and "CIT_PREF_RATE" in formula
                for formula in year1_formulas
            ),
            "no CIT rate formula switches at the preferential window boundary",
        )

    def test_standard_regime_rate_formula_is_unchanged(self):
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)

        self.assertNotIn("CIT_PREF_RATE", workbook.defined_names)
        self.assertEqual(result["derivation"]["cit"]["regime"], "standard_with_holiday")

        sheet = workbook["Pro Forma (Audit)"]
        rate_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value == "Applicable CIT rate"
        )
        formula = sheet.cell(row=rate_row, column=4).value
        self.assertNotIn("CIT_PREF", formula)
        # legacy formula preserved bit-for-bit for the conservative ESCO default
        self.assertIn("CIT_RATE*CIT_REDUCED_FRACTION,CIT_RATE))", formula)


class FxSensitivitySheetTests(TestCase):

    def test_table_carries_engine_scenarios_and_live_irr_formulas(self):
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["FX Sensitivity"]
        engine = calculate_fx_sensitivity(result)

        for index, scenario in enumerate(engine):
            row = 5 + index
            self.assertAlmostEqual(
                sheet.cell(row=row, column=1).value,
                scenario["vnd_depreciation_rate"],
            )
            self.assertTrue(str(sheet.cell(row=row, column=2).value).startswith("=IRR("))
            self.assertAlmostEqual(
                sheet.cell(row=row, column=4).value,
                scenario["equity_irr_fraction"],
            )
            self.assertAlmostEqual(
                sheet.cell(row=row, column=5).value, scenario["npv_usd"]
            )


class UsdDebtAuditTests(TestCase):

    def _header_labels(self, sheet, header_row):
        return [
            sheet.cell(row=header_row, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

    def test_vnd_default_defines_no_debt_currency_cell_or_dscr_column(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)

        self.assertNotIn("DEBT_CURRENCY", workbook.defined_names)
        fx = workbook["FX Sensitivity"]
        headers = self._header_labels(fx, 4)
        self.assertNotIn("Min DSCR (USD)", headers)
        self.assertNotIn("Min DSCR (engine)", headers)
        # Legacy note preserved bit-for-bit (DSCR unchanged claim intact).
        self.assertIn("DSCR is unchanged", fx.cell(row=2, column=1).value)

    def test_usd_debt_defines_currency_cell(self):
        workbook = build_vietnam_esco_workbook(_usd_debt_result(), assumptions=ESCO_ASSUMPTIONS)

        self.assertIn("DEBT_CURRENCY", workbook.defined_names)

    def test_usd_debt_surfaces_resolved_rate_in_debt_rate_cell(self):
        result = _usd_debt_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Assumptions"]
        rate_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=2).value == "Debt interest rate"
        )
        # The existing DEBT_RATE cell already carries the engine-resolved 5% USD
        # rate — no separate resolved-rate cell is needed.
        self.assertAlmostEqual(sheet.cell(row=rate_row, column=3).value, 0.05)

    def test_usd_debt_fx_sheet_has_live_min_dscr_column(self):
        result = _usd_debt_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        fx = workbook["FX Sensitivity"]
        engine = calculate_fx_sensitivity(result)

        headers = self._header_labels(fx, 4)
        self.assertIn("Min DSCR (USD)", headers)
        self.assertIn("Min DSCR (engine)", headers)
        min_dscr_col = headers.index("Min DSCR (USD)") + 1
        engine_col = headers.index("Min DSCR (engine)") + 1
        for index, scenario in enumerate(engine):
            row = 5 + index
            self.assertTrue(
                str(fx.cell(row=row, column=min_dscr_col).value).startswith("=MIN(")
            )
            self.assertAlmostEqual(
                fx.cell(row=row, column=engine_col).value, scenario["min_dscr"]
            )

    def test_usd_debt_model_basis_discloses_conventions(self):
        workbook = build_vietnam_esco_workbook(_usd_debt_result(), assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Model Basis"]
        text = "\n".join(
            str(sheet.cell(row=row, column=col).value)
            for row in range(1, sheet.max_row + 1)
            for col in range(1, 4)
            if sheet.cell(row=row, column=col).value
        )

        self.assertIn("USD", text)
        # The three disclosed conventions: fixed USD debt service under FX drift,
        # CIT not recomputed, and no VAS FX revaluation on outstanding principal.
        self.assertIn("revaluation", text.lower())

    def test_vnd_default_model_basis_omits_usd_debt_disclosures(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Model Basis"]
        text = "\n".join(
            str(sheet.cell(row=row, column=col).value)
            for row in range(1, sheet.max_row + 1)
            for col in range(1, 4)
            if sheet.cell(row=row, column=col).value
        )

        self.assertNotIn("revaluation", text.lower())


class DscrSizingAuditTests(TestCase):

    @staticmethod
    def _named_cell(workbook, name):
        destination = next(iter(workbook.defined_names[name].destinations))
        sheet_name, coord = destination
        return workbook[sheet_name][coord]

    @staticmethod
    def _proforma_labels(workbook):
        sheet = workbook["Pro Forma (Audit)"]
        return [
            sheet.cell(row=row, column=1).value
            for row in range(1, sheet.max_row + 1)
        ]

    def test_fraction_default_defines_no_sizing_cells_and_keeps_debt_formula(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)

        for name in ("TARGET_MIN_DSCR", "SUPPORTED_DEBT", "FRACTION_DEBT"):
            self.assertNotIn(name, workbook.defined_names)
        self.assertEqual(
            self._named_cell(workbook, "DEBT_PRINCIPAL").value,
            "=TOTAL_CAPEX*DEBT_FRACTION",
        )
        self.assertNotIn(
            "DSCR-sized debt fixed point", self._proforma_labels(workbook)
        )

    def test_dscr_sized_defines_sizing_cells_and_min_debt_formula(self):
        result = _dscr_sized_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)

        for name in ("TARGET_MIN_DSCR", "SUPPORTED_DEBT", "FRACTION_DEBT"):
            self.assertIn(name, workbook.defined_names)
        # DEBT_PRINCIPAL reproduces the min() sizing rule live.
        self.assertEqual(
            self._named_cell(workbook, "DEBT_PRINCIPAL").value,
            "=MIN(FRACTION_DEBT,SUPPORTED_DEBT)",
        )
        # The supported-debt cell carries the engine's converged fixed point.
        sizing = result["derivation"]["debt_sizing"]
        self.assertAlmostEqual(
            self._named_cell(workbook, "SUPPORTED_DEBT").value,
            sizing["supported_principal_usd"],
        )
        self.assertAlmostEqual(
            self._named_cell(workbook, "TARGET_MIN_DSCR").value,
            sizing["target_min_dscr"],
        )

    def test_dscr_sized_adds_fixed_point_tie_out_row(self):
        workbook = build_vietnam_esco_workbook(_dscr_sized_result(), assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Pro Forma (Audit)"]
        labels = self._proforma_labels(workbook)
        self.assertIn("DSCR-sized debt fixed point", labels)
        row = labels.index("DSCR-sized debt fixed point") + 1
        status_formula = sheet.cell(row=row, column=6).value
        # The status is a live formula branching on whether the covenant binds.
        self.assertIn("SUPPORTED_DEBT<FRACTION_DEBT", status_formula)
        self.assertIn("TARGET_MIN_DSCR", status_formula)

    def test_dscr_sized_model_basis_discloses_conventions(self):
        workbook = build_vietnam_esco_workbook(_dscr_sized_result(), assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Model Basis"]
        text = "\n".join(
            str(sheet.cell(row=row, column=col).value)
            for row in range(1, sheet.max_row + 1)
            for col in range(1, 4)
            if sheet.cell(row=row, column=col).value
        )
        self.assertIn("min(fraction-based, DSCR-supported)", text)
        self.assertIn("level-payment sizing", text)

    def test_fraction_default_model_basis_omits_sizing_disclosures(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)
        sheet = workbook["Model Basis"]
        text = "\n".join(
            str(sheet.cell(row=row, column=col).value)
            for row in range(1, sheet.max_row + 1)
            for col in range(1, 4)
            if sheet.cell(row=row, column=col).value
        )
        self.assertNotIn("DSCR-supported", text)
        self.assertNotIn("level-payment sizing", text)


class BatteryReplacementCapitalizationAuditTests(TestCase):
    """Circular 45 replacement capitalization on the audit workbook. The default
    (capitalize) fixture ``_esco_result`` carries a year-11 replacement, so its
    workbook gains a per-replacement depreciation row, the replacement-aware total
    depreciation and EBT formulas, and the capitalized Model Basis disclosure. The
    legacy ``"expense"`` fixture keeps the two-class depreciation / EBT formulas
    and the expensing register line byte-for-byte.
    """

    REPL_DEP_LABEL = "Replacement depreciation (in-service yr 11)"

    def _labels(self, workbook):
        sheet = workbook["Pro Forma (Audit)"]
        return {
            sheet.cell(row=row, column=1).value: row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value
        }

    def _model_basis_text(self, workbook):
        sheet = workbook["Model Basis"]
        return "\n".join(
            str(sheet.cell(row=row, column=col).value)
            for row in range(1, sheet.max_row + 1)
            for col in range(1, 4)
            if sheet.cell(row=row, column=col).value
        )

    def test_capitalize_adds_replacement_row_and_replacement_aware_formulas(self):
        result = _esco_result()  # capitalize default, year-11 replacement
        self.assertIn("battery_replacement", result["derivation"])
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        # A live replacement-depreciation row exists for the year-11 asset.
        self.assertIn(self.REPL_DEP_LABEL, labels)
        repl_dep_formula = sheet.cell(
            row=labels[self.REPL_DEP_LABEL], column=4
        ).value
        # Active only within the 8-year in-service window, cost / BESS_DEP_YEARS.
        self.assertIn("BESS_DEP_YEARS", repl_dep_formula)
        self.assertIn(">=11", repl_dep_formula)

        # Total depreciation ties in the replacement schedule (PV + BESS + repl =
        # two '+' terms) and EBT adds the replacement cost back (the only mode
        # whose EBT formula carries a '+').
        total_dep = sheet.cell(row=labels["Total depreciation"], column=4).value
        self.assertEqual(total_dep.count("+"), 2)
        ebt = sheet.cell(
            row=labels["Taxable income before loss relief (EBT)"], column=4
        ).value
        repl_col_ref = f"D{labels['Battery replacement (engine schedule)']}"
        self.assertIn("+" + repl_col_ref, ebt)

    def test_expense_mode_keeps_legacy_depreciation_ebt_and_register(self):
        result = _esco_expense_result()  # legacy expense treatment
        self.assertNotIn("battery_replacement", result["derivation"])
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        # No replacement-depreciation rows; the two-class total depreciation and
        # the legacy EBT (no add-back) are preserved bit-for-bit.
        self.assertNotIn(self.REPL_DEP_LABEL, labels)
        self.assertFalse(
            any(str(label).startswith("Replacement depreciation (in-service yr")
                for label in labels)
        )
        total_dep = sheet.cell(row=labels["Total depreciation"], column=4).value
        self.assertEqual(total_dep.count("+"), 1)
        ebt = sheet.cell(
            row=labels["Taxable income before loss relief (EBT)"], column=4
        ).value
        self.assertNotIn("+", ebt)

        # The expensing simplification register line is kept; the capitalized
        # Model Basis disclosure is absent.
        text = self._model_basis_text(workbook)
        self.assertIn("expensed in the replacement year", text)
        self.assertNotIn("CAPITALIZED, not expensed", text)

    def test_capitalize_model_basis_discloses_capitalization_and_truncation(self):
        workbook = build_vietnam_esco_workbook(_esco_result(), assumptions=ESCO_ASSUMPTIONS)
        text = self._model_basis_text(workbook)

        self.assertIn("CAPITALIZED, not expensed", text)
        self.assertIn("truncated at the analysis horizon", text)
        # The expensing simplification line is removed in the default treatment.
        self.assertNotIn("expensed in the replacement year", text)

    def test_no_replacement_case_keeps_legacy_formulas_and_register(self):
        # Capitalize default but an all-zero replacement schedule: no block is
        # emitted, so the workbook is byte-for-byte the legacy expensed shape.
        result = _esco_result(replacement_costs_by_year=[0.0] * 30)
        self.assertNotIn("battery_replacement", result["derivation"])
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        labels = self._labels(workbook)
        sheet = workbook["Pro Forma (Audit)"]

        self.assertNotIn(self.REPL_DEP_LABEL, labels)
        total_dep = sheet.cell(row=labels["Total depreciation"], column=4).value
        self.assertEqual(total_dep.count("+"), 1)
        self.assertIn("expensed in the replacement year", self._model_basis_text(workbook))


class CoverSheetTests(TestCase):

    def test_cover_aggregates_check_statuses_when_audit_sheets_exist(self):
        result = _esco_result()
        workbook = build_vietnam_esco_workbook(result, assumptions=ESCO_ASSUMPTIONS)
        cover = workbook["Cover"]

        status_formula = None
        for row in range(1, cover.max_row + 1):
            value = cover.cell(row=row, column=3).value
            if isinstance(value, str) and value.startswith("=IF(COUNTIF"):
                status_formula = value
        self.assertIsNotNone(status_formula)
        self.assertIn("Pro Forma (Audit)", status_formula)
        self.assertIn("FX Sensitivity", status_formula)

    def test_cover_omits_status_row_without_derivation(self):
        workbook = build_vietnam_esco_workbook(
            {"summary": {}, "annual_cash_flows": []},
            assumptions={"esco_energy_discount_fraction": 0.9},
        )
        cover = workbook["Cover"]
        for row in range(1, cover.max_row + 1):
            value = cover.cell(row=row, column=3).value
            if isinstance(value, str):
                self.assertFalse(value.startswith("=IF(COUNTIF"))
